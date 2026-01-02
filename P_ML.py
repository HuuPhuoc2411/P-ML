

# ============================================================================
# ARDUINO OPTIMIZATION SETTINGS
# ============================================================================
# Cau hinh toi uu hoa cho Arduino Mega (256KB Flash, 8KB RAM) / Uno (32KB Flash, 2KB RAM)

# TARGET BOARD: "uno", "mega", "esp32" (tu dong cau hinh optimize)
TARGET_BOARD = "mega"             # Chon board muc tieu de auto-config

# === AUTO-CONFIG theo board ===
# Uno: 32KB Flash, 2KB RAM - Rat han che, chi models don gian nhat
# Mega: 256KB Flash, 8KB RAM - Co the chay SVM nho/medium
# ESP32: 4MB Flash, 520KB RAM - Khong can optimize

ARDUINO_OPTIMIZE = True           # True: Bat PROGMEM + float optimization cho AVR
USE_FLOAT_PRECISION = True        # True = float (4 bytes), False = double (8 bytes)
COEFFICIENT_PRECISION = 6         # So chu so thap phan (6 cho Mega, 10 cho ESP32)

# === GIOI HAN KICH THUOC MODEL ===
MAX_SUPPORT_VECTORS = 150         # Gioi han so SVs (Uno: 30-50, Mega: 100-200, ESP32: 999999)
MAX_MLP_NEURONS = 100             # Gioi han tong so neurons trong MLP hidden layers
MAX_KNN_SAMPLES = 50              # Gioi han so training samples cho KNN (Uno: 20, Mega: 50-100)
MAX_TREE_DEPTH = 10               # Gioi han do sau cay quyet dinh

# === SIZE ESTIMATION LIMITS (bytes) ===
MAX_FLASH_BYTES = 200000          # Mega: ~200KB kha dung, Uno: ~30KB
MAX_RAM_BYTES = 6000              # Mega: ~6KB kha dung, Uno: ~1.5KB

# MEGA TRAINING (FIND TRUE SVM THAT FITS FLASH)
# ============================================================================
# Neu ban muon "SVM kernel that" chay duoc tren Arduino Mega, cach dung la:
#   train sao cho so support vectors du nho de fit 256KB Flash.
# Cac knob duoi day se them penalty vao Optuna de uu tien model nho hon.
TRAIN_FOR_MEGA = True             # True: Optuna se uu tien SVM it SV de fit Mega
MEGA_FLASH_BUDGET_KB = 180        # Ngan sach Flash danh cho 1 model SVM (KB) - Mega: 180KB, Uno: 25KB
SVM_FLASH_PENALTY_WEIGHT = 0.9    # Cang lon => cang uu tien model nho (0.8-0.95 khuyen nghi)

# === AUTO-CONFIGURE theo TARGET_BOARD ===
def auto_configure_for_board():
    """Tu dong cau hinh settings theo board muc tieu"""
    global ARDUINO_OPTIMIZE, USE_FLOAT_PRECISION, COEFFICIENT_PRECISION
    global MAX_SUPPORT_VECTORS, MAX_MLP_NEURONS, MAX_KNN_SAMPLES, MAX_TREE_DEPTH
    global MAX_FLASH_BYTES, MAX_RAM_BYTES, MEGA_FLASH_BUDGET_KB, TRAIN_FOR_MEGA
    
    board = TARGET_BOARD.lower()
    
    if board == "uno":
        # Arduino Uno: 32KB Flash, 2KB RAM - Rat han che
        ARDUINO_OPTIMIZE = True
        USE_FLOAT_PRECISION = True
        COEFFICIENT_PRECISION = 4           # Giam precision de tiet kiem Flash
        MAX_SUPPORT_VECTORS = 30            # Rat it SVs
        MAX_MLP_NEURONS = 30                # MLP rat nho
        MAX_KNN_SAMPLES = 20                # KNN sieu nho
        MAX_TREE_DEPTH = 6                  # Cay nong
        MAX_FLASH_BYTES = 28000             # ~28KB kha dung
        MAX_RAM_BYTES = 1500                # ~1.5KB kha dung
        MEGA_FLASH_BUDGET_KB = 25           # Ngan sach cho 1 model
        TRAIN_FOR_MEGA = True
        print("🎯 Auto-config cho Arduino UNO (32KB Flash, 2KB RAM)")
        print("     Only very small models will fit! Recommended: LinearSVC, LDA, GaussianNB")
        
    elif board == "mega":
        # Arduino Mega: 256KB Flash, 8KB RAM
        ARDUINO_OPTIMIZE = True
        USE_FLOAT_PRECISION = True
        COEFFICIENT_PRECISION = 6           # Can bang precision va size
        MAX_SUPPORT_VECTORS = 150           # Medium SVs
        MAX_MLP_NEURONS = 100               # MLP vua
        MAX_KNN_SAMPLES = 50                # KNN vua
        MAX_TREE_DEPTH = 10                 # Cay medium
        MAX_FLASH_BYTES = 200000            # ~200KB kha dung
        MAX_RAM_BYTES = 6000                # ~6KB kha dung
        MEGA_FLASH_BUDGET_KB = 180          # Ngan sach cho 1 model
        TRAIN_FOR_MEGA = True
        print("🎯 Auto-config cho Arduino MEGA (256KB Flash, 8KB RAM)")
        print("   ✅ Can run small/medium SVM, MLP, Decision Tree")
        
    elif board == "esp32":
        # ESP32: 4MB Flash, 520KB RAM - Khong can optimize
        ARDUINO_OPTIMIZE = False
        USE_FLOAT_PRECISION = True          # Van dung float cho toc do
        COEFFICIENT_PRECISION = 10          # Full precision
        MAX_SUPPORT_VECTORS = 999999        # Khong gioi han
        MAX_MLP_NEURONS = 999999
        MAX_KNN_SAMPLES = 999999
        MAX_TREE_DEPTH = 999999
        MAX_FLASH_BYTES = 3500000           # ~3.5MB kha dung
        MAX_RAM_BYTES = 400000              # ~400KB kha dung
        MEGA_FLASH_BUDGET_KB = 3000         # 3MB Ngan sach
        TRAIN_FOR_MEGA = False
        print("🎯 Auto-config cho ESP32 (4MB Flash, 520KB RAM)")
        print("   ✅ No optimization needed - enough memory for all models")
    else:
        print(f"  Board '{board}' not recognized. Using Mega config as default.")

# Goi auto-config khi import module
auto_configure_for_board()

# Luu y: 
# - Tat ARDUINO_OPTIMIZE neu dung ESP32 (4MB Flash, 520KB RAM)
# - Double precision chi can thiet cho cac bai toan cuc ky nhay cam

# ============================================================================
# INSTALL REQUIRED PACKAGES (For Google Colab)
# ============================================================================
import sys

# Kiểm tra xem có đang chạy trên Colab không
try:
    import google.colab
    IN_COLAB_CHECK = True
except:
    IN_COLAB_CHECK = False

# Kiểm tra xem có đang chạy trên Kaggle không
try:
    import kaggle
    IN_KAGGLE_CHECK = True
except:
    IN_KAGGLE_CHECK = False

# Cài đặt các thư viện cần thiết
if IN_COLAB_CHECK or IN_KAGGLE_CHECK:
    print(" Installing: optuna, openpyxl, gradio...")
    def _maybe_install_deps():
        """Notebook-friendly dependency install.

        - In Colab/Jupyter: runs `pip install -q ...` via IPython.
        - In normal Python execution: does nothing (assumes deps are preinstalled).
        """
        try:
            ip = get_ipython()  # type: ignore[name-defined]
        except Exception:
            ip = None
        if ip is None:
            return
        try:
            ip.run_line_magic("pip", "install -q optuna openpyxl gradio")
        except Exception:
            # If pip magic isn't available, just skip.
            return


    _maybe_install_deps()
    print("✅ Installation complete!")

# ============================================================================
# IMPORT LIBRARIES
# ============================================================================
print("📦 Importing libraries (may take 30-60 seconds first time)...")

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
print("   📚 pandas, numpy, matplotlib, seaborn")

from sklearn.model_selection import StratifiedKFold, cross_validate
from sklearn.svm import SVC, NuSVC, LinearSVC
from sklearn.tree import DecisionTreeClassifier
from sklearn.ensemble import (RandomForestClassifier, ExtraTreesClassifier, 
                               GradientBoostingClassifier, AdaBoostClassifier)
from sklearn.neural_network import MLPClassifier
from sklearn.neighbors import KNeighborsClassifier
from sklearn.discriminant_analysis import LinearDiscriminantAnalysis, QuadraticDiscriminantAnalysis
from sklearn.naive_bayes import GaussianNB, BernoulliNB
from sklearn.preprocessing import StandardScaler
from sklearn.kernel_approximation import RBFSampler
from sklearn.pipeline import Pipeline
from sklearn.metrics import (accuracy_score, precision_score, recall_score, 
                             f1_score, confusion_matrix, classification_report)
print("   ✅ scikit-learn")

import optuna
from optuna.samplers import TPESampler
from optuna.pruners import MedianPruner
print("   ✅ optuna")

import json
import os
import time
import pickle
import shutil
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')
warnings.filterwarnings('ignore', category=optuna.exceptions.ExperimentalWarning)

# For Google Colab
try:
    from google.colab import files, drive
    IN_COLAB = True
except:
    IN_COLAB = False

# For Kaggle
try:
    import kaggle
    IN_KAGGLE = True
except:
    IN_KAGGLE = False

# Check if running in cloud environment
IN_CLOUD = IN_COLAB or IN_KAGGLE

# Import Gradio cho giao diện web
print("   ⏳ Importing gradio (may take 20-40 seconds)...")
try:
    import gradio as gr
    GRADIO_AVAILABLE = True
    print("   ✅ gradio")
except ImportError:
    gr = None
    GRADIO_AVAILABLE = False
    if not IN_CLOUD:
        print("⚠️  Gradio not installed. Run: pip install gradio")

print("✅ Import libraries complete!\n")

# ============================================================================
# CONFIGURATION - NGƯỜI DÙNG TỰ CẤU HÌNH Ở ĐÂY
# ============================================================================

# --- Đường dẫn file (chỉ dùng khi chạy CLI mode, Gradio mode tải qua giao diện) ---
SINGLE_FILE_PATH = ""  # Để trống nếu dùng Gradio
TRAIN_FILE_PATH = ""   # Để trống nếu dùng Gradio  
TEST_FILE_PATH = ""    # Để trống nếu dùng Gradio

# --- Cấu hình chia dữ liệu ---
TRAIN_RATIO = 0.7  # Tỷ lệ train (70%), test sẽ là 30%
RANDOM_STATE = 42

# --- Cấu hình Optuna ---
N_TRIALS = 50  # Số lần thử tối ưu hóa (tăng lên để kết quả tốt hơn nhưng mất thời gian)
CV_FOLDS = 5   # Số fold cho cross-validation
OPTUNA_TIMEOUT = 180  # Timeout cho mỗi model (giây) - giảm xuống để train nhanh hơn
MIN_TRIALS_BEFORE_PRUNING = 3  # S? trial t?i thi?u tru?c khi b?t d?u prune
PRUNING_WARMUP_STEPS = 1  # S? fold t?i thi?u tru?c khi prune trial

# --- C?u h?nh output ---
OUTPUT_DIR = "ML_Output"  # Thu m?c ch?a k?t qu?
ARDUINO_LIB_NAME = "MLPredictor"  # T?n thu vi?n Arduino
FUNC_NAME_SUFFIX = ""  # H?u t? th?m v?o t?n h?m (vd: "_v1"  predict_SVM_v1)
SEPARATE_MODEL_FILES = True  # True: M?i model 1 file ri?ng (t?i uu ROM), False: T?t c? trong 1 file

def get_safe_name(model_name):
    """Convert model name to safe C identifier"""
    return model_name.replace('-', '_').replace(' ', '_')

def get_func_name(model_name):
    """Get function name with suffix: safe_model_name + FUNC_NAME_SUFFIX"""
    safe_name = get_safe_name(model_name)
    return safe_name + FUNC_NAME_SUFFIX if FUNC_NAME_SUFFIX else safe_name

# ============================================================================
# SPXY ALGORITHM - KENNARD-STONE ALGORITHM
# ============================================================================

def calculate_distance_matrix(X):
    """T?nh ma tr?n kho?ng c?ch Euclidean"""
    n = X.shape[0]
    dist_matrix = np.zeros((n, n))
    for i in range(n):
        for j in range(i+1, n):
            dist = np.linalg.norm(X[i] - X[j])
            dist_matrix[i, j] = dist
            dist_matrix[j, i] = dist
    return dist_matrix

def spxy_split(X, y, train_size=0.7, random_state=None):
    """
    SPXY algorithm d? chia d? li?u gi? t? l? c?c l?p
    
    Parameters:
    -----------
    X : array-like, shape (n_samples, n_features)
    y : array-like, shape (n_samples,)
    train_size : float, t? l? d? li?u train
    random_state : int, seed cho random
    
    Returns:
    --------
    X_train, X_test, y_train, y_test
    """
    if random_state:
        np.random.seed(random_state)
    
    # L?y c?c l?p duy nh?t
    classes = np.unique(y)
    
    train_indices = []
    test_indices = []
    
    # Áp dụng SPXY cho từng lớp
    for cls in classes:
        cls_indices = np.where(y == cls)[0]
        X_cls = X[cls_indices]
        
        n_train = int(len(cls_indices) * train_size)
        
        if n_train < 2:
            # N?u l?p qu? nh?, chia ng?u nhi?n
            np.random.shuffle(cls_indices)
            train_indices.extend(cls_indices[:n_train])
            test_indices.extend(cls_indices[n_train:])
            continue
        
        # T?nh ma tr?n kho?ng c?ch
        dist_matrix = calculate_distance_matrix(X_cls)
        
        # Ch?n 2 m?u xa nh?t
        max_dist_idx = np.unravel_index(dist_matrix.argmax(), dist_matrix.shape)
        selected = list(max_dist_idx)
        remaining = list(set(range(len(X_cls))) - set(selected))
        
        # Ch?n c?c m?u ti?p theo
        while len(selected) < n_train and remaining:
            max_min_dist = -1
            max_min_idx = -1
            
            for idx in remaining:
                min_dist = min([dist_matrix[idx, s] for s in selected])
                if min_dist > max_min_dist:
                    max_min_dist = min_dist
                    max_min_idx = idx
            
            if max_min_idx != -1:
                selected.append(max_min_idx)
                remaining.remove(max_min_idx)
            else:
                break
        
        # N?u chua d?, th?m ng?u nhi?n
        if len(selected) < n_train:
            add_more = np.random.choice(remaining, n_train - len(selected), replace=False)
            selected.extend(add_more)
            remaining = list(set(remaining) - set(add_more))
        
        # Th?m v?o danh s?ch train v? test
        train_indices.extend(cls_indices[selected])
        test_indices.extend(cls_indices[remaining])
    
    # Tr?n d? li?u
    np.random.shuffle(train_indices)
    np.random.shuffle(test_indices)
    
    return X[train_indices], X[test_indices], y[train_indices], y[test_indices]

# ============================================================================
# DATA LOADING AND PREPROCESSING
# ============================================================================

def load_data():
    """Đọc dữ liệu từ Excel"""
    print("=" * 80)
    print("START READING DATA")
    print("=" * 80)
    
    # Ki?m tra tru?ng h?p n?o
    if SINGLE_FILE_PATH:
        print(f" Case 1: Reading from single file: {SINGLE_FILE_PATH}")
        df = pd.read_excel(SINGLE_FILE_PATH)
        
        # Ki?m tra c?t Class
        if 'Class' not in df.columns:
            raise ValueError("File ph?i c? c?t 'Class'!")
        
        print(f"✅ Successfully read {len(df)} samples")
        print(f"ℹ️ Columns: {list(df.columns)}")
        print(f"ℹ️ Class distribution:\n{df['Class'].value_counts().sort_index()}")
        
        # T?ch Class
        y = df['Class'].values
        
        # Get features and convert to numeric
        print("\n Converting data to numeric...")
        feature_df = df.drop('Class', axis=1).apply(pd.to_numeric, errors='coerce')
        feature_names = feature_df.columns.tolist()
        
        # Ki?m tra NaN
        nan_count = feature_df.isna().sum().sum()
        if nan_count > 0:
            print(f"  Detected {nan_count} NaN values - replacing with median")
            for col in feature_df.columns:
                median_val = feature_df[col].median()
                if pd.isna(median_val):
                    median_val = 0
                feature_df[col].fillna(median_val, inplace=True)
        
        X = feature_df.values.astype(np.float64)
        print(f"✅ Conversion successful: {X.shape[1]} features")
        
        print(f"\n📊 Splitting data with SPXY (Train: {TRAIN_RATIO*100}%, Test: {(1-TRAIN_RATIO)*100}%)")
        X_train, X_test, y_train, y_test = spxy_split(X, y, train_size=TRAIN_RATIO, random_state=RANDOM_STATE)
        
        print(f"📊 Train: {len(X_train)} samples, Test: {len(X_test)} samples")
        
        # Luu file train v? test
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        
        train_df = pd.DataFrame(X_train, columns=feature_names)
        train_df['Class'] = y_train
        train_df.to_excel(f"{OUTPUT_DIR}/train_data.xlsx", index=False)
        print(f"✓ Saved: {OUTPUT_DIR}/train_data.xlsx")
        
        test_df = pd.DataFrame(X_test, columns=feature_names)
        test_df['Class'] = y_test
        test_df.to_excel(f"{OUTPUT_DIR}/test_data.xlsx", index=False)
        print(f"✓ Saved: {OUTPUT_DIR}/test_data.xlsx")
        
    elif TRAIN_FILE_PATH and TEST_FILE_PATH:
        print(f"ℹ️ Case 2: Reading from 2 pre-split files")
        print(f"   Train: {TRAIN_FILE_PATH}")
        print(f"   Test: {TEST_FILE_PATH}")
        
        train_df = pd.read_excel(TRAIN_FILE_PATH)
        test_df = pd.read_excel(TEST_FILE_PATH)
        
        # Ki?m tra c?t Class
        if 'Class' not in train_df.columns or 'Class' not in test_df.columns:
            raise ValueError("C? 2 file ph?i c? c?t 'Class'!")
        
        print(f"📊 Train: {len(train_df)} samples")
        print(f"📊 Test: {len(test_df)} samples")
        
        # T?ch features v? class
        y_train = train_df['Class'].values
        y_test = test_df['Class'].values
        
        # L?y feature columns
        feature_cols = train_df.drop('Class', axis=1)
        feature_names = feature_cols.columns.tolist()
        
        # Convert all features to numeric, replace non-numeric with NaN
        print("\n Converting data to numeric...")
        X_train_df = train_df.drop('Class', axis=1).apply(pd.to_numeric, errors='coerce')
        X_test_df = test_df.drop('Class', axis=1).apply(pd.to_numeric, errors='coerce')
        
        # Ki?m tra NaN values
        train_nans = X_train_df.isna().sum().sum()
        test_nans = X_test_df.isna().sum().sum()
        
        if train_nans > 0 or test_nans > 0:
            print(f"  Detected {train_nans} NaN in train and {test_nans} NaN in test")
            print("⚠️  NaN values will be replaced with feature median")
            
            # Thay th? NaN b?ng median
            for col in X_train_df.columns:
                median_val = X_train_df[col].median()
                if pd.isna(median_val):  # N?u c? c?t d?u l? NaN
                    median_val = 0
                X_train_df[col].fillna(median_val, inplace=True)
                X_test_df[col].fillna(median_val, inplace=True)
        
        # Convert to numpy arrays
        X_train = X_train_df.values.astype(np.float64)
        X_test = X_test_df.values.astype(np.float64)
        
        print(f"✅ Conversion successful: {X_train.shape[1]} features")
        
    else:
        raise ValueError("Vui l?ng cung c?p du?ng d?n file!")
    
    return X_train, X_test, y_train, y_test, feature_names

def plot_data_distribution(X_train, X_test, y_train, y_test, feature_names):
    """Vẽ biểu đồ Bar chart Train vs Test (Mean ± Std) cho tất cả features"""
    print("\n" + "=" * 80)
    print("CREATE DATA DISTRIBUTION CHART")
    print("=" * 80)
    
    # Create output folder n?u chua c?
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    
    n_features = len(feature_names)
    
    # T?nh mean v? std cho train v? test
    train_means = []
    train_stds = []
    test_means = []
    test_stds = []
    
    for i in range(n_features):
        train_means.append(X_train[:, i].mean())
        train_stds.append(X_train[:, i].std())
        test_means.append(X_test[:, i].mean())
        test_stds.append(X_test[:, i].std())
    
    # T?o figure
    fig, ax = plt.subplots(figsize=(20, 6))
    
    # V? tr? c?c bars
    x = np.arange(n_features)
    width = 0.35
    
    # Vẽ bars với error bars
    bars1 = ax.bar(x - width/2, train_means, width, yerr=train_stds,
                   label='Train (Mean ± Std)', color='cornflowerblue', 
                   capsize=3, alpha=0.8, edgecolor='black', linewidth=0.5)
    bars2 = ax.bar(x + width/2, test_means, width, yerr=test_stds,
                   label='Test (Mean ± Std)', color='coral',
                   capsize=3, alpha=0.8, edgecolor='black', linewidth=0.5)
    
    # Thiết lập labels và title
    ax.set_xlabel('Features', fontsize=12, fontweight='bold')
    ax.set_ylabel('Values', fontsize=12, fontweight='bold')
    ax.set_title('Data Distribution: Train vs Test (Mean ± Std)', fontsize=14, fontweight='bold')
    ax.set_xticks(x)
    ax.set_xticklabels(feature_names, rotation=45, ha='right', fontsize=9)
    ax.legend(loc='upper right', fontsize=11)
    ax.grid(axis='y', alpha=0.3, linestyle='--')
    
    # Tight layout
    plt.tight_layout()
    plt.savefig(f"{OUTPUT_DIR}/data_distribution.png", dpi=300, bbox_inches='tight')
    print(f"✓ Saved chart: {OUTPUT_DIR}/data_distribution.png")
    plt.close()

# ============================================================================
# MODEL TRAINING WITH OPTUNA
# ============================================================================

def get_model_configs():
    """Định nghĩa t?t c? các mô hình ML"""
    return {
        # Support Vector Machines
        "SVM_linear": {"family": "SVM", "model_class": SVC, "kernel": "linear"},
        "SVM_rbf": {"family": "SVM", "model_class": SVC, "kernel": "rbf"},
        "SVM_poly": {"family": "SVM", "model_class": SVC, "kernel": "poly"},
        "SVM_sigmoid": {"family": "SVM", "model_class": SVC, "kernel": "sigmoid"},
        "NuSVC_linear": {"family": "SVM", "model_class": NuSVC, "kernel": "linear"},
        "NuSVC_rbf": {"family": "SVM", "model_class": NuSVC, "kernel": "rbf"},
        "NuSVC_poly": {"family": "SVM", "model_class": NuSVC, "kernel": "poly"},
        "NuSVC_sigmoid": {"family": "SVM", "model_class": NuSVC, "kernel": "sigmoid"},
        "LinearSVC": {"family": "SVM", "model_class": LinearSVC, "kernel": None},

        # Mega-friendly kernel approximation (near RBF-SVM accuracy, tiny memory)
        # RFF = Random Fourier Features: RBFSampler -> LinearSVC
        "RFF_RBF_LinearSVC": {"family": "SVM_RFF", "model_class": LinearSVC, "feature_map": "rbf_rff"},
        
        # Tree-based Models
        "DecisionTree": {"family": "Tree", "model_class": DecisionTreeClassifier},
        "RandomForest": {"family": "Tree", "model_class": RandomForestClassifier},
        "ExtraTrees": {"family": "Tree", "model_class": ExtraTreesClassifier},
        "GradientBoosting": {"family": "Tree", "model_class": GradientBoostingClassifier},
        "AdaBoost": {"family": "Tree", "model_class": AdaBoostClassifier},
        
        # Neural Networks
        "MLP_relu": {"family": "Neural Network", "model_class": MLPClassifier, "activation": "relu"},
        "MLP_tanh": {"family": "Neural Network", "model_class": MLPClassifier, "activation": "tanh"},
        "MLP_logistic": {"family": "Neural Network", "model_class": MLPClassifier, "activation": "logistic"},
        
        # K-Nearest Neighbors
        "KNN_uniform": {"family": "KNN", "model_class": KNeighborsClassifier, "weights": "uniform"},
        "KNN_distance": {"family": "KNN", "model_class": KNeighborsClassifier, "weights": "distance"},
        
        # Discriminant Analysis
        "LDA": {"family": "Discriminant", "model_class": LinearDiscriminantAnalysis},
        "QDA": {"family": "Discriminant", "model_class": QuadraticDiscriminantAnalysis},
        
        # Naive Bayes
        "GaussianNB": {"family": "Naive Bayes", "model_class": GaussianNB},
        "BernoulliNB": {"family": "Naive Bayes", "model_class": BernoulliNB},
    }

def estimate_svm_flash_kb(model, n_features: int, n_classes: int) -> float:
    """Rough estimate of flash usage for exported kernel SVM (float, PROGMEM).

    This is NOT an exact Arduino compile size, but it correlates strongly with SV count.
    """
    try:
        if not hasattr(model, 'support_vectors_'):
            return 0.0
        n_sv = int(len(model.support_vectors_))
        # arrays we export (float): support vectors + dual coefs + intercepts + scaler
        sv_bytes = n_sv * int(n_features) * 4
        if n_classes <= 2:
            dual_bytes = n_sv * 4
            intercept_bytes = 4
            nsup_bytes = 0
        else:
            dual_bytes = (n_classes - 1) * n_sv * 4
            intercept_bytes = int(getattr(model, 'intercept_', np.zeros(1)).size) * 4
            nsup_bytes = n_classes * 2  # int16 approx
        scaler_bytes = int(n_features) * 2 * 4
        total = sv_bytes + dual_bytes + intercept_bytes + nsup_bytes + scaler_bytes
        return float(total) / 1024.0
    except Exception:
        return 0.0

def estimate_model_size(model, model_name, scaler, n_features, n_classes):
    """
    U?c lu?ng k?ch thu?c Flash v? RAM cho m?t model.
    Returns: (flash_bytes, ram_bytes, warning_message)
    """
    flash_bytes = 0
    ram_bytes = 0
    warning = ""
    
    family = ""
    if hasattr(model, 'support_vectors_'):
        family = "SVM"
    elif hasattr(model, 'coefs_'):
        family = "MLP"
    elif hasattr(model, '_fit_X'):
        family = "KNN"
    elif hasattr(model, 'tree_'):
        family = "Tree"
    elif hasattr(model, 'coef_'):
        family = "Linear"
    elif hasattr(model, 'theta_'):
        family = "NaiveBayes"
    elif hasattr(model, 'means_'):
        family = "Discriminant"
    
    bytes_per_float = 4 if USE_FLOAT_PRECISION else 8
    
    # Scaler lu?n c?n: mean + scale
    scaler_bytes = n_features * 2 * bytes_per_float
    flash_bytes += scaler_bytes
    
    if family == "SVM":
        n_sv = len(model.support_vectors_)
        # Support vectors
        sv_bytes = n_sv * n_features * bytes_per_float
        flash_bytes += sv_bytes
        
        # Dual coefficients
        if n_classes <= 2:
            dual_bytes = n_sv * bytes_per_float
        else:
            dual_bytes = (n_classes - 1) * n_sv * bytes_per_float
        flash_bytes += dual_bytes
        
        # Intercepts
        if n_classes <= 2:
            intercept_bytes = bytes_per_float
        else:
            n_pairs = n_classes * (n_classes - 1) // 2
            intercept_bytes = n_pairs * bytes_per_float
        flash_bytes += intercept_bytes
        
        # RAM: scaled features + temporary arrays
        ram_bytes = n_features * bytes_per_float * 2
        
        if n_sv > MAX_SUPPORT_VECTORS:
            warning = f" {n_sv} SVs vượt quá giới hạn {MAX_SUPPORT_VECTORS}. Sẽ cắt bớt!"
            
    elif family == "MLP":
        total_neurons = 0
        for i, coef in enumerate(model.coefs_):
            n_in, n_out = coef.shape
            # Weights
            flash_bytes += n_in * n_out * bytes_per_float
            # Biases
            flash_bytes += n_out * bytes_per_float
            total_neurons += n_out
        
        # RAM: layer buffers
        max_layer_size = max([c.shape[1] for c in model.coefs_])
        ram_bytes = max_layer_size * 2 * bytes_per_float + n_features * bytes_per_float
        
        if total_neurons > MAX_MLP_NEURONS:
            warning = f" MLP c? {total_neurons} neurons vượt quá giới hạn {MAX_MLP_NEURONS}"
            
    elif family == "KNN":
        fit_X = getattr(model, '_fit_X', None)
        if fit_X is not None:
            n_samples = min(len(fit_X), MAX_KNN_SAMPLES)
            # Training data
            flash_bytes += n_samples * n_features * bytes_per_float
            # Labels
            flash_bytes += n_samples * 2  # int16
            
            # RAM: distances array + k-nearest
            k = getattr(model, 'n_neighbors', 5)
            ram_bytes = n_samples * bytes_per_float + k * (bytes_per_float + 4)
            
            if len(fit_X) > MAX_KNN_SAMPLES:
                warning = f"⚠️ KNN c? {len(fit_X)} samples, sẽ cắt xuống {MAX_KNN_SAMPLES}"
                
    elif family == "Linear":
        # LinearSVC, LDA coefficients
        coef = model.coef_
        flash_bytes += coef.size * bytes_per_float
        # Intercepts
        flash_bytes += len(model.intercept_) * bytes_per_float
        ram_bytes = n_features * bytes_per_float * 2
        
    elif family == "NaiveBayes":
        if hasattr(model, 'theta_'):
            flash_bytes += model.theta_.size * bytes_per_float
        if hasattr(model, 'var_'):
            flash_bytes += model.var_.size * bytes_per_float
        if hasattr(model, 'class_log_prior_'):
            flash_bytes += len(model.class_log_prior_) * bytes_per_float
        ram_bytes = n_features * bytes_per_float + n_classes * bytes_per_float
        
    elif family == "Discriminant":
        # Means
        flash_bytes += n_classes * n_features * bytes_per_float
        # Priors
        flash_bytes += n_classes * bytes_per_float
        ram_bytes = n_features * bytes_per_float + n_classes * bytes_per_float
    
    # Ki?m tra giới hạn
    if flash_bytes > MAX_FLASH_BYTES:
        warning += f"\n⚠️ Flash {flash_bytes/1024:.1f}KB vượt quá {MAX_FLASH_BYTES/1024:.1f}KB!"
    if ram_bytes > MAX_RAM_BYTES:
        warning += f"\n⚠️ RAM {ram_bytes/1024:.1f}KB vượt quá {MAX_RAM_BYTES/1024:.1f}KB!"
    
    return flash_bytes, ram_bytes, warning.strip()

def limit_support_vectors(model, max_sv=None):
    """
    giới hạn s? Support Vectors b?ng c?ch gi? l?i nh?ng SVs quan tr?ng nh?t.
    Ch? gi? SVs c? |alpha| (dual coefficients) l?n nh?t.
    
    Returns: (model_modified, n_original, n_kept)
    """
    if max_sv is None:
        max_sv = MAX_SUPPORT_VECTORS
        
    if not hasattr(model, 'support_vectors_'):
        return model, 0, 0
    
    n_sv = len(model.support_vectors_)
    if n_sv <= max_sv:
        return model, n_sv, n_sv
    
    print(f"   ⚠️ Model has {n_sv} SVs, exceeding limit of {max_sv}")
    print(f"    Keeping {max_sv} most important SVs...")
    
    # T?nh importance score d?a tr?n |alpha| (dual coefficients)
    dual_coef = model.dual_coef_
    
    # T?ng |alpha| cho m?i SV (qua t?t c? binary classifiers)
    importance = np.sum(np.abs(dual_coef), axis=0)
    
    # Ch?n top max_sv SVs theo importance
    top_indices = np.argsort(importance)[-max_sv:]
    top_indices = np.sort(top_indices)  # Gi? th? t? ban d?u
    
    # C?p nh?t model (t?o b?n copy d? tr?nh modify original)
    import copy
    model_limited = copy.deepcopy(model)
    
    model_limited.support_vectors_ = model.support_vectors_[top_indices]
    model_limited.dual_coef_ = model.dual_coef_[:, top_indices]
    model_limited.support_ = model.support_[top_indices] if hasattr(model, 'support_') else top_indices
    
    # C?p nh?t n_support_ cho multi-class
    if hasattr(model, 'n_support_'):
        # Đếm lại s? SVs cho m?i class
        new_n_support = []
        cumsum = np.cumsum([0] + list(model.n_support_))
        for i in range(len(model.n_support_)):
            start, end = cumsum[i], cumsum[i+1]
            count = np.sum((top_indices >= start) & (top_indices < end))
            new_n_support.append(count)
        model_limited.n_support_ = np.array(new_n_support)
    
    print(f"   ? Reduced from {n_sv} to {max_sv} SVs")
    
    # U?c lu?ng gi?m k?ch thu?c
    n_features = model.support_vectors_.shape[1]
    old_size = n_sv * n_features * 4 / 1024
    new_size = max_sv * n_features * 4 / 1024
    print(f"    SVs size: {old_size:.1f}KB → {new_size:.1f}KB (reduced {(1-new_size/old_size)*100:.1f}%)")
    
    return model_limited, n_sv, max_sv

def print_size_optimization_info():
    """In th?ng tin v? c?i d?t t?i uu k?ch thu?c hi?n t?i"""
    print("\n" + "=" * 80)
    print("  SIZE OPTIMIZATION SETTINGS FOR ARDUINO")
    print("=" * 80)
    print(f"�� Target Board: {TARGET_BOARD.upper()}")
    print(f"   ? ARDUINO_OPTIMIZE: {ARDUINO_OPTIMIZE} {'(PROGMEM enabled)' if ARDUINO_OPTIMIZE else '(RAM mode)'}")
    print(f"   ? Float precision: {'float (4 bytes)' if USE_FLOAT_PRECISION else 'double (8 bytes)'}")
    print(f"   ? Coefficient precision: {COEFFICIENT_PRECISION} decimal places")
    print()
    print(f" MODEL LIMITS:")
    print(f"   ? Max Support Vectors: {MAX_SUPPORT_VECTORS}")
    print(f"   ? Max MLP Neurons: {MAX_MLP_NEURONS}")
    print(f"   ? Max KNN Samples: {MAX_KNN_SAMPLES}")
    print(f"   ? Max Tree Depth: {MAX_TREE_DEPTH}")
    print()
    print(f" MEMORY LIMITS:")
    print(f"   ? Max Flash: {MAX_FLASH_BYTES/1024:.1f} KB")
    print(f"   ? Max RAM: {MAX_RAM_BYTES/1024:.1f} KB")
    print(f"   ? Train for Mega: {TRAIN_FOR_MEGA}")
    print("=" * 80 + "\n")

def objective_rff_linearsvc(trial, X_train, y_train):
    """Objective: RBFSampler (RFF) + LinearSVC. Designed to fit Arduino Mega."""
    n_components = trial.suggest_categorical('n_components', [16, 32, 64])
    gamma = trial.suggest_float('gamma', 1e-4, 10.0, log=True)
    C = trial.suggest_float('C', 0.1, 100, log=True)
    loss = trial.suggest_categorical('loss', ['hinge', 'squared_hinge'])

    pipeline = Pipeline([
        ('rff', RBFSampler(gamma=gamma, n_components=n_components, random_state=RANDOM_STATE)),
        ('lin', LinearSVC(C=C, loss=loss, max_iter=20000, random_state=RANDOM_STATE)),
    ])

    cv = StratifiedKFold(n_splits=CV_FOLDS, shuffle=True, random_state=RANDOM_STATE)
    scores = []
    for fold, (train_idx, val_idx) in enumerate(cv.split(X_train, y_train)):
        X_tr, X_val = X_train[train_idx], X_train[val_idx]
        y_tr, y_val = y_train[train_idx], y_train[val_idx]
        pipeline.fit(X_tr, y_tr)
        score = pipeline.score(X_val, y_val)
        scores.append(score)
        trial.report(score, fold)
        if trial.should_prune():
            raise optuna.TrialPruned()

    return float(sum(scores) / len(scores))

def objective_svc(trial, X_train, y_train, kernel):
    """Objective function cho SVC v?i pruning"""
    params = {
        'C': trial.suggest_float('C', 0.1, 100, log=True),
        'kernel': kernel,
        'random_state': RANDOM_STATE
    }
    
    if kernel == 'rbf' or kernel == 'sigmoid':
        params['gamma'] = trial.suggest_categorical('gamma', ['scale', 'auto'])
    elif kernel == 'poly':
        params['degree'] = trial.suggest_int('degree', 2, 5)
        params['gamma'] = trial.suggest_categorical('gamma', ['scale', 'auto'])
        params['coef0'] = trial.suggest_float('coef0', 0.0, 10.0)
    
    model = SVC(**params)
    cv = StratifiedKFold(n_splits=CV_FOLDS, shuffle=True, random_state=RANDOM_STATE)
    
    # Cross-validation v?i pruning
    scores = []
    n_classes = int(len(np.unique(y_train)))
    n_features = int(X_train.shape[1])
    for fold, (train_idx, val_idx) in enumerate(cv.split(X_train, y_train)):
        X_tr, X_val = X_train[train_idx], X_train[val_idx]
        y_tr, y_val = y_train[train_idx], y_train[val_idx]
        
        model.fit(X_tr, y_tr)
        score = model.score(X_val, y_val)

        # Mega constraint: penalize oversized kernel SVM so Optuna finds small-SV solutions
        if TRAIN_FOR_MEGA:
            est_kb = estimate_svm_flash_kb(model, n_features=n_features, n_classes=n_classes)
            if est_kb > float(MEGA_FLASH_BUDGET_KB):
                # N?u vượt quá Ngan sach Flash, d?ng trial ngay l?p t?c
                trial.report(0.0, fold)
                raise optuna.TrialPruned()
            overflow = max(0.0, (est_kb - float(MEGA_FLASH_BUDGET_KB)) / max(1.0, float(MEGA_FLASH_BUDGET_KB)))
            score = float(score) - float(SVM_FLASH_PENALTY_WEIGHT) * float(overflow)

        scores.append(score)
        
        # Report intermediate value cho pruning
        trial.report(score, fold)
        
        # Pruning: d?ng s?m n?u k?t qu? k?m
        if trial.should_prune():
            raise optuna.TrialPruned()
    
    return sum(scores) / len(scores)

def objective_nusvc(trial, X_train, y_train, kernel):
    """Objective function cho NuSVC v?i pruning"""
    params = {
        'nu': trial.suggest_float('nu', 0.01, 0.99),
        'kernel': kernel,
        'random_state': RANDOM_STATE
    }
    
    if kernel == 'rbf' or kernel == 'sigmoid':
        params['gamma'] = trial.suggest_categorical('gamma', ['scale', 'auto'])
    elif kernel == 'poly':
        params['degree'] = trial.suggest_int('degree', 2, 5)
        params['gamma'] = trial.suggest_categorical('gamma', ['scale', 'auto'])
        params['coef0'] = trial.suggest_float('coef0', 0.0, 10.0)
    
    model = NuSVC(**params)
    cv = StratifiedKFold(n_splits=CV_FOLDS, shuffle=True, random_state=RANDOM_STATE)
    
    scores = []
    n_classes = int(len(np.unique(y_train)))
    n_features = int(X_train.shape[1])
    for fold, (train_idx, val_idx) in enumerate(cv.split(X_train, y_train)):
        X_tr, X_val = X_train[train_idx], X_train[val_idx]
        y_tr, y_val = y_train[train_idx], y_train[val_idx]
        
        model.fit(X_tr, y_tr)
        score = model.score(X_val, y_val)

        if TRAIN_FOR_MEGA:
            est_kb = estimate_svm_flash_kb(model, n_features=n_features, n_classes=n_classes)
            if est_kb > float(MEGA_FLASH_BUDGET_KB):
                trial.report(0.0, fold)
                raise optuna.TrialPruned()
            overflow = max(0.0, (est_kb - float(MEGA_FLASH_BUDGET_KB)) / max(1.0, float(MEGA_FLASH_BUDGET_KB)))
            score = float(score) - float(SVM_FLASH_PENALTY_WEIGHT) * float(overflow)

        scores.append(score)
        
        trial.report(score, fold)
        if trial.should_prune():
            raise optuna.TrialPruned()
    
    return sum(scores) / len(scores)

def objective_tree(trial, X_train, y_train, model_class):
    """Objective function cho Tree-based models"""
    if model_class == DecisionTreeClassifier:
        params = {
            'max_depth': trial.suggest_int('max_depth', 3, 20),
            'min_samples_split': trial.suggest_int('min_samples_split', 2, 20),
            'min_samples_leaf': trial.suggest_int('min_samples_leaf', 1, 10),
            'criterion': trial.suggest_categorical('criterion', ['gini', 'entropy']),
            'random_state': RANDOM_STATE
        }
    elif model_class in [RandomForestClassifier, ExtraTreesClassifier]:
        params = {
            'n_estimators': trial.suggest_int('n_estimators', 50, 300),
            'max_depth': trial.suggest_int('max_depth', 3, 20),
            'min_samples_split': trial.suggest_int('min_samples_split', 2, 20),
            'min_samples_leaf': trial.suggest_int('min_samples_leaf', 1, 10),
            'criterion': trial.suggest_categorical('criterion', ['gini', 'entropy']),
            'random_state': RANDOM_STATE,
            'n_jobs': -1
        }
    elif model_class == GradientBoostingClassifier:
        params = {
            'n_estimators': trial.suggest_int('n_estimators', 50, 200),
            'learning_rate': trial.suggest_float('learning_rate', 0.01, 0.3, log=True),
            'max_depth': trial.suggest_int('max_depth', 3, 10),
            'min_samples_split': trial.suggest_int('min_samples_split', 2, 20),
            'min_samples_leaf': trial.suggest_int('min_samples_leaf', 1, 10),
            'random_state': RANDOM_STATE
        }
    else:  # AdaBoost
        params = {
            'n_estimators': trial.suggest_int('n_estimators', 50, 200),
            'learning_rate': trial.suggest_float('learning_rate', 0.01, 2.0, log=True),
            'random_state': RANDOM_STATE
        }
    
    model = model_class(**params)
    cv = StratifiedKFold(n_splits=CV_FOLDS, shuffle=True, random_state=RANDOM_STATE)
    scores = cross_validate(model, X_train, y_train, cv=cv, scoring='accuracy', n_jobs=-1)
    
    return scores['test_score'].mean()

def objective_mlp(trial, X_train, y_train, activation):
    """Objective function cho MLP"""
    hidden_layer_sizes = []
    n_layers = trial.suggest_int('n_layers', 1, 3)
    for i in range(n_layers):
        hidden_layer_sizes.append(trial.suggest_int(f'n_units_l{i}', 10, 200))
    
    params = {
        'hidden_layer_sizes': tuple(hidden_layer_sizes),
        'activation': activation,
        'solver': trial.suggest_categorical('solver', ['adam', 'sgd']),
        'alpha': trial.suggest_float('alpha', 1e-5, 1e-1, log=True),
        'learning_rate': trial.suggest_categorical('learning_rate', ['constant', 'adaptive']),
        'max_iter': 1000,
        'random_state': RANDOM_STATE
    }
    
    model = MLPClassifier(**params)
    cv = StratifiedKFold(n_splits=CV_FOLDS, shuffle=True, random_state=RANDOM_STATE)
    scores = cross_validate(model, X_train, y_train, cv=cv, scoring='accuracy', n_jobs=-1)
    
    return scores['test_score'].mean()

def objective_knn(trial, X_train, y_train, weights):
    """Objective function cho KNN"""
    params = {
        'n_neighbors': trial.suggest_int('n_neighbors', 3, 30),
        'weights': weights,
        'metric': trial.suggest_categorical('metric', ['euclidean', 'manhattan', 'minkowski']),
        'n_jobs': -1
    }
    
    if params['metric'] == 'minkowski':
        params['p'] = trial.suggest_int('p', 1, 5)
    
    model = KNeighborsClassifier(**params)
    cv = StratifiedKFold(n_splits=CV_FOLDS, shuffle=True, random_state=RANDOM_STATE)
    scores = cross_validate(model, X_train, y_train, cv=cv, scoring='accuracy', n_jobs=-1)
    
    return scores['test_score'].mean()

def objective_lda(trial, X_train, y_train):
    """Objective function cho LDA"""
    params = {
        'solver': trial.suggest_categorical('solver', ['svd', 'lsqr', 'eigen']),
    }
    
    if params['solver'] in ['lsqr', 'eigen']:
        params['shrinkage'] = trial.suggest_float('shrinkage', 0.0, 1.0)
    
    model = LinearDiscriminantAnalysis(**params)
    cv = StratifiedKFold(n_splits=CV_FOLDS, shuffle=True, random_state=RANDOM_STATE)
    scores = cross_validate(model, X_train, y_train, cv=cv, scoring='accuracy', n_jobs=-1)
    
    return scores['test_score'].mean()

def objective_qda(trial, X_train, y_train):
    """Objective function cho QDA"""
    params = {
        'reg_param': trial.suggest_float('reg_param', 0.0, 1.0),
    }
    
    model = QuadraticDiscriminantAnalysis(**params)
    cv = StratifiedKFold(n_splits=CV_FOLDS, shuffle=True, random_state=RANDOM_STATE)
    scores = cross_validate(model, X_train, y_train, cv=cv, scoring='accuracy', n_jobs=-1)
    
    return scores['test_score'].mean()

def objective_naive_bayes(trial, X_train, y_train, model_class):
    """Objective function cho Naive Bayes"""
    if model_class == GaussianNB:
        params = {
            'var_smoothing': trial.suggest_float('var_smoothing', 1e-10, 1e-5, log=True)
        }
    elif model_class == BernoulliNB:
        params = {
            'alpha': trial.suggest_float('alpha', 0.1, 10.0, log=True),
            'binarize': trial.suggest_float('binarize', 0.0, 1.0)
        }
    
    model = model_class(**params)
    cv = StratifiedKFold(n_splits=CV_FOLDS, shuffle=True, random_state=RANDOM_STATE)
    scores = cross_validate(model, X_train, y_train, cv=cv, scoring='accuracy', n_jobs=-1)
    
    return scores['test_score'].mean()

def objective_linearsvc(trial, X_train, y_train):
    """Objective function cho LinearSVC v?i pruning"""
    params = {
        'C': trial.suggest_float('C', 0.1, 100, log=True),
        'loss': trial.suggest_categorical('loss', ['hinge', 'squared_hinge']),
        'max_iter': 10000,
        'random_state': RANDOM_STATE
    }
    
    model = LinearSVC(**params)
    cv = StratifiedKFold(n_splits=CV_FOLDS, shuffle=True, random_state=RANDOM_STATE)
    
    scores = []
    for fold, (train_idx, val_idx) in enumerate(cv.split(X_train, y_train)):
        X_tr, X_val = X_train[train_idx], X_train[val_idx]
        y_tr, y_val = y_train[train_idx], y_train[val_idx]
        
        model.fit(X_tr, y_tr)
        score = model.score(X_val, y_val)
        scores.append(score)
        
        trial.report(score, fold)
        if trial.should_prune():
            raise optuna.TrialPruned()
    
    return sum(scores) / len(scores)

def train_single_model(model_name, config, X_train, y_train, X_test, y_test, scaler):
    """Train m?t model v?i Optuna"""
    print(f"\n{'='*60}")
    print(f"Training: {model_name}")
    print(f"{'='*60}")
    
    try:
        # T?o study Optuna v?i Pruner d? early stopping
        study = optuna.create_study(
            direction='maximize',
            sampler=TPESampler(seed=RANDOM_STATE),
            pruner=MedianPruner(
                n_startup_trials=MIN_TRIALS_BEFORE_PRUNING,
                n_warmup_steps=PRUNING_WARMUP_STEPS,
                interval_steps=1
            )
        )
        
        # Optimize d?a tr?n family
        model_class = config['model_class']

        # Special: RFF (RBFSampler) + LinearSVC
        if config.get('family') == 'SVM_RFF' and config.get('feature_map') == 'rbf_rff':
            study.optimize(
                lambda trial: objective_rff_linearsvc(trial, X_train, y_train),
                n_trials=N_TRIALS,
                timeout=OPTUNA_TIMEOUT,
                show_progress_bar=False
            )

            best_params = study.best_params
            n_complete = len([t for t in study.trials if t.state == optuna.trial.TrialState.COMPLETE])
            n_pruned = len([t for t in study.trials if t.state == optuna.trial.TrialState.PRUNED])
            n_failed = len([t for t in study.trials if t.state == optuna.trial.TrialState.FAIL])
            print(f"📊 Trials: {n_complete} completed, {n_pruned} pruned (early stopped), {n_failed} failed")
            print(f"🧩 Best params: {best_params}")
            print(f"🏅 Best CV score: {study.best_value:.4f}")

            pipeline = Pipeline([
                ('rff', RBFSampler(
                    gamma=float(best_params['gamma']),
                    n_components=int(best_params['n_components']),
                    random_state=RANDOM_STATE
                )),
                ('lin', LinearSVC(
                    C=float(best_params['C']),
                    loss=str(best_params['loss']),
                    max_iter=20000,
                    random_state=RANDOM_STATE
                )),
            ])

            pipeline.fit(X_train, y_train)

            cv = StratifiedKFold(n_splits=CV_FOLDS, shuffle=True, random_state=RANDOM_STATE)
            cv_results = cross_validate(
                pipeline, X_train, y_train, cv=cv,
                scoring=['accuracy', 'precision_macro', 'recall_macro', 'f1_macro'],
                n_jobs=-1,
                return_train_score=True
            )

            y_pred = pipeline.predict(X_test)
            start_time = time.time()
            _ = pipeline.predict(X_test)
            elapsed_time = time.time() - start_time
            pred_speed = len(X_test) / elapsed_time if elapsed_time > 0 else 0
            model_size = len(pickle.dumps(pipeline))
            cm = confusion_matrix(y_test, y_pred)

            results = {
                'Algorithm Family': 'SVM_RFF',
                'Model': model_name,
                'Kernel/Type': 'rbf_rff',
                'Best params (JSON)': json.dumps(best_params),
                'CV Train Accuracy': cv_results['train_accuracy'].mean(),
                'CV Train Precision': cv_results['train_precision_macro'].mean(),
                'CV Train Recall': cv_results['train_recall_macro'].mean(),
                'CV Train F1': cv_results['train_f1_macro'].mean(),
                'CV Val Accuracy': cv_results['test_accuracy'].mean(),
                'CV Val Precision': cv_results['test_precision_macro'].mean(),
                'CV Val Recall': cv_results['test_recall_macro'].mean(),
                'CV Val F1': cv_results['test_f1_macro'].mean(),
                'Test Accuracy': accuracy_score(y_test, y_pred),
                'Test Precision': precision_score(y_test, y_pred, average='macro'),
                'Test Recall': recall_score(y_test, y_pred, average='macro'),
                'Test F1': f1_score(y_test, y_pred, average='macro'),
                'Prediction speed (obs/sec)': pred_speed,
                'Model size (bytes)': model_size
            }

            print(f"📊 CV Train: Acc={results['CV Train Accuracy']*100:.2f}%, Prec={results['CV Train Precision']*100:.2f}%")
            print(f"📊 CV Val  : Acc={results['CV Val Accuracy']*100:.2f}%, Prec={results['CV Val Precision']*100:.2f}%")
            print(f"🎯 Test    : Acc={results['Test Accuracy']*100:.2f}%, Prec={results['Test Precision']*100:.2f}%")

            return pipeline, results, cm
        
        if model_class in [SVC, NuSVC]:
            study.optimize(
                lambda trial: objective_svc(trial, X_train, y_train, config['kernel']) if model_class == SVC 
                else objective_nusvc(trial, X_train, y_train, config['kernel']),
                n_trials=N_TRIALS,
                timeout=OPTUNA_TIMEOUT,
                show_progress_bar=False  # T?t d? gi?m spam
            )
        elif model_class == LinearSVC:
            study.optimize(
                lambda trial: objective_linearsvc(trial, X_train, y_train),
                n_trials=N_TRIALS,
                timeout=OPTUNA_TIMEOUT,
                show_progress_bar=False
            )
        elif config['family'] == 'Tree':
            study.optimize(
                lambda trial: objective_tree(trial, X_train, y_train, model_class),
                n_trials=N_TRIALS,
                timeout=OPTUNA_TIMEOUT,
                show_progress_bar=False
            )
        elif model_class == MLPClassifier:
            study.optimize(
                lambda trial: objective_mlp(trial, X_train, y_train, config['activation']),
                n_trials=N_TRIALS,
                timeout=OPTUNA_TIMEOUT,
                show_progress_bar=False
            )
        elif model_class == KNeighborsClassifier:
            study.optimize(
                lambda trial: objective_knn(trial, X_train, y_train, config['weights']),
                n_trials=N_TRIALS,
                timeout=OPTUNA_TIMEOUT,
                show_progress_bar=False
            )
        elif model_class == LinearDiscriminantAnalysis:
            study.optimize(
                lambda trial: objective_lda(trial, X_train, y_train),
                n_trials=N_TRIALS,
                timeout=OPTUNA_TIMEOUT,
                show_progress_bar=False
            )
        elif model_class == QuadraticDiscriminantAnalysis:
            study.optimize(
                lambda trial: objective_qda(trial, X_train, y_train),
                n_trials=N_TRIALS,
                timeout=OPTUNA_TIMEOUT,
                show_progress_bar=False
            )
        elif model_class in [GaussianNB, BernoulliNB]:
            study.optimize(
                lambda trial: objective_naive_bayes(trial, X_train, y_train, model_class),
                n_trials=N_TRIALS,
                timeout=OPTUNA_TIMEOUT,
                show_progress_bar=False
            )
        
        # L?y best params
        best_params = study.best_params
        
        # Th?ng k? trials
        n_complete = len([t for t in study.trials if t.state == optuna.trial.TrialState.COMPLETE])
        n_pruned = len([t for t in study.trials if t.state == optuna.trial.TrialState.PRUNED])
        n_failed = len([t for t in study.trials if t.state == optuna.trial.TrialState.FAIL])
        
        print(f"📊 Trials: {n_complete} completed, {n_pruned} pruned (early stopped), {n_failed} failed")
        print(f"🧩 Best params: {best_params}")
        print(f"🏅 Best CV score: {study.best_value:.4f}")
        
        # Train model v?i best params
        if 'kernel' in config and config['kernel']:
            best_params['kernel'] = config['kernel']
        if 'activation' in config:
            best_params['activation'] = config['activation']
        if 'weights' in config:
            best_params['weights'] = config['weights']
        if model_class == LinearSVC:
            best_params['max_iter'] = 10000
        if model_class == MLPClassifier:
            best_params['max_iter'] = 1000
        
        # Lo?i b? c?c params kh?ng h?p l?
        params_to_remove = []
        
        # MLP: n_layers v? n_units_l* l? params trung gian, kh?ng ph?i params c?a MLPClassifier
        if model_class == MLPClassifier:
            params_to_remove.extend([k for k in best_params.keys() if k.startswith('n_units_l') or k == 'n_layers'])
        
        # KNN kh?ng c? random_state
        if model_class == KNeighborsClassifier:
            params_to_remove.append('random_state')
        
        # Naive Bayes kh?ng c? random_state
        if model_class in [GaussianNB, BernoulliNB]:
            params_to_remove.append('random_state')
        
        # LDA/QDA kh?ng c? random_state
        if model_class in [LinearDiscriminantAnalysis, QuadraticDiscriminantAnalysis]:
            params_to_remove.append('random_state')
        
        # X?a c?c params kh?ng h?p l?
        for param in params_to_remove:
            best_params.pop(param, None)
        
        # Th?m random_state cho c?c model support
        if model_class not in [KNeighborsClassifier, GaussianNB, BernoulliNB, 
                                LinearDiscriminantAnalysis, QuadraticDiscriminantAnalysis]:
            best_params['random_state'] = RANDOM_STATE
        
        model = model_class(**best_params)
        model.fit(X_train, y_train)

        svm_flash_kb = None
        if model_class in [SVC, NuSVC]:
            try:
                n_classes = int(len(np.unique(y_train)))
                n_features = int(X_train.shape[1])
                svm_flash_kb = float(estimate_svm_flash_kb(model, n_features=n_features, n_classes=n_classes))
                if TRAIN_FOR_MEGA and svm_flash_kb > float(MEGA_FLASH_BUDGET_KB):
                    print(f"⚠️  Model {model_name} exceeds Mega Flash: {svm_flash_kb:.1f} KB > {float(MEGA_FLASH_BUDGET_KB):.1f} KB. Skipping this model!")
                    return None, None, None
            except Exception:
                svm_flash_kb = None
        
        # Cross-validation metrics
        cv = StratifiedKFold(n_splits=CV_FOLDS, shuffle=True, random_state=RANDOM_STATE)
        cv_results = cross_validate(
            model, X_train, y_train, cv=cv,
            scoring=['accuracy', 'precision_macro', 'recall_macro', 'f1_macro'],
            n_jobs=-1,
            return_train_score=True
        )
        
        # Test metrics
        y_pred = model.predict(X_test)
        
        # Prediction speed
        start_time = time.time()
        _ = model.predict(X_test)
        elapsed_time = time.time() - start_time
        pred_speed = len(X_test) / elapsed_time if elapsed_time > 0 else 0
        
        # Model size
        model_size = len(pickle.dumps(model))
        
        # Confusion matrix
        cm = confusion_matrix(y_test, y_pred)
        
        results = {
            'Algorithm Family': config.get('family', 'SVM'),
            'Model': model_name,
            'Kernel/Type': config.get('kernel') or config.get('activation') or config.get('weights') or 'N/A',
            'Best params (JSON)': json.dumps(best_params),
            'CV Train Accuracy': cv_results['train_accuracy'].mean(),
            'CV Train Precision': cv_results['train_precision_macro'].mean(),
            'CV Train Recall': cv_results['train_recall_macro'].mean(),
            'CV Train F1': cv_results['train_f1_macro'].mean(),
            'CV Val Accuracy': cv_results['test_accuracy'].mean(),
            'CV Val Precision': cv_results['test_precision_macro'].mean(),
            'CV Val Recall': cv_results['test_recall_macro'].mean(),
            'CV Val F1': cv_results['test_f1_macro'].mean(),
            'Test Accuracy': accuracy_score(y_test, y_pred),
            'Test Precision': precision_score(y_test, y_pred, average='macro'),
            'Test Recall': recall_score(y_test, y_pred, average='macro'),
            'Test F1': f1_score(y_test, y_pred, average='macro'),
            'Prediction speed (obs/sec)': pred_speed,
            'Model size (bytes)': model_size
        }

        if svm_flash_kb is not None:
            results['SVM Flash est (KB)'] = svm_flash_kb
        
        print(f"📊 CV Train: Acc={results['CV Train Accuracy']*100:.2f}%, Prec={results['CV Train Precision']*100:.2f}%")
        print(f"📊 CV Val  : Acc={results['CV Val Accuracy']*100:.2f}%, Prec={results['CV Val Precision']*100:.2f}%")
        print(f"🎯 Test    : Acc={results['Test Accuracy']*100:.2f}%, Prec={results['Test Precision']*100:.2f}%")
        
        return model, results, cm
    
    except Exception as e:
        print(f"\n Error training {model_name}!")
        print(f"   Error type: {type(e).__name__}")
        print(f"   Details: {str(e)[:200]}")
        import traceback
        print(f"   Traceback (shortened):")
        tb_lines = traceback.format_exc().split('\n')
        for line in tb_lines[-5:]:  # Ch? hi?n 5 d?ng cu?i
            if line.strip():
                print(f"     {line}")
        
        # Tr? v? None d? b? qua model n?y
        return None, None, None

def train_all_models(X_train, y_train, X_test, y_test):
    """Train tất cả các models"""
    print("\n" + "=" * 80)
    print("START TRAINING ALL MODELS")
    print("=" * 80)
    
    # In th?ng tin c?i d?t t?i uu k?ch thu?c
    print_size_optimization_info()
    
    # T?t optuna logging d? gi?m spam
    optuna.logging.set_verbosity(optuna.logging.WARNING)
    
    # Chu?n h?a d? li?u
    scaler = StandardScaler()
    X_train_scaled = scaler.fit_transform(X_train)
    X_test_scaled = scaler.transform(X_test)
    
    model_configs = get_model_configs()
    
    all_results = []
    trained_models = {}
    confusion_matrices = {}
    
    n_features = X_train.shape[1]
    n_classes = len(np.unique(y_train))
    
    for model_name, config in model_configs.items():
        try:
            model, results, cm = train_single_model(
                model_name, config, X_train_scaled, y_train, X_test_scaled, y_test, scaler
            )
            
            # Ch? th?m v?o k?t qu? n?u train th?nh c?ng
            if model is not None and results is not None:
                # Áp dụng giới hạn Support Vectors cho SVM models
                if hasattr(model, 'support_vectors_') and len(model.support_vectors_) > MAX_SUPPORT_VECTORS:
                    model, n_orig, n_kept = limit_support_vectors(model, MAX_SUPPORT_VECTORS)
                    results['SVs_original'] = n_orig
                    results['SVs_limited'] = n_kept
                
                # U?c lu?ng k?ch thu?c model
                flash_est, ram_est, size_warning = estimate_model_size(
                    model, model_name, scaler, n_features, n_classes
                )
                results['Flash_est_KB'] = flash_est / 1024
                results['RAM_est_KB'] = ram_est / 1024
                
                if size_warning:
                    print(f"   {size_warning}")
                    # Bỏ qua model nếu vượt quá giới hạn Flash nghiêm trọng
                    if flash_est > MAX_FLASH_BYTES * 1.5:  # Cho ph?p 50% buffer
                        print(f"   ⚠️  Model too large, skipping!")
                        continue
                
                all_results.append(results)
                trained_models[model_name] = {'model': model, 'scaler': scaler, 'family': config.get('family', 'SVM')}
                confusion_matrices[model_name] = cm
            else:
                print(f"⚠️  Skipping {model_name} due to train failure\n")
            
        except Exception as e:
            print(f"⚠️  Exception error training {model_name}: {e}\n")
            continue
    
    # In t?ng k?t k?ch thu?c c?c models
    print("\n" + "=" * 80)
    print("📊 MODEL SIZE SUMMARY (estimated)")
    print("=" * 80)
    print(f"{'Model':<30} {'Flash (KB)':<12} {'RAM (KB)':<12} {'Status':<20}")
    print("-" * 80)
    
    total_flash = 0
    for result in all_results:
        model_name = result['Model']
        flash_kb = result.get('Flash_est_KB', 0)
        ram_kb = result.get('RAM_est_KB', 0)
        total_flash += flash_kb
        
        status = "✅ OK"
        if flash_kb > MAX_FLASH_BYTES / 1024:
            status = " Vượt Flash!"
        elif ram_kb > MAX_RAM_BYTES / 1024:
            status = " Vượt RAM!"
            
        print(f"{model_name:<30} {flash_kb:<12.1f} {ram_kb:<12.1f} {status:<20}")
    
    print("-" * 80)
    print(f"{'TOTAL':<30} {total_flash:<12.1f} {'---':<12}")
    print(f"\n  Note: This is an estimate. Compile in Arduino IDE to see actual size.")
    print("=" * 80 + "\n")
    
    return all_results, trained_models, confusion_matrices

def save_results(all_results, confusion_matrices):
    """Save training results to Excel and plot confusion matrices"""
    print("\n" + "=" * 80)
    print("SAVE RESULTS")
    print("=" * 80)
    
    # Create output folder
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    
    # Save results to Excel
    results_df = pd.DataFrame(all_results)
    # Results will be appended to results.xlsx when using Gradio
    
    # Hi?n th? top 10 models theo Test Precision
    print("\nTop 10 models theo Test Precision:")
    top_models = results_df.nlargest(10, 'Test Precision')[['Model', 'CV Train Accuracy', 'CV Val Accuracy', 'Test Accuracy', 'Test Precision']]
    print(top_models.to_string(index=False))
    
    # V? confusion matrices
    cm_dir = f"{OUTPUT_DIR}/confusion_matrices"
    os.makedirs(cm_dir, exist_ok=True)
    
    for model_name, cm in confusion_matrices.items():
        plt.figure(figsize=(8, 6))
        sns.heatmap(cm, annot=True, fmt='d', cmap='Blues')
        plt.title(f'Confusion Matrix - {model_name}')
        plt.ylabel('True Label')
        plt.xlabel('Predicted Label')
        plt.tight_layout()
        
        safe_name = model_name.replace('-', '_').replace(' ', '_')
        plt.savefig(f"{cm_dir}/{safe_name}_confusion_matrix.png", dpi=100, bbox_inches='tight')
        plt.close()
    
    print(f"✓ Saved {len(confusion_matrices)} confusion matrices at: {cm_dir}/")

def select_models_for_arduino(all_results, trained_models):
    """Cho ph?p ngu?i d?ng ch?n models d? t?o Arduino library"""
    print("\n" + "=" * 80)
    print("SELECT MODELS FOR ARDUINO LIBRARY")
    print("=" * 80)
    
    # S?p x?p theo Test Precision
    results_df = pd.DataFrame(all_results)
    results_df = results_df.sort_values('Test Precision', ascending=False).reset_index(drop=True)
    
    # Hi?n th? b?ng v?i format d?p hon
    print("\n📋 ALL MODELS LIST (sorted by Test Precision):")
    print("=" * 130)
    print(f"{'STT':<5} {'Model':<25} {'Family':<15} {'CV Train':<12} {'CV Val':<12} {'Test Acc':<12} {'Test Prec':<12}")
    print("-" * 130)
    
    for idx, row in results_df.iterrows():
        print(f"{idx+1:<5} {row['Model']:<25} {row['Algorithm Family']:<15} "
              f"{row['CV Train Accuracy']:<12.4f} {row['CV Val Accuracy']:<12.4f} {row['Test Accuracy']:<12.4f} {row['Test Precision']:<12.4f}")
    
    print("=" * 130)
    
    # Ch?n models
    selected_models = {}
    
    if IN_COLAB:
        # GOOGLE COLAB: Sử dụng input() - sẽ hiện input box ngay dưới cell
        print("\n MODEL SELECTION GUIDE:")
        print("   ➡️ Press Enter (empty): Auto-select 10 best models")
        print("   🔢 Enter numbers: Select models by index (any number)")
        print("     Example: '1 2 3 4 5' or '1,2,3,4,5' (select first 5)")
        print("     Example: '1 5 10 15' (select specific 4 models)")
        print()
        
        user_input = input(" Nhập STT các models (cách nhau bởi dấu cách hoặc dấu phẩy): ").strip()
        
        # Xử lý input
        if not user_input:
            # AUTO MODE - không nhập gì
            top_n = min(10, len(results_df))
            print(f"\n✅ AUTO mode: Selected {top_n} models with highest Test Precision")
            print("-" * 80)
            
            for i in range(top_n):
                model_name = results_df.iloc[i]['Model']
                selected_models[model_name] = trained_models[model_name]
                print(f"   {i+1:2d}. {model_name:<30} • Precision: {results_df.iloc[i]['Test Precision']:.4f}")
            
            print("-" * 80)
        else:
            # MANUAL MODE - nhập các số
            try:
                # Parse numbers - hỗ trợ cả dấu cách và dấu phẩy
                user_input = user_input.replace(',', ' ')
                numbers = []
                for x in user_input.split():
                    x = x.strip()
                    if x.isdigit():
                        num = int(x)
                        if 1 <= num <= len(results_df):
                            if num not in numbers:  # Tránh trùng lặp
                                numbers.append(num)
                        else:
                            print(f"⚠️  Skipping number {num} (out of range 1-{len(results_df)})")
                
                if numbers:
                    print(f"\n✅ Selected {len(numbers)} models:")
                    print("-" * 80)
                    for num in sorted(numbers):
                        model_name = results_df.iloc[num - 1]['Model']
                        selected_models[model_name] = trained_models[model_name]
                        print(f"   {num:2d}. {model_name:<30} • Precision: {results_df.iloc[num - 1]['Test Precision']:.4f}")
                    print("-" * 80)
                else:
                    # Không có số hợp lệ -> fallback AUTO
                    print(f"\n⚠️  No valid numbers! Using AUTO mode (10 best models)...")
                    print("-" * 80)
                    top_n = min(10, len(results_df))
                    for i in range(top_n):
                        model_name = results_df.iloc[i]['Model']
                        selected_models[model_name] = trained_models[model_name]
                        print(f"   {i+1:2d}. {model_name:<30} • Precision: {results_df.iloc[i]['Test Precision']:.4f}")
                    print("-" * 80)
                        
            except Exception as e:
                print(f"\n⚠️  Error processing input: {e}")
                print("   Using AUTO mode (10 best models)...")
                print("-" * 80)
                top_n = min(10, len(results_df))
                for i in range(top_n):
                    model_name = results_df.iloc[i]['Model']
                    selected_models[model_name] = trained_models[model_name]
                    print(f"   {i+1:2d}. {model_name:<30} • Precision: {results_df.iloc[i]['Test Precision']:.4f}")
                print("-" * 80)
    
    else:
        # LOCAL MODE: T? d?ng ch?n AUTO
        print("\n  [Local mode] Auto-selecting 10 best models using AUTO mode")
        top_n = min(10, len(results_df))
        
        for i in range(top_n):
            model_name = results_df.iloc[i]['Model']
            selected_models[model_name] = trained_models[model_name]
        print(f"✅ COMPLETED: Selected {len(selected_models)} models to create Arduino library")
    
    # T?ng k?t
    print("\n" + "=" * 100)
    print(f"✅ COMPLETED: Selected {len(selected_models)} models to create Arduino library")
    print("=" * 100 + "\n")
    
    return selected_models

def generate_svm_c_code(model, model_name, scaler, n_features, n_classes):
    """
    Generate C code cho SVM models (linear, poly, rbf, sigmoid)
    Thay th? m2cgen v?i custom lightweight implementation
    """
    safe_name = get_safe_name(model_name)
    func_name = get_func_name(model_name)
    
    # Ki?m tra kernel type
    if hasattr(model, 'kernel'):
        kernel = model.kernel
    else:
        # LinearSVC kh?ng c? kernel attribute
        kernel = 'linear'
    
    # L?y gamma
    if kernel != 'linear' and hasattr(model, 'gamma'):
        gamma = model.gamma if isinstance(model.gamma, float) else model._gamma
        if gamma == 'scale':
            gamma = 1.0 / (n_features * model.support_vectors_.var())
        elif gamma == 'auto':
            gamma = 1.0 / n_features
    else:
        gamma = 1.0
    
    coef0 = model.coef0 if hasattr(model, 'coef0') else 0.0
    degree = model.degree if hasattr(model, 'degree') else 3
    
    code_parts = []
    
    # X?c d?nh ki?u d? li?u v? precision d?a tr?n settings
    data_type = "float" if USE_FLOAT_PRECISION else "double"
    precision = COEFFICIENT_PRECISION if ARDUINO_OPTIMIZE else 10
    type_suffix = "f" if USE_FLOAT_PRECISION else ""
    progmem = " PROGMEM" if ARDUINO_OPTIMIZE else ""
    
    # Scaler parameters - T?i uu h?a memory n?u b?t ARDUINO_OPTIMIZE
    opt_note = " (Memory optimized)" if ARDUINO_OPTIMIZE else ""
    code_parts.append(f"""
// ========================================
// Model: {model_name} (Kernel: {kernel}){opt_note}
// ========================================

// Scaler parameters (stored in {"FLASH" if ARDUINO_OPTIMIZE else "RAM"})
static const {data_type} scaler_mean_{safe_name}[{n_features}]{progmem} = {{
    {', '.join(f'{x:.{precision}f}{type_suffix}' for x in scaler.mean_)}
}};

static const {data_type} scaler_scale_{safe_name}[{n_features}]{progmem} = {{
    {', '.join(f'{x:.{precision}f}{type_suffix}' for x in scaler.scale_)}
}};

// PROGMEM-safe readers (AVR needs pgm_read_*, others can dereference directly)
static inline {data_type} read_{safe_name}(const {data_type}* p) {{
#if defined(ARDUINO_ARCH_AVR)
    // On AVR, double == float (4 bytes) so pgm_read_float is safe for both.
    return ({data_type})pgm_read_float((const float*)p);
#else
    return *p;
#endif
}}

static inline int read_int_{safe_name}(const int* p) {{
#if defined(ARDUINO_ARCH_AVR)
    return (int)pgm_read_word((const uint16_t*)p);
#else
    return *p;
#endif
}}

// Standardize features
static inline void standardize_{safe_name}(const float* features, {data_type}* scaled) {{""")
    
    # Code d?c t? PROGMEM ho?c tr?c ti?p t? RAM
    if ARDUINO_OPTIMIZE:
        code_parts.append(f"""
    for (int i = 0; i < {n_features}; i++) {{
        {data_type} mean_val = read_{safe_name}(&scaler_mean_{safe_name}[i]);
        {data_type} scale_val = read_{safe_name}(&scaler_scale_{safe_name}[i]);
        scaled[i] = (features[i] - mean_val) / scale_val;
    }}
}}
""")
    else:
        code_parts.append(f"""
    for (int i = 0; i < {n_features}; i++) {{
        scaled[i] = (({data_type})features[i] - scaler_mean_{safe_name}[i]) / scaler_scale_{safe_name}[i];
    }}
}}
""")
    
    # Kernel functions - T?i uu h?a theo settings
    if kernel == 'poly':
        if ARDUINO_OPTIMIZE:
            code_parts.append(f"""
// Polynomial kernel (PROGMEM optimized)
static inline {data_type} poly_kernel_{safe_name}(const {data_type}* x, const {data_type}* sv_progmem, int sv_idx) {{
    {data_type} dot = 0.0{type_suffix};
    for (int i = 0; i < {n_features}; i++) {{
        {data_type} sv_val = read_{safe_name}(&sv_progmem[sv_idx * {n_features} + i]);
        dot += x[i] * sv_val;
    }}
    {data_type} result = {gamma:.{precision}f}{type_suffix} * dot + {coef0:.{precision}f}{type_suffix};
    {data_type} powered = result;
    for (int d = 1; d < {degree}; d++) powered *= result;
    return powered;
}}
""")
        else:
            code_parts.append(f"""
// Polynomial kernel
static inline {data_type} poly_kernel_{safe_name}(const {data_type}* x, const {data_type}* sv) {{
    {data_type} dot = 0.0{type_suffix};
    for (int i = 0; i < {n_features}; i++) dot += x[i] * sv[i];
    {data_type} result = {gamma:.{precision}f}{type_suffix} * dot + {coef0:.{precision}f}{type_suffix};
    {data_type} powered = result;
    for (int d = 1; d < {degree}; d++) powered *= result;
    return powered;
}}
""")
    elif kernel == 'rbf':
        if ARDUINO_OPTIMIZE:
            code_parts.append(f"""
// RBF kernel (PROGMEM optimized)
static inline {data_type} rbf_kernel_{safe_name}(const {data_type}* x, const {data_type}* sv_progmem, int sv_idx) {{
    {data_type} dist_sq = 0.0{type_suffix};
    for (int i = 0; i < {n_features}; i++) {{
        {data_type} sv_val = read_{safe_name}(&sv_progmem[sv_idx * {n_features} + i]);
        {data_type} diff = x[i] - sv_val;
        dist_sq += diff * diff;
    }}
    return {'expf' if USE_FLOAT_PRECISION else 'exp'}(-{gamma:.{precision}f}{type_suffix} * dist_sq);
}}
""")
        else:
            code_parts.append(f"""
// RBF kernel
static inline {data_type} rbf_kernel_{safe_name}(const {data_type}* x, const {data_type}* sv) {{
    {data_type} dist_sq = 0.0{type_suffix};
    for (int i = 0; i < {n_features}; i++) {{
        {data_type} diff = x[i] - sv[i];
        dist_sq += diff * diff;
    }}
    return {'expf' if USE_FLOAT_PRECISION else 'exp'}(-{gamma:.{precision}f}{type_suffix} * dist_sq);
}}
""")
    elif kernel == 'sigmoid':
        if ARDUINO_OPTIMIZE:
            code_parts.append(f"""
// Sigmoid kernel (PROGMEM optimized)
static inline {data_type} sigmoid_kernel_{safe_name}(const {data_type}* x, const {data_type}* sv_progmem, int sv_idx) {{
    {data_type} dot = 0.0{type_suffix};
    for (int i = 0; i < {n_features}; i++) {{
        {data_type} sv_val = read_{safe_name}(&sv_progmem[sv_idx * {n_features} + i]);
        dot += x[i] * sv_val;
    }}
    return {'tanhf' if USE_FLOAT_PRECISION else 'tanh'}({gamma:.{precision}f}{type_suffix} * dot + {coef0:.{precision}f}{type_suffix});
}}
""")
        else:
            code_parts.append(f"""
// Sigmoid kernel
static inline {data_type} sigmoid_kernel_{safe_name}(const {data_type}* x, const {data_type}* sv) {{
    {data_type} dot = 0.0{type_suffix};
    for (int i = 0; i < {n_features}; i++) dot += x[i] * sv[i];
    return {'tanhf' if USE_FLOAT_PRECISION else 'tanh'}({gamma:.{precision}f}{type_suffix} * dot + {coef0:.{precision}f}{type_suffix});
}}
""")
    elif kernel == 'linear':
        if ARDUINO_OPTIMIZE:
            code_parts.append(f"""
// Linear kernel (dot product, PROGMEM optimized)
static inline {data_type} linear_kernel_{safe_name}(const {data_type}* x, const {data_type}* sv_progmem, int sv_idx) {{
    {data_type} dot = 0.0{type_suffix};
    for (int i = 0; i < {n_features}; i++) {{
        {data_type} sv_val = read_{safe_name}(&sv_progmem[sv_idx * {n_features} + i]);
        dot += x[i] * sv_val;
    }}
    return dot;
}}
""")
        else:
            code_parts.append(f"""
// Linear kernel (dot product)
static inline {data_type} linear_kernel_{safe_name}(const {data_type}* x, const {data_type}* sv) {{
    {data_type} dot = 0.0{type_suffix};
    for (int i = 0; i < {n_features}; i++) dot += x[i] * sv[i];
    return dot;
}}
""")
    
    # Support vectors v? prediction
    if hasattr(model, 'support_vectors_'):
        # SVC/NuSVC
        sv = model.support_vectors_
        n_sv_original = len(sv)
        
        # ⚠️ CHỈ giới hạn SVs khi ARDUINO_OPTIMIZE = True (Arduino Mega)
        # ESP32 có đủ bộ nhớ, không cần limiting!
        if ARDUINO_OPTIMIZE and n_sv_original > MAX_SUPPORT_VECTORS:
            print(f"   ⚠️ Arduino Mega: Model has {n_sv_original} SVs, limiting to {MAX_SUPPORT_VECTORS}...")
            # Ch?n SVs quan tr?ng nh?t (c? |alpha| l?n nh?t)
            dual_coef = model.dual_coef_
            importance = np.sum(np.abs(dual_coef), axis=0)
            top_indices = np.argsort(importance)[-MAX_SUPPORT_VECTORS:]
            top_indices = np.sort(top_indices)
            
            sv = sv[top_indices]
            if n_classes == 2:
                limited_alpha = model.dual_coef_[0][top_indices]
            else:
                limited_dual_coef = model.dual_coef_[:, top_indices]
            
            print(f"   ✂️ Reduced SVs: {n_sv_original} → {len(sv)}")
        
        # Flatten SVs th?nh 1D array d? d?ng PROGMEM hi?u qu? hon
        sv_flat = sv.flatten()
        n_sv_total = len(sv)
        
        # T?nh k?ch thu?c u?c lu?ng
        size_sv = n_sv_total * n_features * 4  # float = 4 bytes
        size_kb = size_sv / 1024
        print(f"    SVs size: {n_sv_total} vectors ? {n_features} features ? 4 bytes = {size_sv:,} bytes ({size_kb:.1f} KB)")
        
        if size_sv > MAX_FLASH_BYTES:
            print(f"     Model still too large for {TARGET_BOARD.upper()}! ({size_kb:.1f}KB > {MAX_FLASH_BYTES/1024:.1f}KB)")
            print(f"   💡 Suggestion: Reduce MAX_SUPPORT_VECTORS or choose different model (LinearSVC, LDA)")
        elif size_sv > MAX_FLASH_BYTES * 0.7:
            print(f"   ⚠️  Model uses >70% Flash. May not have room for other code.")
        
        # Chia thành nhiều dòng cho dễ đọc
        chunk_size = 10
        sv_chunks = [sv_flat[i:i+chunk_size] for i in range(0, len(sv_flat), chunk_size)]
        # Remove trailing comma by joining properly
        sv_lines_list = [', '.join(f'{x:.{precision}f}{type_suffix}' for x in chunk) for chunk in sv_chunks]
        sv_lines = ',\n    '.join(sv_lines_list)
        
        if n_classes == 2:
            # Binary classification
            alpha = limited_alpha if ('limited_alpha' in locals()) else model.dual_coef_[0]
            intercept = model.intercept_[0]
            
            if ARDUINO_OPTIMIZE:
                code_parts.append(f"""
// Support vectors (flattened, {n_sv_total} vectors ? {n_features} features)
static const {data_type} sv_{safe_name}[{len(sv_flat)}]{progmem} = {{
    {sv_lines}
}};

// Dual coefficients
static const {data_type} alpha_{safe_name}[{len(sv)}]{progmem} = {{
    {', '.join(f'{a:.{precision}f}{type_suffix}' for a in alpha)}
}};

int predict_{func_name}(const float* features) {{
    {data_type} scaled[{n_features}];
    standardize_{safe_name}(features, scaled);
    
    {data_type} decision = {intercept:.{precision}f}{type_suffix};
    for (int i = 0; i < {len(sv)}; i++) {{
        {data_type} alpha_val = read_{safe_name}(&alpha_{safe_name}[i]);
        decision += alpha_val * {kernel}_kernel_{safe_name}(scaled, sv_{safe_name}, i);
    }}
    
    return decision >= 0.0{type_suffix} ? {model.classes_[1]} : {model.classes_[0]};
}}
""")
            else:
                # Non-optimized version (ESP32)
                sv_lines = ',\n'.join('    {' + ', '.join(f'{x:.{precision}f}{type_suffix}' for x in vec) + '}' for vec in sv)
                code_parts.append(f"""
// Support vectors
static const {data_type} sv_{safe_name}[{n_sv_total}][{n_features}] = {{
{sv_lines}
}};

// Dual coefficients
static const {data_type} alpha_{safe_name}[{len(sv)}] = {{
    {', '.join(f'{a:.{precision}f}{type_suffix}' for a in alpha)}
}};

int predict_{func_name}(const float* features) {{
    {data_type} scaled[{n_features}];
    standardize_{safe_name}(features, scaled);
    
    {data_type} decision = {intercept:.{precision}f}{type_suffix};
    for (int i = 0; i < {len(sv)}; i++) {{
        decision += alpha_{safe_name}[i] * {kernel}_kernel_{safe_name}(scaled, sv_{safe_name}[i]);
    }}
    
    return decision >= 0.0{type_suffix} ? {model.classes_[1]} : {model.classes_[0]};
}}
""")
        else:
            # Multi-class OvO
            # Cho ESP32: dùng trực tiếp model data, KHÔNG limiting
            # Cho Arduino Mega: dùng limited data nếu đã được tạo
            if ARDUINO_OPTIMIZE:
                # Arduino Mega - có thể có limiting
                if 'limited_dual_coef' in locals():
                    dual_coef_to_use = limited_dual_coef
                    # C?n t?nh l?i n_support_ d?a tr?n top_indices
                    n_sv = []
                    cumsum = np.cumsum([0] + list(model.n_support_))
                    for i in range(len(model.n_support_)):
                        start, end = cumsum[i], cumsum[i+1]
                        count = np.sum((top_indices >= start) & (top_indices < end))
                        n_sv.append(count)
                    n_sv = np.array(n_sv)
                else:
                    dual_coef_to_use = model.dual_coef_
                    n_sv = model.n_support_
                
                # Flatten dual_coef for OPTIMIZED path
                dual_coef_flat = dual_coef_to_use.flatten()
                dual_coef_chunks = [dual_coef_flat[i:i+chunk_size] for i in range(0, len(dual_coef_flat), chunk_size)]
                dual_coef_lines_list = [', '.join(f'{x:.{precision}f}{type_suffix}' for x in chunk) for chunk in dual_coef_chunks]
                dual_coef_lines = ',\n    '.join(dual_coef_lines_list)
            
                code_parts.append(f"""
// Support vectors (flattened, {n_sv_total} vectors × {n_features} features)
static const {data_type} sv_{safe_name}[{len(sv_flat)}]{progmem} = {{
    {sv_lines}
}};

// Number of support vectors per class
static const int n_support_{safe_name}[{n_classes}]{progmem} = {{{', '.join(map(str, n_sv))}}};

// Dual coefficients (flattened from [{n_classes - 1}][{n_sv_total}])
static const {data_type} dual_coef_{safe_name}[{len(dual_coef_flat)}]{progmem} = {{
    {dual_coef_lines}
}};

// Intercepts for each binary classifier
static const {data_type} intercepts_{safe_name}[{len(model.intercept_)}]{progmem} = {{
    {', '.join(f'{x:.{precision}f}{type_suffix}' for x in model.intercept_)}
}};

int predict_{func_name}(const float* features) {{
    {data_type} scaled[{n_features}];
    standardize_{safe_name}(features, scaled);
    
    int votes[{n_classes}] = {{0}};
    int clf_idx = 0;
    
    // Calculate support vector start indices
    int sv_start[{n_classes + 1}];
    sv_start[0] = 0;
    for (int i = 0; i < {n_classes}; i++) {{
        sv_start[i + 1] = sv_start[i] + read_int_{safe_name}(&n_support_{safe_name}[i]);
    }}
    
    // One-vs-One classification
    for (int i = 0; i < {n_classes}; i++) {{
        for (int j = i + 1; j < {n_classes}; j++) {{
            {data_type} decision = read_{safe_name}(&intercepts_{safe_name}[clf_idx]);
            
            // Support vectors from class i
            for (int k = sv_start[i]; k < sv_start[i + 1]; k++) {{
                {data_type} coef = read_{safe_name}(&dual_coef_{safe_name}[(j - 1) * {n_sv_total} + k]);
                decision += coef * {kernel}_kernel_{safe_name}(scaled, sv_{safe_name}, k);
            }}
            
            // Support vectors from class j
            for (int k = sv_start[j]; k < sv_start[j + 1]; k++) {{
                {data_type} coef = read_{safe_name}(&dual_coef_{safe_name}[i * {n_sv_total} + k]);
                decision += coef * {kernel}_kernel_{safe_name}(scaled, sv_{safe_name}, k);
            }}
            
            if (decision > 0.0{type_suffix}) votes[i]++; else votes[j]++;
            clf_idx++;
        }}
    }}
    
    // Return class with most votes
    int max_votes = votes[0], predicted = 0;
    for (int i = 1; i < {n_classes}; i++) {{
        if (votes[i] > max_votes) {{
            max_votes = votes[i];
            predicted = i;
        }}
    }}
    
    static const int class_labels[{n_classes}] = {{{', '.join(map(str, model.classes_))}}};
    return class_labels[predicted];
}}
""")
            else:
                # ESP32 - NON-OPTIMIZED, giống g15B.py
                # Dùng trực tiếp model data, KHÔNG limiting
                sv_lines = ',\n'.join('    {' + ', '.join(f'{x:.{precision}f}{type_suffix}' for x in vec) + '}' for vec in sv)
                dual_coef_lines = ',\n'.join('    {' + ', '.join(f'{x:.{precision}f}{type_suffix}' for x in row) + '}' for row in model.dual_coef_)
                
                code_parts.append(f"""
// Support vectors
static const {data_type} sv_{safe_name}[{n_sv_total}][{n_features}] = {{
{sv_lines}
}};

// Number of support vectors per class
static const int n_support_{safe_name}[{n_classes}] = {{{', '.join(map(str, model.n_support_))}}};

// Dual coefficients
static const {data_type} dual_coef_{safe_name}[{n_classes - 1}][{n_sv_total}] = {{
{dual_coef_lines}
}};

// Intercepts
static const {data_type} intercepts_{safe_name}[{len(model.intercept_)}] = {{
    {', '.join(f'{x:.{precision}f}{type_suffix}' for x in model.intercept_)}
}};

int predict_{func_name}(const float* features) {{
    {data_type} scaled[{n_features}];
    standardize_{safe_name}(features, scaled);
    
    int votes[{n_classes}] = {{0}};
    int clf_idx = 0;
    
    // Calculate support vector start indices
    int sv_start[{n_classes + 1}];
    sv_start[0] = 0;
    for (int i = 0; i < {n_classes}; i++) {{
        sv_start[i + 1] = sv_start[i] + n_support_{safe_name}[i];
    }}
    
    // One-vs-One classification
    for (int i = 0; i < {n_classes}; i++) {{
        for (int j = i + 1; j < {n_classes}; j++) {{
            {data_type} decision = intercepts_{safe_name}[clf_idx];
            
            // Support vectors from class i
            for (int k = sv_start[i]; k < sv_start[i + 1]; k++) {{
                decision += dual_coef_{safe_name}[j - 1][k] * {kernel}_kernel_{safe_name}(scaled, sv_{safe_name}[k]);
            }}
            
            // Support vectors from class j
            for (int k = sv_start[j]; k < sv_start[j + 1]; k++) {{
                decision += dual_coef_{safe_name}[i][k] * {kernel}_kernel_{safe_name}(scaled, sv_{safe_name}[k]);
            }}
            
            if (decision > 0.0{type_suffix}) votes[i]++; else votes[j]++;
            clf_idx++;
        }}
    }}
    
    // Return class with most votes
    int max_votes = votes[0], predicted = 0;
    for (int i = 1; i < {n_classes}; i++) {{
        if (votes[i] > max_votes) {{
            max_votes = votes[i];
            predicted = i;
        }}
    }}
    
    static const int class_labels[{n_classes}] = {{{', '.join(map(str, model.classes_))}}};
    return class_labels[predicted];
}}
""")
    else:
        # LinearSVC - s? d?ng coef_ v? intercept_
        coef = model.coef_
        intercept = model.intercept_
        
        # T?nh k?ch thu?c (d?ng data_type size)
        bytes_per_value = 4 if USE_FLOAT_PRECISION else 8
        size_coef = coef.size * bytes_per_value
        print(f"   📊 LinearSVC size: {coef.size} coefficients × {bytes_per_value} bytes = {size_coef:,} bytes")
        
        if n_classes == 2:
            # Binary classification
            if ARDUINO_OPTIMIZE:
                code_parts.append(f"""
// Coefficients for linear model
static const {data_type} coef_{safe_name}[{n_features}]{progmem} = {{
    {', '.join(f'{x:.{precision}f}{type_suffix}' for x in coef[0])}
}};

int predict_{func_name}(const float* features) {{
    {data_type} scaled[{n_features}];
    standardize_{safe_name}(features, scaled);
    
    {data_type} decision = {intercept[0]:.{precision}f}{type_suffix};
    for (int i = 0; i < {n_features}; i++) {{
        {data_type} coef_val = read_{safe_name}(&coef_{safe_name}[i]);
        decision += coef_val * scaled[i];
    }}
    
    return decision >= 0.0{type_suffix} ? {model.classes_[1]} : {model.classes_[0]};
}}
""")
            else:
                code_parts.append(f"""
// Coefficients for linear model
static const {data_type} coef_{safe_name}[{n_features}] = {{
    {', '.join(f'{x:.{precision}f}{type_suffix}' for x in coef[0])}
}};

int predict_{func_name}(const float* features) {{
    {data_type} scaled[{n_features}];
    standardize_{safe_name}(features, scaled);
    
    {data_type} decision = {intercept[0]:.{precision}f}{type_suffix};
    for (int i = 0; i < {n_features}; i++) {{
        decision += coef_{safe_name}[i] * scaled[i];
    }}
    
    return decision >= 0.0{type_suffix} ? {model.classes_[1]} : {model.classes_[0]};
}}
""")
        else:
            # Multi-class LinearSVC
            coef_flat = coef.flatten()
            chunk_size = 10
            coef_chunks = [coef_flat[i:i+chunk_size] for i in range(0, len(coef_flat), chunk_size)]
            coef_lines_list = [', '.join(f'{x:.{precision}f}{type_suffix}' for x in chunk) for chunk in coef_chunks]
            coef_lines = ',\n    '.join(coef_lines_list)
            
            if ARDUINO_OPTIMIZE:
                code_parts.append(f"""
// Coefficients for {n_classes}-class linear model (flattened from [{n_classes}][{n_features}])
static const {data_type} coef_{safe_name}[{len(coef_flat)}]{progmem} = {{
    {coef_lines}
}};

// Intercepts for each class
static const {data_type} intercepts_{safe_name}[{n_classes}]{progmem} = {{
    {', '.join(f'{x:.{precision}f}{type_suffix}' for x in intercept)}
}};

int predict_{func_name}(const float* features) {{
    {data_type} scaled[{n_features}];
    standardize_{safe_name}(features, scaled);
    
    {data_type} scores[{n_classes}];
    for (int c = 0; c < {n_classes}; c++) {{
        scores[c] = read_{safe_name}(&intercepts_{safe_name}[c]);
        for (int i = 0; i < {n_features}; i++) {{
            {data_type} coef_val = read_{safe_name}(&coef_{safe_name}[c * {n_features} + i]);
            scores[c] += coef_val * scaled[i];
        }}
    }}
    
    int predicted = 0;
    {data_type} max_score = scores[0];
    for (int c = 1; c < {n_classes}; c++) {{
        if (scores[c] > max_score) {{
            max_score = scores[c];
            predicted = c;
        }}
    }}
    
    static const int class_labels[{n_classes}] = {{{', '.join(map(str, model.classes_))}}};
    return class_labels[predicted];
}}
""")
            else:
                # Non-optimized version (ESP32)
                coef_lines = ',\n'.join('    {' + ', '.join(f'{x:.{precision}f}{type_suffix}' for x in row) + '}' for row in coef)
                code_parts.append(f"""
// Coefficients for {n_classes}-class linear model
static const {data_type} coef_{safe_name}[{n_classes}][{n_features}] = {{
{coef_lines}
}};

// Intercepts for each class
static const {data_type} intercepts_{safe_name}[{n_classes}] = {{
    {', '.join(f'{x:.{precision}f}{type_suffix}' for x in intercept)}
}};

int predict_{func_name}(const float* features) {{
    {data_type} scaled[{n_features}];
    standardize_{safe_name}(features, scaled);
    
    {data_type} scores[{n_classes}];
    for (int c = 0; c < {n_classes}; c++) {{
        scores[c] = intercepts_{safe_name}[c];
        for (int i = 0; i < {n_features}; i++) {{
            scores[c] += coef_{safe_name}[c][i] * scaled[i];
        }}
    }}
    
    int predicted = 0;
    {data_type} max_score = scores[0];
    for (int c = 1; c < {n_classes}; c++) {{
        if (scores[c] > max_score) {{
            max_score = scores[c];
            predicted = c;
        }}
    }}
    
    static const int class_labels[{n_classes}] = {{{', '.join(map(str, model.classes_))}}};
    return class_labels[predicted];
}}
""")
    
    return ''.join(code_parts)

def generate_rff_svm_c_code(model, model_name, scaler, n_features, n_classes):
    """Generate C code for RBFSampler (RFF) + LinearSVC pipeline (Mega-friendly)."""
    safe_name = get_safe_name(model_name)
    func_name = get_func_name(model_name)

    if not hasattr(model, 'named_steps') or 'rff' not in model.named_steps or 'lin' not in model.named_steps:
        raise ValueError("RFF model must be a Pipeline with steps 'rff' and 'lin'")

    rff = model.named_steps['rff']
    lin = model.named_steps['lin']

    W = np.asarray(getattr(rff, 'random_weights_', None))
    b = np.asarray(getattr(rff, 'random_offset_', None))
    if W is None or b is None:
        raise ValueError("RBFSampler not fitted (missing random_weights_/random_offset_)")

    n_components = int(W.shape[1])
    scale = float(np.sqrt(2.0 / n_components))

    coef = np.asarray(lin.coef_)
    intercept = np.asarray(lin.intercept_)
    classes = np.asarray(lin.classes_)

    progmem = " PROGMEM" if ARDUINO_OPTIMIZE else ""

    # Flatten W for PROGMEM
    W_flat = W.astype(np.float32).flatten()
    b_flat = b.astype(np.float32).flatten()

    chunk_size = 10
    def _fmt_lines(arr):
        chunks = [arr[i:i+chunk_size] for i in range(0, len(arr), chunk_size)]
        lines_list = [', '.join(f"{float(x):.{COEFFICIENT_PRECISION}f}f" for x in ch) for ch in chunks]
        return ',\n    '.join(lines_list)

    W_lines = _fmt_lines(W_flat)
    b_lines = _fmt_lines(b_flat)

    code_parts = []
    code_parts.append(f"""
// ========================================
// {model_name} (RFF RBF -> LinearSVC, {n_components} components)
// ========================================

// Scaler parameters (match training StandardScaler)
static const float scaler_mean_{safe_name}[{n_features}]{progmem} = {{
    {', '.join(f'{float(x):.{COEFFICIENT_PRECISION}f}f' for x in getattr(scaler, 'mean_', [0]*n_features))}
}};

static const float scaler_scale_{safe_name}[{n_features}]{progmem} = {{
    {', '.join(f'{float(x):.{COEFFICIENT_PRECISION}f}f' for x in getattr(scaler, 'scale_', [1]*n_features))}
}};

// RFF parameters (W: [{n_features} x {n_components}], b: [{n_components}])
static const float rff_W_{safe_name}[{len(W_flat)}]{progmem} = {{
    {W_lines}
}};

static const float rff_b_{safe_name}[{len(b_flat)}]{progmem} = {{
    {b_lines}
}};

static inline float rff_read_f_{safe_name}(const float* p) {{
#if MLP_HAVE_PGM
    return pgm_read_float(p);
#else
    return *p;
#endif
}}

static inline void standardize_{safe_name}(const float* features, float* scaled) {{
    for (int i = 0; i < {n_features}; i++) {{
        float mean_v, scale_v;
#if MLP_HAVE_PGM
        mean_v = pgm_read_float(&scaler_mean_{safe_name}[i]);
        scale_v = pgm_read_float(&scaler_scale_{safe_name}[i]);
#else
        mean_v = scaler_mean_{safe_name}[i];
        scale_v = scaler_scale_{safe_name}[i];
#endif
        scaled[i] = (features[i] - mean_v) / scale_v;
    }}
}}

static inline void rff_transform_{safe_name}(const float* x_scaled, float* z) {{
    const float s = {scale:.{COEFFICIENT_PRECISION}f}f;
    for (int j = 0; j < {n_components}; j++) {{
        float dot = 0.0f;
        int base = j; // column-major access via flat (i*n_components + j)
        for (int i = 0; i < {n_features}; i++) {{
            float wij = rff_read_f_{safe_name}(&rff_W_{safe_name}[i * {n_components} + j]);
            dot += x_scaled[i] * wij;
        }}
        float bj = rff_read_f_{safe_name}(&rff_b_{safe_name}[j]);
        z[j] = s * cosf(dot + bj);
    }}
}}
""")

    # Emit linear classifier weights
    coef_flat = coef.astype(np.float32).flatten()
    intercept_flat = intercept.astype(np.float32).flatten()
    coef_lines = _fmt_lines(coef_flat)
    intercept_lines = _fmt_lines(intercept_flat)

    code_parts.append(f"""
// Linear classifier on RFF features
static const float lin_coef_{safe_name}[{len(coef_flat)}]{progmem} = {{
    {coef_lines}
}};

static const float lin_intercept_{safe_name}[{len(intercept_flat)}]{progmem} = {{
    {intercept_lines}
}};

int predict_{func_name}(const float* features) {{
    float x_scaled[{n_features}];
    float z[{n_components}];
    standardize_{safe_name}(features, x_scaled);
    rff_transform_{safe_name}(x_scaled, z);

    // Scores
""")

    if len(classes) == 2 and coef.shape[0] == 1:
        # Binary LinearSVC often stores one row
        code_parts.append(f"""
    float decision = 0.0f;
    for (int j = 0; j < {n_components}; j++) {{
        float wj = rff_read_f_{safe_name}(&lin_coef_{safe_name}[j]);
        decision += wj * z[j];
    }}
    decision += rff_read_f_{safe_name}(&lin_intercept_{safe_name}[0]);
    static const int class_labels[2] = {{{int(classes[0])}, {int(classes[1])}}};
    return (decision >= 0.0f) ? class_labels[1] : class_labels[0];
""")
    else:
        # Multi-class OvR
        class_labels = ', '.join(str(int(c)) for c in classes)
        code_parts.append(f"""
    float best = -1e30f;
    int best_idx = 0;
    for (int c = 0; c < {len(classes)}; c++) {{
        float score = rff_read_f_{safe_name}(&lin_intercept_{safe_name}[c]);
        int base = c * {n_components};
        for (int j = 0; j < {n_components}; j++) {{
            float w = rff_read_f_{safe_name}(&lin_coef_{safe_name}[base + j]);
            score += w * z[j];
        }}
        if (c == 0 || score > best) {{
            best = score;
            best_idx = c;
        }}
    }}
    static const int class_labels[{len(classes)}] = {{{class_labels}}};
    return class_labels[best_idx];
}}
""")

    return ''.join(code_parts)

def generate_tree_c_code(model, model_name, scaler, n_features, n_classes):
    """Tạo code C cho Decision Tree-based models"""
    safe_name = get_safe_name(model_name)
    func_name = get_func_name(model_name)
    
    code_parts = []
    code_parts.append(f"""
// ========================================
// {model_name}
// ========================================

""")
    
    # Th?m scaler parameters
    code_parts.append(f"static const float {safe_name}_scaler_mean[{n_features}] = {{\n")
    for i, val in enumerate(scaler.mean_):
        code_parts.append(f"    {val:.10f}f{',' if i < len(scaler.mean_)-1 else ''}\n")
    code_parts.append("};\n\n")
    
    code_parts.append(f"static const float {safe_name}_scaler_scale[{n_features}] = {{\n")
    for i, val in enumerate(scaler.scale_):
        code_parts.append(f"    {val:.10f}f{',' if i < len(scaler.scale_)-1 else ''}\n")
    code_parts.append("};\n\n")
    
    # Tree-based models: extract decision tree structure
    # For single tree (DecisionTreeClassifier)
    if hasattr(model, 'tree_'):
        tree = model.tree_
        n_nodes = tree.node_count
        
        # Extract tree arrays
        feature = tree.feature
        threshold = tree.threshold
        children_left = tree.children_left
        children_right = tree.children_right
        
        # For classification, value contains class counts per node
        # We need to get the predicted class for each leaf
        value = tree.value
        
        # Tree feature array (which feature to split on, -2 for leaf)
        code_parts.append(f"static const int {safe_name}_tree_feature[{n_nodes}] = {{\n    ")
        for i, f in enumerate(feature):
            code_parts.append(f"{int(f)}")
            if i < n_nodes - 1:
                code_parts.append(", ")
            if (i + 1) % 15 == 0:
                code_parts.append("\n    ")
        code_parts.append("\n};\n\n")
        
        # Tree threshold array
        code_parts.append(f"static const float {safe_name}_tree_threshold[{n_nodes}] = {{\n    ")
        for i, t in enumerate(threshold):
            code_parts.append(f"{float(t):.10f}f")
            if i < n_nodes - 1:
                code_parts.append(", ")
            if (i + 1) % 5 == 0:
                code_parts.append("\n    ")
        code_parts.append("\n};\n\n")
        
        # Children left array
        code_parts.append(f"static const int {safe_name}_tree_children_left[{n_nodes}] = {{\n    ")
        for i, c in enumerate(children_left):
            code_parts.append(f"{int(c)}")
            if i < n_nodes - 1:
                code_parts.append(", ")
            if (i + 1) % 15 == 0:
                code_parts.append("\n    ")
        code_parts.append("\n};\n\n")
        
        # Children right array
        code_parts.append(f"static const int {safe_name}_tree_children_right[{n_nodes}] = {{\n    ")
        for i, c in enumerate(children_right):
            code_parts.append(f"{int(c)}")
            if i < n_nodes - 1:
                code_parts.append(", ")
            if (i + 1) % 15 == 0:
                code_parts.append("\n    ")
        code_parts.append("\n};\n\n")
        
        # Tree value array (predicted class for each node)
        code_parts.append(f"static const int {safe_name}_tree_value[{n_nodes}] = {{\n    ")
        for i in range(n_nodes):
            # value[i] has shape (1, n_classes), get the class with max count
            predicted_class = int(value[i].argmax())
            code_parts.append(f"{predicted_class}")
            if i < n_nodes - 1:
                code_parts.append(", ")
            if (i + 1) % 15 == 0:
                code_parts.append("\n    ")
        code_parts.append("\n};\n\n")
        
        code_parts.append(f"""
static int {safe_name}_predict_tree(const float* features_scaled) {{
    // Simple tree traversal using arrays
    int node = 0;
    while ({safe_name}_tree_feature[node] != -2) {{ // -2 indicates leaf
        if (features_scaled[{safe_name}_tree_feature[node]] <= {safe_name}_tree_threshold[node]) {{
            node = {safe_name}_tree_children_left[node];
        }} else {{
            node = {safe_name}_tree_children_right[node];
        }}
    }}
    // Return class from leaf node
    return {safe_name}_tree_value[node];
}}
""")
    
    # For ensemble models (RandomForest, ExtraTrees, etc.)
    elif hasattr(model, 'estimators_'):
        n_estimators = len(model.estimators_)
        code_parts.append(f"""
// Ensemble model with {n_estimators} trees
// Note: Full implementation would be very large for ESP32
// Using simplified voting approach based on feature importance

static const float {safe_name}_feature_importance[{n_features}] = {{\n""")
        if hasattr(model, 'feature_importances_'):
            importances = model.feature_importances_
            for i, val in enumerate(importances):
                code_parts.append(f"    {val:.10f}f{',' if i < len(importances)-1 else ''}\n")
        else:
            for i in range(n_features):
                code_parts.append(f"    {1.0/n_features:.10f}f{',' if i < n_features-1 else ''}\n")
        code_parts.append("};\n\n")
        
        code_parts.append(f"""
static int {safe_name}_predict_ensemble(const float* features_scaled) {{
    // Simplified ensemble: weighted vote based on feature importance
    float scores[{n_classes}] = {{0.0f}};
    
    for (int c = 0; c < {n_classes}; c++) {{
        for (int f = 0; f < {n_features}; f++) {{
            scores[c] += features_scaled[f] * {safe_name}_feature_importance[f];
        }}
    }}
    
    // Find max score
    int predicted = 0;
    float max_score = scores[0];
    for (int c = 1; c < {n_classes}; c++) {{
        if (scores[c] > max_score) {{
            max_score = scores[c];
            predicted = c;
        }}
    }}
    
    return predicted;
}}
""")
    
    # Main predict function
    code_parts.append(f"""
int predict_{func_name}(const float* features) {{
    // Apply scaling
    float features_scaled[{n_features}];
    for (int i = 0; i < {n_features}; i++) {{
        features_scaled[i] = (features[i] - {safe_name}_scaler_mean[i]) / {safe_name}_scaler_scale[i];
    }}
    
    int predicted;
""")
    
    if hasattr(model, 'tree_'):
        code_parts.append(f"    predicted = {safe_name}_predict_tree(features_scaled);\n")
    else:
        code_parts.append(f"    predicted = {safe_name}_predict_ensemble(features_scaled);\n")
    
    code_parts.append(f"""    
    static const int class_labels[{n_classes}] = {{{', '.join(map(str, model.classes_))}}};
    return class_labels[predicted];
}}
""")
    
    return ''.join(code_parts)

def generate_mlp_c_code(model, model_name, scaler, n_features, n_classes):
    """Tạo code C cho Multi-Layer Perceptron"""
    safe_name = get_safe_name(model_name)
    func_name = get_func_name(model_name)
    
    code_parts = []
    code_parts.append(f"""
// ========================================
// {model_name}
// ========================================

""")
    
    # Scaler parameters
    code_parts.append(f"static const float {safe_name}_scaler_mean[{n_features}] = {{\n")
    for i, val in enumerate(scaler.mean_):
        code_parts.append(f"    {val:.10f}f{',' if i < len(scaler.mean_)-1 else ''}\n")
    code_parts.append("};\n\n")
    
    code_parts.append(f"static const float {safe_name}_scaler_scale[{n_features}] = {{\n")
    for i, val in enumerate(scaler.scale_):
        code_parts.append(f"    {val:.10f}f{',' if i < len(scaler.scale_)-1 else ''}\n")
    code_parts.append("};\n\n")
    
    # Activation function
    activation = getattr(model, 'activation', 'relu')  # Default to relu if not found
    code_parts.append(f"""
static inline float {safe_name}_activation(float x) {{""")
    
    if activation == 'relu':
        code_parts.append("""
    return (x > 0.0f) ? x : 0.0f;
}
""")
    elif activation == 'tanh':
        code_parts.append("""
    return tanhf(x);
}
""")
    elif activation == 'logistic':
        code_parts.append("""
    return 1.0f / (1.0f + expf(-x));
}
""")
    else:  # identity
        code_parts.append("""
    return x;
}
""")
    
    # Network weights and biases
    n_layers = len(model.coefs_)
    
    for layer_idx in range(n_layers):
        coef = model.coefs_[layer_idx]
        intercept = model.intercepts_[layer_idx]
        
        n_inputs = coef.shape[0]
        n_outputs = coef.shape[1]
        
        # Weight matrix
        code_parts.append(f"\nstatic const float {safe_name}_layer{layer_idx}_weights[{n_inputs}][{n_outputs}] = {{\n")
        for i in range(n_inputs):
            line = "    {" + ", ".join(f"{coef[i][j]:.10f}f" for j in range(n_outputs)) + "}"
            if i < n_inputs - 1:
                line += ","
            code_parts.append(line + "\n")
        code_parts.append("};\n")
        
        # Bias vector
        code_parts.append(f"\nstatic const float {safe_name}_layer{layer_idx}_bias[{n_outputs}] = {{\n")
        for i, val in enumerate(intercept):
            code_parts.append(f"    {val:.10f}f{',' if i < len(intercept)-1 else ''}\n")
        code_parts.append("};\n")
    
    # Predict function
    # Calculate max hidden layer size
    max_hidden_size = max([coef.shape[1] for coef in model.coefs_[:-1]] + [n_features])
    
    code_parts.append(f"""
int predict_{func_name}(const float* features) {{
    // Apply scaling
    float features_scaled[{n_features}];
    for (int i = 0; i < {n_features}; i++) {{
        features_scaled[i] = (features[i] - {safe_name}_scaler_mean[i]) / {safe_name}_scaler_scale[i];
    }}
    
    // Forward pass through network
    float layer_input[{max_hidden_size}];
    float layer_output[{max_hidden_size}];
    
    // Copy input
    for (int i = 0; i < {n_features}; i++) {{
        layer_input[i] = features_scaled[i];
    }}
""")
    
    # Generate forward pass for each layer
    for layer_idx in range(n_layers):
        n_inputs = model.coefs_[layer_idx].shape[0]
        n_outputs = model.coefs_[layer_idx].shape[1]
        is_output_layer = (layer_idx == n_layers - 1)
        
        code_parts.append(f"""    // Layer {layer_idx}
    for (int j = 0; j < {n_outputs}; j++) {{
        layer_output[j] = {safe_name}_layer{layer_idx}_bias[j];
        for (int i = 0; i < {n_inputs}; i++) {{
            layer_output[j] += layer_input[i] * {safe_name}_layer{layer_idx}_weights[i][j];
        }}
""")
        
        if not is_output_layer:
            code_parts.append(f"        layer_output[j] = {safe_name}_activation(layer_output[j]);\n")
        
        code_parts.append("    }\n")
        
        if not is_output_layer:
            code_parts.append(f"""    
    // Copy to next layer input
    for (int i = 0; i < {n_outputs}; i++) {{
        layer_input[i] = layer_output[i];
    }}
    
""")
    
    code_parts.append(f"""    
    // Find class with max output
    int predicted = 0;
    float max_output = layer_output[0];
    for (int i = 1; i < {n_classes}; i++) {{
        if (layer_output[i] > max_output) {{
            max_output = layer_output[i];
            predicted = i;
        }}
    }}
    
    static const int class_labels[{n_classes}] = {{{', '.join(map(str, model.classes_))}}};
    return class_labels[predicted];
}}
""")
    
    return ''.join(code_parts)

def generate_knn_c_code(model, model_name, scaler, n_features, n_classes):
    """Tạo code C cho K-Nearest Neighbors"""
    safe_name = get_safe_name(model_name)
    func_name = get_func_name(model_name)
    
    # KNN requires storing training data - can be very large
    # We'll store a subset or use a simplified approach.
    # NOTE: Mega SRAM is tiny; constants MUST be in PROGMEM when ARDUINO_OPTIMIZE=True.
    # Defensive: some sklearn versions store fit data in _fit_X; if missing, try _fit_method-related attrs.
    fit_X = getattr(model, '_fit_X', None)
    fit_y = getattr(model, '_y', None)
    if fit_X is None or fit_y is None:
        raise ValueError(f"KNN generator: missing training buffers (_fit_X/_y) for {model_name}")

    n_training_total = len(fit_X)
    # S? d?ng MAX_KNN_SAMPLES t? config
    n_training = min(n_training_total, MAX_KNN_SAMPLES)
    if n_training <= 0:
        raise ValueError(f"KNN generator: no training samples available for {model_name}")
    
    # U?c lu?ng k?ch thu?c
    knn_size = n_training * n_features * 4  # float = 4 bytes
    print(f"   ⚠️ KNN samples: {n_training_total} → {n_training} (limited by MAX_KNN_SAMPLES={MAX_KNN_SAMPLES})")
    print(f"   ⚠️ KNN data size: {knn_size/1024:.1f} KB")
    
    if knn_size > MAX_FLASH_BYTES * 0.5:
        print(f"   ⚠️ KNN uses >50% Flash! Consider reducing MAX_KNN_SAMPLES")
    
    code_parts = []
    code_parts.append(f"""
// ========================================
// {model_name}
// ========================================
// Note: Using {n_training}/{n_training_total} training samples (MAX_KNN_SAMPLES={MAX_KNN_SAMPLES})
// Estimated size: {knn_size/1024:.1f} KB

""")
    
    progmem = " PROGMEM" if ARDUINO_OPTIMIZE else ""

    # Store labels as CLASS INDEX (0..n_classes-1) to make voting correct even if labels are {1,2,3,...}
    # scikit-learn KNN may store model._y already label-encoded (0..C-1) even when classes_ are original labels.
    # So we support BOTH cases:
    # - if y value exists in classes_ -> map it
    # - else if y is in [0..C-1] -> treat it as encoded index
    class_to_index = {}
    for i, c in enumerate(model.classes_):
        try:
            class_to_index[int(c)] = i
        except Exception:
            # Non-integer labels aren't supported by Arduino int output here.
            pass

    y_idx = []
    for i in range(n_training):
        y_raw = fit_y[i]
        try:
            y_int = int(y_raw)
        except Exception:
            # Fallback: if cannot convert, use 0 to avoid crashing generation
            y_int = 0

        if y_int in class_to_index:
            y_idx.append(int(class_to_index[y_int]))
        elif 0 <= y_int < n_classes:
            # Already encoded index
            y_idx.append(int(y_int))
        else:
            # Unknown label - map to 0 as a safe fallback
            y_idx.append(0)

    # Scaler (Flash-safe when ARDUINO_OPTIMIZE)
    code_parts.append(f"static const float {safe_name}_scaler_mean[{n_features}]{progmem} = {{\n")
    scaler_mean = list(getattr(scaler, 'mean_', []))
    scaler_scale = list(getattr(scaler, 'scale_', []))
    if len(scaler_mean) != n_features or len(scaler_scale) != n_features:
        # Pad/truncate defensively to avoid generation-time crashes
        scaler_mean = (scaler_mean + [0.0] * n_features)[:n_features]
        scaler_scale = (scaler_scale + [1.0] * n_features)[:n_features]

    for i, val in enumerate(scaler_mean):
        code_parts.append(f"    {float(val):.10f}f{',' if i < len(scaler_mean)-1 else ''}\n")
    code_parts.append("};\n\n")

    code_parts.append(f"static const float {safe_name}_scaler_scale[{n_features}]{progmem} = {{\n")
    for i, val in enumerate(scaler_scale):
        code_parts.append(f"    {float(val):.10f}f{',' if i < len(scaler_scale)-1 else ''}\n")
    code_parts.append("};\n\n")

    # Training data (flattened for PROGMEM efficiency)
    train_flat = []
    for i in range(n_training):
        for j in range(n_features):
            train_flat.append(float(fit_X[i][j]))

    code_parts.append(f"// Training data: {n_training} samples × {n_features} features (flattened)\n")
    code_parts.append(f"static const float {safe_name}_training_data[{n_training * n_features}]{progmem} = {{\n")
    chunk = 10
    for start in range(0, len(train_flat), chunk):
        row = train_flat[start:start+chunk]
        code_parts.append("    " + ", ".join(f"{v:.10f}f" for v in row))
        if start + chunk < len(train_flat):
            code_parts.append(",\n")
        else:
            code_parts.append("\n")
    code_parts.append("};\n\n")

    code_parts.append(f"// Training labels as class indices (0..{n_classes-1})\n")
    code_parts.append(f"static const int16_t {safe_name}_training_labels[{n_training}]{progmem} = {{\n")
    for i in range(n_training):
        code_parts.append(f"    {int(y_idx[i])}{',' if i < n_training-1 else ''}\n")
    code_parts.append("};\n\n")
    
    k = min(int(getattr(model, 'n_neighbors', 5)), int(n_training))
    metric = getattr(model, 'metric', 'euclidean')  # Default to euclidean

    if metric == 'euclidean':
        dist_accum_line = "dist += diff * diff;"
        dist_finalize_line = "distances[i] = sqrtf(dist);"
    elif metric == 'manhattan':
        dist_accum_line = "dist += fabsf(diff);"
        dist_finalize_line = "distances[i] = dist;"
    else:
        p = int(getattr(model, 'p', 2))
        dist_accum_line = f"dist += powf(fabsf(diff), {p}.0f);"
        dist_finalize_line = f"distances[i] = powf(dist, 1.0f/{p}.0f);"

    # Predict function
    # Notes:
    # - distances[] is on stack (ok)
    # - training arrays stored in PROGMEM when MLP_HAVE_PGM
    # - read scaler params from PROGMEM too
    code_parts.append(f"""
int predict_{func_name}(const float* features) {{
    // Apply scaling
    float features_scaled[{n_features}];
    for (int i = 0; i < {n_features}; i++) {{
        float mean_v;
        float scale_v;
#if MLP_HAVE_PGM
        mean_v = pgm_read_float(&{safe_name}_scaler_mean[i]);
        scale_v = pgm_read_float(&{safe_name}_scaler_scale[i]);
#else
        mean_v = {safe_name}_scaler_mean[i];
        scale_v = {safe_name}_scaler_scale[i];
#endif
        features_scaled[i] = (features[i] - mean_v) / scale_v;
    }}
    
    // Calculate distances to all training samples
    float distances[{n_training}];
    for (int i = 0; i < {n_training}; i++) {{
        float dist = 0.0f;
        int base = i * {n_features};
        for (int j = 0; j < {n_features}; j++) {{
            float xj = features_scaled[j];
            float tj;
#if MLP_HAVE_PGM
            tj = pgm_read_float(&{safe_name}_training_data[base + j]);
#else
            tj = {safe_name}_training_data[base + j];
#endif
            float diff = xj - tj;
            {dist_accum_line}
        }}
        {dist_finalize_line}
    }}
    
    // Find k nearest neighbors using simple selection
    int k_indices[{k}];
    float k_distances[{k}];
    
    // Initialize with first k samples
    for (int i = 0; i < {k}; i++) {{
        k_indices[i] = i;
        k_distances[i] = distances[i];
    }}
    
    // Find k smallest distances
    for (int i = {k}; i < {n_training}; i++) {{
        // Find max in current k
        int max_idx = 0;
        for (int j = 1; j < {k}; j++) {{
            if (k_distances[j] > k_distances[max_idx]) {{
                max_idx = j;
            }}
        }}
        
        // Replace if current is smaller
        if (distances[i] < k_distances[max_idx]) {{
            k_indices[max_idx] = i;
            k_distances[max_idx] = distances[i];
        }}
    }}
    
    // Vote among k neighbors
    int votes[{n_classes}] = {{0}};
    for (int i = 0; i < {k}; i++) {{
        int idx = k_indices[i];
        int16_t label_idx;
#if MLP_HAVE_PGM
        label_idx = (int16_t)pgm_read_word(&{safe_name}_training_labels[idx]);
#else
        label_idx = {safe_name}_training_labels[idx];
#endif
        if (label_idx >= 0 && label_idx < {n_classes}) votes[label_idx]++;
    }}
    
    // Find class with most votes
    int predicted = 0;
    int max_votes = votes[0];
    for (int c = 1; c < {n_classes}; c++) {{
        if (votes[c] > max_votes) {{
            max_votes = votes[c];
            predicted = c;
        }}
    }}
    
    static const int class_labels[{n_classes}] = {{{', '.join(map(str, model.classes_))}}};
    return class_labels[predicted];
}}
""")
    
    return ''.join(code_parts)

def generate_discriminant_c_code(model, model_name, scaler, n_features, n_classes):
    """Tạo code C cho LDA/QDA"""
    safe_name = get_safe_name(model_name)
    func_name = get_func_name(model_name)
    is_lda = 'Linear' in model_name or 'LDA' in model_name
    
    code_parts = []
    code_parts.append(f"""
// ========================================
// {model_name}
// ========================================

""")
    
    # Scaler
    code_parts.append(f"static const float {safe_name}_scaler_mean[{n_features}] = {{\n")
    for i, val in enumerate(scaler.mean_):
        code_parts.append(f"    {val:.10f}f{',' if i < len(scaler.mean_)-1 else ''}\n")
    code_parts.append("};\n\n")
    
    code_parts.append(f"static const float {safe_name}_scaler_scale[{n_features}] = {{\n")
    for i, val in enumerate(scaler.scale_):
        code_parts.append(f"    {val:.10f}f{',' if i < len(scaler.scale_)-1 else ''}\n")
    code_parts.append("};\n\n")
    
    # Class means
    code_parts.append(f"static const float {safe_name}_means[{n_classes}][{n_features}] = {{\n")
    for i in range(n_classes):
        line = "    {" + ", ".join(f"{model.means_[i][j]:.10f}f" for j in range(n_features)) + "}"
        if i < n_classes - 1:
            line += ","
        code_parts.append(line + "\n")
    code_parts.append("};\n\n")
    
    # Priors
    code_parts.append(f"static const float {safe_name}_priors[{n_classes}] = {{\n")
    for i, val in enumerate(model.priors_):
        code_parts.append(f"    {val:.10f}f{',' if i < len(model.priors_)-1 else ''}\n")
    code_parts.append("};\n\n")
    
    # Predict function - simplified using Mahalanobis distance
    code_parts.append(f"""
int predict_{func_name}(const float* features) {{
    // Apply scaling
    float features_scaled[{n_features}];
    for (int i = 0; i < {n_features}; i++) {{
        features_scaled[i] = (features[i] - {safe_name}_scaler_mean[i]) / {safe_name}_scaler_scale[i];
    }}
    
    // Calculate discriminant scores for each class
    float scores[{n_classes}];
    
    for (int c = 0; c < {n_classes}; c++) {{
        // Simple Euclidean distance to class mean
        float dist = 0.0f;
        for (int i = 0; i < {n_features}; i++) {{
            float diff = features_scaled[i] - {safe_name}_means[c][i];
            dist += diff * diff;
        }}
        
        // Discriminant score = -distance + log(prior)
        scores[c] = -dist + logf({safe_name}_priors[c]);
    }}
    
    // Find class with highest score
    int predicted = 0;
    float max_score = scores[0];
    for (int c = 1; c < {n_classes}; c++) {{
        if (scores[c] > max_score) {{
            max_score = scores[c];
            predicted = c;
        }}
    }}
    
    static const int class_labels[{n_classes}] = {{{', '.join(map(str, model.classes_))}}};
    return class_labels[predicted];
}}
""")
    
    return ''.join(code_parts)

def generate_naive_bayes_c_code(model, model_name, scaler, n_features, n_classes, debug_log_path=None):
    """Tạo code C cho Naive Bayes"""
    safe_name = get_safe_name(model_name)
    func_name = get_func_name(model_name)
    is_gaussian = 'Gaussian' in model_name
    
    def log(msg):
        if debug_log_path:
            with open(debug_log_path, 'a', encoding='utf-8') as f:
                f.write(msg + '\n')
    
    code_parts = []
    code_parts.append(f"""
// ========================================
// {model_name}
// ========================================

""")
    
    # Scaler
    code_parts.append(f"static const float {safe_name}_scaler_mean[{n_features}] = {{\n")
    for i, val in enumerate(scaler.mean_):
        code_parts.append(f"    {val:.10f}f{',' if i < len(scaler.mean_)-1 else ''}\n")
    code_parts.append("};\n\n")
    
    code_parts.append(f"static const float {safe_name}_scaler_scale[{n_features}] = {{\n")
    for i, val in enumerate(scaler.scale_):
        code_parts.append(f"    {val:.10f}f{',' if i < len(scaler.scale_)-1 else ''}\n")
    code_parts.append("};\n\n")
    
    # Class priors - GaussianNB c? class_prior_, BernoulliNB c? class_log_prior_
    class_log_prior = getattr(model, 'class_log_prior_', None)
    if class_log_prior is None:
        # GaussianNB: t?nh log t? class_prior_
        class_prior = model.class_prior_
        class_log_prior = np.log(class_prior)
        log(f"GaussianNB: Computing log prior from class_prior_")
    
    code_parts.append(f"static const float {safe_name}_class_log_prior[{n_classes}] = {{\n")
    for i, val in enumerate(class_log_prior):
        code_parts.append(f"    {val:.10f}f{',' if i < len(class_log_prior)-1 else ''}\n")
    code_parts.append("};\n\n")
    
    if is_gaussian:
        # Gaussian: store theta (mean) and var (variance)
        # Note: GaussianNB stores means in theta_ and variances in var_ or sigma_
        log(f"DEBUG GaussianNB - model_name: {model_name}")
        log(f"Has theta_: {hasattr(model, 'theta_')}")
        log(f"Has var_: {hasattr(model, 'var_')}")
        log(f"Has sigma_: {hasattr(model, 'sigma_')}")
        log(f"Has epsilon_: {hasattr(model, 'epsilon_')}")
        
        # Li?t k? T?T C? attributes c? trailing underscore (fitted attributes)
        fitted_attrs = [attr for attr in dir(model) if attr.endswith('_') and not attr.startswith('_')]
        log(f"All fitted attributes: {fitted_attrs}")
        
        theta = getattr(model, 'theta_', None)
        var = getattr(model, 'var_', getattr(model, 'sigma_', None))
        
        if theta is None:
            raise AttributeError(f"GaussianNB model '{model_name}' missing theta_ attribute. Available: {fitted_attrs}")
        if var is None:
            # Try epsilon_ as fallback (some sklearn versions)
            var = getattr(model, 'epsilon_', None)
            if var is None:
                raise AttributeError(f"GaussianNB model '{model_name}' missing var_/sigma_/epsilon_ attribute. Available: {fitted_attrs}")
        
        log(f"theta shape: {theta.shape if theta is not None else 'None'}")
        log(f"var shape: {var.shape if var is not None else 'None'}")
        
        code_parts.append(f"static const float {safe_name}_theta[{n_classes}][{n_features}] = {{\n")
        for i in range(n_classes):
            line = "    {" + ", ".join(f"{theta[i][j]:.10f}f" for j in range(n_features)) + "}"
            if i < n_classes - 1:
                line += ","
            code_parts.append(line + "\n")
        code_parts.append("};\n\n")
        
        code_parts.append(f"static const float {safe_name}_var[{n_classes}][{n_features}] = {{\n")
        for i in range(n_classes):
            line = "    {" + ", ".join(f"{var[i][j]:.10f}f" for j in range(n_features)) + "}"
            if i < n_classes - 1:
                line += ","
            code_parts.append(line + "\n")
        code_parts.append("};\n\n")
        
        code_parts.append(f"""
static inline float {safe_name}_gaussian_log_prob(float x, float mean, float var) {{
    float diff = x - mean;
    return -0.5f * logf(2.0f * 3.14159265359f * var) - (diff * diff) / (2.0f * var);
}}
""")
    else:
        # Bernoulli: store theta (feature log probabilities)
        code_parts.append(f"static const float {safe_name}_theta[{n_classes}][{n_features}] = {{\n")
        for i in range(n_classes):
            theta_vals = []
            for j in range(n_features):
                theta_val = model.feature_log_prob_[i][j] if hasattr(model, 'feature_log_prob_') else model.theta_[i][j]
                theta_vals.append(f"{theta_val:.10f}f")
            line = "    {" + ", ".join(theta_vals) + "}"
            if i < n_classes - 1:
                line += ","
            code_parts.append(line + "\n")
        code_parts.append("};\n\n")
    
    # Predict function
    code_parts.append(f"""
int predict_{func_name}(const float* features) {{
    // Apply scaling
    float features_scaled[{n_features}];
    for (int i = 0; i < {n_features}; i++) {{
        features_scaled[i] = (features[i] - {safe_name}_scaler_mean[i]) / {safe_name}_scaler_scale[i];
    }}
    
    // Calculate log probability for each class
    float log_probs[{n_classes}];
    
    for (int c = 0; c < {n_classes}; c++) {{
        log_probs[c] = {safe_name}_class_log_prior[c];
        
        for (int i = 0; i < {n_features}; i++) {{
""")
    
    if is_gaussian:
        code_parts.append(f"""            log_probs[c] += {safe_name}_gaussian_log_prob(
                features_scaled[i], 
                {safe_name}_theta[c][i], 
                {safe_name}_var[c][i]
            );
""")
    else:
        # Bernoulli - use log probabilities directly
        code_parts.append(f"""            // Bernoulli: use feature log probabilities
            if (features_scaled[i] > 0.5f) {{
                log_probs[c] += {safe_name}_theta[c][i];  // log P(x_i=1|y=c)
            }} else {{
                // For P(x=0), use log1p and exp more safely
                float theta_val = {safe_name}_theta[c][i];
                if (theta_val < -10.0f) {{
                    log_probs[c] += 0.0f;  // exp(theta) ? 0, so 1-exp(theta) ? 1, log(1) = 0
                }} else {{
                    log_probs[c] += logf(1.0f - expf(theta_val) + 1e-10f);
                }}
            }}
""")
    
    code_parts.append(f"""        }}
    }}
    
    // Find class with highest log probability
    int predicted = 0;
    float max_log_prob = log_probs[0];
    for (int c = 1; c < {n_classes}; c++) {{
        if (log_probs[c] > max_log_prob) {{
            max_log_prob = log_probs[c];
            predicted = c;
        }}
    }}
    
    static const int class_labels[{n_classes}] = {{{', '.join(map(str, model.classes_))}}};
    return class_labels[predicted];
}}
""")
    
    return ''.join(code_parts)


def print_memory_report(trained_models):
    """ Memory không thể ước tính chính xác t? Python!
    Compile example trong Arduino IDE d? xem Flash/RAM th?c t?.
    """
    print("\n" + "=" * 100)
    print("TRAINED MODELS SUMMARY")
    print("=" * 100)
    print(f"Total models: {len(trained_models)}")
    print("\nModel list:")
    for i, model_name in enumerate(trained_models.keys(), 1):
        print(f"  {i}. {model_name}")
    
    print("\n" + "=" * 100)
    print("📏  TO MEASURE ACTUAL SIZE:")
    print("\n1. Open Arduino IDE")
    print("2. Open: File → Examples → MLPredictor → SerialPredict (or SensorPredict)")
    print("3. Select board ESP32/ESP32-S3/etc...")
    print("4. Compile (Verify) to see actual Flash/RAM usage in output")
    print("\n📋 Compiler output will show:")
    print("   - Sketch uses XXXXX bytes (XX%) of program storage space (Flash)")
    print("   - Global variables use XXXX bytes (XX%) of dynamic memory (RAM)")
    print("   - Execution time of each model")
    print("=" * 100 + "\n")

def create_memory_readme(lib_dir, model_c_codes, n_features, n_classes):
    """T?o README.md v?i hu?ng d?n do memory (chỉ dùng models đã generate)"""
    
    readme_content = f"""# {ARDUINO_LIB_NAME}

Auto-generated Machine Learning library for Arduino/ESP32
Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

## 📚 Library Information

- **Number of Models**: {len(model_c_codes)}
- **Number of Features**: {n_features}
- **Number of Classes**: {n_classes}

##  Memory Measurement

**Python CANNOT accurately calculate Flash/RAM usage!**

To measure real memory usage:

1. Open Arduino IDE
2. Open: `File ? Examples ? {ARDUINO_LIB_NAME} ? SerialPredict` (or SensorPredict)
3. Select your board (ESP32/ESP32-S3/etc.)
4. Click **Verify** to compile and check output for Flash/RAM:
   ```
   Sketch uses XXXXX bytes (XX%) of program storage space
   Global variables use XXXX bytes (XX%) of dynamic memory
   ```
6. Upload and open Serial Monitor (115200) for detailed report

## 📋 Models Included

"""    
    for i, model_name in enumerate(model_c_codes.keys(), 1):
        readme_content += f"{i}. {model_name}\n"
    
    readme_content += f"""

##  Target Microcontrollers

| MCU | Flash | RAM | Notes |
|-----|-------|-----|-------|
"""
    
    # Simplified MCU list
    mcu_list = [
        ("Arduino Uno", "32 KB", "2 KB", "Too small for ML"),
        ("Arduino Mega 2560", "256 KB", "8 KB", "May fit small models"),
        ("ESP8266", "4 MB", "80 KB", "Good for ML"),
        ("ESP32", "4-16 MB", "520 KB", "Excellent for ML"),
        ("ESP32-S2", "4-8 MB", "320 KB", "Good for ML"),
        ("ESP32-S3", "4-16 MB", "512 KB", "Excellent for ML"),
        ("ESP32-C3", "4-8 MB", "400 KB", "Good for ML"),
        ("Raspberry Pi Pico", "2 MB", "264 KB", "Good for ML"),
        ("STM32F4", "512 KB-2 MB", "128-256 KB", "Good for ML"),
        ("STM32H7", "1-2 MB", "512 KB-1 MB", "Excellent for ML"),
    ]
    
    for mcu_name, flash, ram, notes in mcu_list:
        readme_content += f"| {mcu_name} | {flash} | {ram} | {notes} |\n"
    
    readme_content += f"""

**Compile your sketch to check if your board has enough space!**

##  Usage Example

```cpp
#include <{ARDUINO_LIB_NAME}.h>

float features[NUM_FEATURES];

void setup() {{
    Serial.begin(115200);
    
    // Print feature information
    print_feature_info();
    
    // Set your feature values
    features[0] = value1;
    features[1] = value2;
    // ... set all {n_features} features
    
    // Make prediction (using first model as example)
"""
    
    first_model = list(model_c_codes.keys())[0]
    safe_first = first_model.replace('-', '_').replace(' ', '_')
    
    readme_content += f"""    int prediction = predict_{safe_first}(features);
    
    Serial.print("Predicted class: ");
    Serial.println(prediction);
}}

void loop() {{
    // Your code here
}}
```

## 📋 Available Models

"""
    
    for i, model_name in enumerate(model_c_codes.keys(), 1):
        func_name = get_func_name(model_name)
        readme_content += f"{i}. `predict_{func_name}()` - {model_name}\n"
    
    readme_content += f"""

##  Important Notes

1. **Feature Order**: Features must be provided in the exact order specified
2. **Data Type**: All features must be `float` (32-bit)
3. **Scaling**: StandardScaler is applied automatically inside the library
4. **Memory**: Compile your sketch to see actual Flash/RAM usage
5. **Concurrency**: Models can be called sequentially or use voting

## ? ROM Optimization

**Separate Model Files Mode** {'(ENABLED )' if SEPARATE_MODEL_FILES else '(DISABLED ❌)'}

This library generates each model as a **separate .cpp file** in the `src/` folder.

**Benefits:**
- ? **Linker only compiles models you actually use**
- ? If you call 2 models ? only those 2 are compiled into binary
- ? Significantly reduces Flash/ROM usage on small MCUs

**Example:**
```cpp
// sketch.ino - Only use 2 models
#include <MLPredictor.h>

void setup() {{
    // Only these 2 functions are called
    int pred1 = predict_SVM_rbf_ovo(features);
    int pred2 = predict_Tree_DecisionTree(features);
    
    // Other 8 models are NOT compiled ? Save ROM!
}}
```

**Result:** Binary size ? size of 2 models only (typically 20-40 KB)

**To verify:** After compiling, check Arduino IDE output:
```
Sketch uses XXXXX bytes (XX%) of program storage space
```

##  See Examples

Check the `examples/` folder for:
- `SerialPredict/` - Read data from Serial and predict
- `SensorPredict/` - Read from sensors and predict

---
*Auto-generated by ML Pipeline*
"""
    
    with open(f"{lib_dir}/README.md", 'w', encoding='utf-8') as f:
        f.write(readme_content)
    print(f"✓ Created: README.md (with memory information)")

def generate_arduino_library(trained_models, feature_names, class_labels, X_train, X_test):
    """T?o Arduino/ESP32 library"""
    print("\n" + "=" * 80)
    print("CREATE ARDUINO/ESP32 LIBRARY")
    print("=" * 80)
    
    # ⚠️ CRITICAL: Set ARDUINO_OPTIMIZE based on target device
    global ARDUINO_OPTIMIZE
    target_device = gradio_results.get('target_device', 'ESP32')
    ARDUINO_OPTIMIZE = (target_device.lower() == "arduino mega")
    print(f"📌 Target device: {target_device}")
    print(f"📌 ARDUINO_OPTIMIZE (PROGMEM): {ARDUINO_OPTIMIZE}")
    
    lib_dir = f"{OUTPUT_DIR}/{ARDUINO_LIB_NAME}"
    os.makedirs(lib_dir, exist_ok=True)
    
    n_features = len(feature_names)
    n_classes = len(class_labels)
    
    # Dictionary d? luu C code v? memory info
    model_c_codes = {}
    memory_info = {}
    
    # T?nh value ranges t? data
    value_ranges = {}
    for i, fname in enumerate(feature_names):
        value_ranges[fname] = {
            'min': float(min(X_train[:, i].min(), X_test[:, i].min())),
            'max': float(max(X_train[:, i].max(), X_test[:, i].max()))
        }
    
    # ========================================
    # BU?C 1: Generate C code cho t?t c? models
    # ========================================
    print("\n?🔧 Generating C code for models...")
    
    source_content = f"""/*
 * {ARDUINO_LIB_NAME} Implementation
 * Auto-generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
 */

#include "{ARDUINO_LIB_NAME}.h"
#include <math.h>

// ========================================
// HELPER FUNCTIONS
// ========================================

void print_feature_info() {{
    Serial.println("\\n========================================");
    Serial.println("FEATURE INFORMATION");
    Serial.println("========================================");
    Serial.print("Number of features: ");
    Serial.println(NUM_FEATURES);
    Serial.print("Number of classes: ");
    Serial.println(NUM_CLASSES);
    Serial.println("\\nFeature order and expected ranges:");
    
    for (int i = 0; i < NUM_FEATURES; i++) {{
        Serial.print("  ");
        Serial.print(i + 1);
        Serial.print(". ");
        Serial.print(FEATURE_NAMES[i]);
        Serial.print(": [");
        Serial.print(FEATURE_RANGES[i][0], 4);
        Serial.print(", ");
        Serial.print(FEATURE_RANGES[i][1], 4);
        Serial.println("]");
    }}
    
    Serial.println("\\nClass labels:");
    for (int i = 0; i < NUM_CLASSES; i++) {{
        Serial.print("  ");
        Serial.println(CLASS_LABELS[i]);
    }}
    Serial.println("========================================\\n");
}}

"""
    
    # T?o file log d? debug (ngu?i d?ng c? th? xem sau)
    debug_log_path = f"{OUTPUT_DIR}/generation_debug.log"
    with open(debug_log_path, 'w', encoding='utf-8') as debug_log:
        debug_log.write(f"=== GENERATION DEBUG LOG ===\n")
        debug_log.write(f"Time: {datetime.now()}\n")
        debug_log.write(f"Models to generate: {list(trained_models.keys())}\n\n")
    
    for model_name, model_data in trained_models.items():
        try:
            model = model_data['model']
            scaler = model_data['scaler']
            family = model_data.get('family', 'SVM')
            
            print(f"  🔧 Generating code for {model_name} ({family})...")
            
            # Log v?o file
            with open(debug_log_path, 'a', encoding='utf-8') as debug_log:
                debug_log.write(f"\n{'='*60}\n")
                debug_log.write(f"Processing: {model_name}\n")
                debug_log.write(f"Family: {family}\n")
                debug_log.write(f"Type: {type(model).__name__}\n")
            
            # Ch?n generator ph? h?p v?i lo?i model
            if family == 'SVM':
                c_code = generate_svm_c_code(model, model_name, scaler, n_features, n_classes)
            elif family == 'SVM_RFF':
                c_code = generate_rff_svm_c_code(model, model_name, scaler, n_features, n_classes)
            elif family == 'Tree':
                c_code = generate_tree_c_code(model, model_name, scaler, n_features, n_classes)
            elif family == 'Neural Network':
                c_code = generate_mlp_c_code(model, model_name, scaler, n_features, n_classes)
            elif family == 'KNN':
                c_code = generate_knn_c_code(model, model_name, scaler, n_features, n_classes)
            elif family == 'Discriminant':
                c_code = generate_discriminant_c_code(model, model_name, scaler, n_features, n_classes)
            elif family == 'Naive Bayes':
                with open(debug_log_path, 'a', encoding='utf-8') as debug_log:
                    debug_log.write(f"Calling generate_naive_bayes_c_code...\n")
                c_code = generate_naive_bayes_c_code(model, model_name, scaler, n_features, n_classes, debug_log_path)
                with open(debug_log_path, 'a', encoding='utf-8') as debug_log:
                    debug_log.write(f"✅ Code generated successfully, length: {len(c_code)}\n")
            else:
                print(f"  Unknown family '{family}' for {model_name}, SKIPPING (no code generated)...")
                continue  # B? qua ho?n to?n - kh?ng th?m v?o model_c_codes v? memory_info
            
            # Luu C code
            model_c_codes[model_name] = c_code
            source_content += c_code + "\n"
            
            #  Không tính memory t? Python - không chính xác!
            # Compile sketch d? do th?c t?
            memory_info[model_name] = {
                'rom_bytes': 0,  # Xem trong compile output
                'ram_bytes': 0,  # Xem trong compile output
                'rom_kb': 0,
                'ram_kb': 0,
                'details': 'Compile sketch to measure'
            }
            
        except Exception as e:
            error_msg = f"❌  Lỗi khi tạo code cho {model_name}: {e}\n"
            error_msg += f"     ? Model family: {family}\n"
            error_msg += f"     ? Model type: {type(model).__name__}\n"
            print(error_msg)
            
            # Log chi ti?t v?o file
            import traceback
            with open(debug_log_path, 'a', encoding='utf-8') as debug_log:
                debug_log.write(f"\n Error:\n")
                debug_log.write(error_msg)
                debug_log.write(f"\nFull traceback:\n")
                debug_log.write(traceback.format_exc())
                debug_log.write(f"\n{'='*60}\n")
            
            traceback.print_exc()
            # SKIP ho?n to?n khi l?i - kh?ng th?m v?o model_c_codes hay memory_info
            continue
    
    # Ki?m tra s? lu?ng models th?c t? vs d? ch?n
    num_selected = len(trained_models)
    num_generated = len(model_c_codes)
    
    if num_generated < num_selected:
        print(f"\n⚠️  Warning: Selected {num_selected} models but only generated {num_generated} models!")
        print(f"   ⚠️ {num_selected - num_generated} model(s) skipped due to code generation error")
        skipped_models = set(trained_models.keys()) - set(model_c_codes.keys())
        for m in skipped_models:
            print(f"      • {m}")
    
    print(f"\n✅ Total: {num_generated}/{num_selected} models generated successfully")
    
    # ========================================
    # BU?C 2: T?o header file (SAU KHI d? c? model_c_codes)
    # ========================================
    print("\n Creating header file...")
    
    feature_ranges = '\n'.join([f"//   {i+1}. {fname}: [{value_ranges[fname]['min']:.4f}, {value_ranges[fname]['max']:.4f}]" 
                                for i, fname in enumerate(feature_names)])
    
    # On AVR (e.g., Arduino Mega), even "const" tables can still consume SRAM unless placed in PROGMEM.
    # The following header generation uses PROGMEM-safe accessors so Mega doesn't overflow RAM.
    def _c_escape_string(s: str) -> str:
        # Escape for a C string literal
        return (
            str(s)
            .replace('\\\\', r'\\\\')
            .replace('"', r'\\"')
            .replace('\n', r'\\n')
            .replace('\r', r'\\r')
            .replace('\t', r'\\t')
        )

    header_content = f"""/*
 * {ARDUINO_LIB_NAME} - Machine Learning Predictor Library
 * Auto-generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
 * 
 * This library contains ML models for prediction on Arduino/ESP32
 * 
 * IMPORTANT: DATA FORMAT REQUIREMENTS
 * ===================================
 * Number of features: {n_features}
 * Number of classes: {n_classes}
 * Classes: {', '.join(map(str, class_labels))}
 * 
 * Feature names (IN THIS EXACT ORDER):
{chr(10).join([f' *   {i+1}. {fname}' for i, fname in enumerate(feature_names)])}
 * 
 * Expected value ranges (from training data):
{feature_ranges}
 * 
 * Data type: float (32-bit floating point)
 * Scaling: Automatic (StandardScaler applied internally)
 * 
 * INPUT FORMAT:
 * - Array of {n_features} float values
 * - Must be in the EXACT order listed above
 * - Supports decimal point (.) and negative numbers
 * - Example: float features[{n_features}] = {{value1, value2, ..., value{n_features}}};
 * 
 * WARNING: Incorrect order or data type will produce incorrect predictions!
 */

#ifndef {ARDUINO_LIB_NAME.upper()}_H
#define {ARDUINO_LIB_NAME.upper()}_H

#include <Arduino.h>
#include <math.h>
#include <string.h>

// PROGMEM helpers (mainly for AVR / Arduino Mega)
#if defined(ARDUINO_ARCH_AVR)
    #include <avr/pgmspace.h>
    #define MLP_HAVE_PGM 1
#else
    #define MLP_HAVE_PGM 0
    #ifndef PROGMEM
        #define PROGMEM
    #endif
#endif

// Model count
#define NUM_MODELS {len(model_c_codes)}
#define NUM_FEATURES {n_features}
#define NUM_CLASSES {n_classes}

// =====================
// METADATA (Flash-safe)
// =====================

// NOTE:
// - On AVR (Mega), store these tables in Flash (PROGMEM) to save SRAM.
// - Use the helper accessors below instead of direct array indexing.

// Class labels (numeric)
static const int16_t CLASS_LABELS[NUM_CLASSES]{' PROGMEM' if ARDUINO_OPTIMIZE else ''} = {{{', '.join(map(str, class_labels))}}};

// Class names (string labels from your data)
// NOTE: Keep names in RAM (even on Arduino Mega) to avoid PROGMEM string pointer issues during Serial.print
{chr(10).join([f'static const char CLASS_NAME_{i}[] = "{_c_escape_string(str(label))}";' for i, label in enumerate(class_labels)])}
static const char* const CLASS_NAMES[NUM_CLASSES] = {{{', '.join([f'CLASS_NAME_{i}' for i in range(len(class_labels))])}}};  

// Feature names
{chr(10).join([f'static const char FEATURE_NAME_{i}[]{" PROGMEM" if ARDUINO_OPTIMIZE else ""} = "{_c_escape_string(fname)}";' for i, fname in enumerate(feature_names)])}
static const char* const FEATURE_NAMES[NUM_FEATURES]{' PROGMEM' if ARDUINO_OPTIMIZE else ''} = {{{', '.join([f'FEATURE_NAME_{i}' for i in range(len(feature_names))])}}};

// Expected value ranges (min, max) for each feature
static const float FEATURE_RANGES[NUM_FEATURES][2]{' PROGMEM' if ARDUINO_OPTIMIZE else ''} = {{
{chr(10).join([f'  {{{value_ranges[fname]["min"]:.6f}f, {value_ranges[fname]["max"]:.6f}f}}, // {fname}' for fname in feature_names])}
}};

// Helper function to get class name from prediction result
static inline const char* getClassName(int16_t classLabel) {{
    for (int i = 0; i < NUM_CLASSES; i++) {{
{f"        if (pgm_read_word(&CLASS_LABELS[i]) == classLabel) return CLASS_NAMES[i];" if ARDUINO_OPTIMIZE else "        if (CLASS_LABELS[i] == classLabel) return CLASS_NAMES[i];"}
    }}
    return "Unknown";
}}

// --------- Accessors (work on AVR + non-AVR) ---------
static inline int16_t mlp_get_class_label(int idx) {{
#if MLP_HAVE_PGM
    return (int16_t)pgm_read_word(&CLASS_LABELS[idx]);
#else
    return CLASS_LABELS[idx];
#endif
}}

static inline float mlp_get_feature_range_min(int idx) {{
#if MLP_HAVE_PGM
    return pgm_read_float(&FEATURE_RANGES[idx][0]);
#else
    return FEATURE_RANGES[idx][0];
#endif
}}

static inline float mlp_get_feature_range_max(int idx) {{
#if MLP_HAVE_PGM
    return pgm_read_float(&FEATURE_RANGES[idx][1]);
#else
    return FEATURE_RANGES[idx][1];
#endif
}}

static inline void mlp_get_feature_name(int idx, char* out, size_t outSize) {{
    if (!out || outSize == 0) return;
#if MLP_HAVE_PGM
    PGM_P p = (PGM_P)pgm_read_ptr(&FEATURE_NAMES[idx]);
    strncpy_P(out, p, outSize);
    out[outSize - 1] = '\0';
#else
    strncpy(out, FEATURE_NAMES[idx], outSize);
    out[outSize - 1] = '\0';
#endif
}}

// Helper function to print feature info
void print_feature_info();

// ========================================
// INPUT NORMALIZATION HELPERS
// ========================================

// Convert numeric arrays to float features (matches training pipeline)
template <typename T>
static inline void mlp_cast_features_to_float(const T* in, float* out, int n) {{
    if (!out) return;
    for (int i = 0; i < n; i++) out[i] = in ? (float)in[i] : 0.0f;
}}

// Parse a pasted line (Excel/CSV) into float features.
// Optimized version - avoids String operations for speed
static inline bool mlp_parse_features(String input, float* out, int expectedCount) {{
    if (!out || expectedCount <= 0) return false;
    
    int len = input.length();
    if (len == 0) return false;
    
    // Fast pre-scan: detect delimiter type
    bool hasTab = false;
    bool hasSemicolon = false;
    for (int i = 0; i < len && !(hasTab && hasSemicolon); i++) {{
        char c = input.charAt(i);
        if (c == '\\t') hasTab = true;
        else if (c == ';') hasSemicolon = true;
    }}
    
    int count = 0;
    int startIdx = 0;
    bool inToken = false;
    char buffer[32];  // Buffer for number parsing
    int bufIdx = 0;
    
    for (int i = 0; i <= len; i++) {{
        char c = (i < len) ? input.charAt(i) : '\\0';
        
        // Decimal comma handling: if TAB/semicolon present, convert comma to dot
        if (c == ',' && (hasTab || hasSemicolon)) {{
            c = '.';
        }}
        
        bool isDelimiter = (i == len) || (c == '\\t') || (c == ' ') || (c == ';') || (c == ',');
        
        if (!isDelimiter && bufIdx < 31) {{
            // Accumulate non-delimiter characters
            buffer[bufIdx++] = c;
            inToken = true;
        }} else if (inToken) {{
            // End of token - parse it
            buffer[bufIdx] = '\\0';
            
            // Skip whitespace-only tokens
            bool hasDigit = false;
            for (int j = 0; j < bufIdx; j++) {{
                if (isdigit(buffer[j]) || buffer[j] == '-' || buffer[j] == '.') {{
                    hasDigit = true;
                    break;
                }}
            }}
            
            if (hasDigit && count < expectedCount) {{
                out[count++] = atof(buffer);
            }}
            
            bufIdx = 0;
            inToken = false;
        }}
    }}
    
    return count == expectedCount;
}}

"""
    
    # T?o function declarations v?i memory info
    header_content += "\n// ========================================\n"
    header_content += "// MODEL FUNCTIONS\n"
    header_content += "// ========================================\n"
    
    for model_name in model_c_codes.keys():
        # Ki?m tra xem model c? trong memory_info kh?ng (defensive programming)
        if model_name not in memory_info:
            print(f"⚠️  Warning: {model_name} has no memory info, skipping")
            continue
        
        safe_name = get_safe_name(model_name)
        func_name = get_func_name(model_name)
        mem = memory_info[model_name]
        header_content += f"\n// {model_name}\n"
        header_content += f"// {mem['details']}\n"
        header_content += f"int predict_{func_name}(const float* features);\n"
        header_content += f"static inline int predict_{func_name}(const double* features) {{\n"
        header_content += f"    float tmp[NUM_FEATURES];\n"
        header_content += f"    mlp_cast_features_to_float(features, tmp, NUM_FEATURES);\n"
        header_content += f"    return predict_{func_name}(tmp);\n"
        header_content += f"}}\n"
        header_content += f"static inline int predict_{func_name}(const int32_t* features) {{\n"
        header_content += f"    float tmp[NUM_FEATURES];\n"
        header_content += f"    mlp_cast_features_to_float(features, tmp, NUM_FEATURES);\n"
        header_content += f"    return predict_{func_name}(tmp);\n"
        header_content += f"}}\n"
        header_content += f"static inline int predict_{func_name}(const String& input) {{\n"
        header_content += f"    float tmp[NUM_FEATURES];\n"
        header_content += f"    if (!mlp_parse_features(input, tmp, NUM_FEATURES)) {{\n"
        header_content += f"        // Parse failed: return first class label (and print hint)\n"
        header_content += f"        Serial.println(String(\"Parse error: expected \" ) + String(NUM_FEATURES) + String(\" values\"));\n"
        header_content += f"    }}\n"
        header_content += f"    return predict_{func_name}(tmp);\n"
        header_content += f"}}\n"
    
    header_content += "\n#endif\n"
    
    # Luu header
    with open(f"{lib_dir}/{ARDUINO_LIB_NAME}.h", 'w') as f:
        f.write(header_content)
    print(f"✓ Created: {ARDUINO_LIB_NAME}.h")
    
    # ========================================
    # BU?C 3: Luu source files
    # ========================================
    if SEPARATE_MODEL_FILES:
        # MODE 1: M?i model 1 file ri?ng (KHUY?N KH?CH - t?i uu ROM)
        print(f"\n Separate files mode: Creating {len(model_c_codes)} individual .cpp files")
        src_dir = f"{lib_dir}/src"
        os.makedirs(src_dir, exist_ok=True)
        
        # T?o main .cpp ch? ch?a helper functions
        main_content = f"""/*
 * {ARDUINO_LIB_NAME} Implementation
 * Auto-generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
 */

#include "{ARDUINO_LIB_NAME}.h"
#include <math.h>

// ========================================
// HELPER FUNCTIONS
// ========================================

void print_feature_info() {{
    Serial.println("\\n========================================");
    Serial.println("FEATURE INFORMATION");
    Serial.println("========================================");
    Serial.print("Number of features: ");
    Serial.println(NUM_FEATURES);
    Serial.print("Number of classes: ");
    Serial.println(NUM_CLASSES);
    
    // Only show first 5 features to avoid slow startup
    Serial.println("\\nFirst 5 features:");
    int displayCount = min(5, NUM_FEATURES);
    for (int i = 0; i < displayCount; i++) {{
        char nameBuf[48];
        mlp_get_feature_name(i, nameBuf, sizeof(nameBuf));
        Serial.print("  ");
        Serial.print(i + 1);
        Serial.print(". ");
        Serial.println(nameBuf);
    }}
    if (NUM_FEATURES > 5) {{
        Serial.print("  ... and ");
        Serial.print(NUM_FEATURES - 5);
        Serial.println(" more features");
    }}
    
    Serial.println("\\nClass labels:");
    for (int i = 0; i < NUM_CLASSES; i++) {{
        Serial.print("  ");
        Serial.println(mlp_get_class_label(i));
    }}
    Serial.println("========================================\\n");
}}
"""
        with open(f"{lib_dir}/{ARDUINO_LIB_NAME}.cpp", 'w') as f:
            f.write(main_content)
        print(f"✓ Created: {ARDUINO_LIB_NAME}.cpp (helper functions)")
        
        # T?o file ri?ng cho t?ng model
        for model_name, c_code in model_c_codes.items():
            # Ki?m tra xem model c? trong memory_info kh?ng
            if model_name not in memory_info:
                print(f"⚠️  Warning: {model_name} has no memory info, using default values")
                memory_info[model_name] = {
                    'rom_bytes': 5000,
                    'ram_bytes': 500,
                    'rom_kb': 5.0,
                    'ram_kb': 0.5,
                    'details': 'Unknown'
                }
            
            safe_name = model_name.replace('-', '_').replace(' ', '_')
            model_file_content = f"""/*
 * {model_name} Implementation
 * Auto-generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
 * 
 * ROM Usage: ~{memory_info[model_name]['rom_kb']:.2f} KB
 * RAM Usage: ~{memory_info[model_name]['ram_kb']:.2f} KB
 */

#include "{ARDUINO_LIB_NAME}.h"
#include <math.h>

{c_code}
"""
            with open(f"{src_dir}/{safe_name}.cpp", 'w') as f:
                f.write(model_file_content)
            print(f"  ✓ {safe_name}.cpp ({memory_info[model_name]['rom_kb']:.1f} KB)")
            
        print(f"\n OPTIMIZATION: Linker will only compile models called in sketch!")
        print(f"   Example: If only 2/10 models called → Binary size only increases ~{sum([memory_info[m]['rom_bytes'] for m in list(model_c_codes.keys())[:2]])/1024:.1f} KB")
        
    else:
        # MODE 2: All in 1 file (simple but not optimized)
        print(f"\n Single file mode: All models in 1 .cpp file")
        with open(f"{lib_dir}/{ARDUINO_LIB_NAME}.cpp", 'w') as f:
            f.write(source_content)
        print(f"✓ Created: {ARDUINO_LIB_NAME}.cpp")
        print(f"⚠️  Warning: Compiler may compile ALL {len(model_c_codes)} models even if only using 1-2")
        print(f"   💡 Set SEPARATE_MODEL_FILES = True to optimize ROM usage")
    
    # T?o README.md v?i memory information
    create_memory_readme(lib_dir, model_c_codes, n_features, n_classes)
    
    # In memory report
    print_memory_report(trained_models)
    
    # T?o library.properties
    properties_content = f"""name={ARDUINO_LIB_NAME}
version=1.0.0
author=Auto-generated
maintainer=Auto-generated
sentence=Machine Learning prediction library for Arduino/ESP32
paragraph=SVM models for classification
category=Data Processing
url=
architectures=*
includes={ARDUINO_LIB_NAME}.h
"""
    
    with open(f"{lib_dir}/library.properties", 'w') as f:
        f.write(properties_content)
    print(f"✓ Created: library.properties")
    
    # T?o examples (chỉ dùng models đã generate th?nh c?ng)
    create_arduino_examples(lib_dir, model_c_codes, feature_names, n_features, n_classes)
    
    return lib_dir

def create_arduino_examples(lib_dir, model_c_codes, feature_names, n_features, n_classes):
    """Tạo file examples cho Arduino (chỉ dùng models đã generate C code)"""
    examples_dir = f"{lib_dir}/examples"
    os.makedirs(examples_dir, exist_ok=True)
    
    # Example 1: Serial Input/Output
    example1_dir = f"{examples_dir}/SerialPredict"
    os.makedirs(example1_dir, exist_ok=True)
    
    model_list = '\n   '.join([f"// {i+1}. {name}" for i, name in enumerate(model_c_codes.keys())])
    first_model = list(model_c_codes.keys())[0]
    safe_first = first_model.replace('-', '_').replace(' ', '_')
    
    example1_content = f"""/*
 * Serial Prediction Example
 *
 * Enter data via Serial Monitor and view the prediction result.
 *
 * INPUT FORMAT OPTIONS:
 * 1. Space-separated: 1.5 2.3 4.1 (use dot for decimal)
 * 2. Tab-separated from Excel: 1,5    2,3    4,1 (comma will be converted to dot)
 * 3. Comma-separated: 1.5,2.3,4.1
 *
 * Available models:
     {model_list}
 */

#include <{ARDUINO_LIB_NAME}.h>

float features[NUM_FEATURES];

void setup() {{
    Serial.begin(115200);
    Serial.setTimeout(100);  // Reduce timeout from 1000ms to 100ms for faster response
    while (!Serial) delay(10);
    
    Serial.println("========================================");
    Serial.println("ML Predictor - Serial Mode");
    Serial.println("========================================");
    
    // Skip feature info on startup to save time
    Serial.print("Ready! ");
    Serial.print(NUM_FEATURES);
    Serial.println(" features expected.");
    Serial.println("Type 'info' for feature details.");
    
    Serial.println("\\nEnter " + String(NUM_FEATURES) + " values:");
    Serial.println("Format: space/tab/comma separated");
    Serial.println("Example: 1.5 2.3 4.1 -0.8 ... (or paste from Excel)\\n");
}}

void loop() {{
    if (Serial.available() > 0) {{
        unsigned long startTime = millis();
        
        // Read the input string.
        String input = Serial.readStringUntil('\\n');
        input.trim();
        
        Serial.print("Read done (");
        Serial.print(millis() - startTime);
        Serial.println("ms)");
        
        // Check for the "info" command.
        if (input.equalsIgnoreCase("info")) {{
            print_feature_info();
            return;
        }}
        
        // Check input length to avoid buffer overflow
        if (input.length() > 512) {{
            Serial.println("Error: Input too long! Max 512 characters.");
            return;
        }}
        
        bool ok = mlp_parse_features(input, features, NUM_FEATURES);
        
        Serial.print("Parse done (");
        Serial.print(millis() - startTime);
        Serial.println("ms). Running predictions...");
        
        if (ok) {{
            // Display parsed input values (only first 3 to save time)
            Serial.print("\\nInput: [");
            int displayCount = min(3, NUM_FEATURES);
            for (int i = 0; i < displayCount; i++) {{
                Serial.print("F");
                Serial.print(i);
                Serial.print("=");
                Serial.print(features[i], 2);  // Reduce precision from 4 to 2 for speed
                Serial.print(", ");
            }}
            if (NUM_FEATURES > 3) {{
                Serial.print("... +");
                Serial.print(NUM_FEATURES - 3);
                Serial.print(" more");
            }}
            Serial.println("]");
            
            // Run predictions with all models.
            Serial.println("\\n--- Predictions ---");
{chr(10).join([f'            unsigned long t{i} = millis();' + chr(10) + f'            int16_t pred{i} = predict_{get_func_name(name)}(features);' + chr(10) + f'            Serial.print("{name}: ");' + chr(10) + f'            Serial.print(getClassName(pred{i}));' + chr(10) + f'            Serial.print(" (");' + chr(10) + f'            Serial.print(millis() - t{i});' + chr(10) + f'            Serial.println("ms)");' for i, name in enumerate(model_c_codes.keys())])}
            
            Serial.print("\\nTotal time: ");
            Serial.print(millis() - startTime);
            Serial.println("ms. Ready for next input.");
        }} else {{
            Serial.print("Error: Need exactly ");
            Serial.print(NUM_FEATURES);
            Serial.println(" values!");
            Serial.println("Tip: paste space/tab/comma separated values");
        }}
    }}
}}
"""
    
    with open(f"{example1_dir}/SerialPredict.ino", 'w') as f:
        f.write(example1_content)
    print(f"✓ Created: examples/SerialPredict/SerialPredict.ino")
    
    # Example 2: Sensor Input
    example2_dir = f"{examples_dir}/SensorPredict"
    os.makedirs(example2_dir, exist_ok=True)
    
    example2_content = f"""/*
 * Sensor Prediction Example
 * 
 * Đọc dữ liệu từ cảm biến và thực hiện prediction
 * Thay đổi hàm readSensors() để đọc từ cảm biến thật
 * 
 * Available models:
   {model_list}
 */

#include <{ARDUINO_LIB_NAME}.h>

float features[NUM_FEATURES];
unsigned long lastPrediction = 0;
const unsigned long PREDICTION_INTERVAL = 1000; // 1 giây

void setup() {{
    Serial.begin(115200);
    while (!Serial) delay(10);
    
    Serial.println("========================================");
    Serial.println("ML Predictor - Sensor Mode");
    Serial.println("========================================");
    
    // Initialize your sensors here.
    initSensors();
}}

void loop() {{
    unsigned long currentMillis = millis();
    
    if (currentMillis - lastPrediction >= PREDICTION_INTERVAL) {{
        lastPrediction = currentMillis;
        
        // Read data from your sensors.
        readSensors();
        
        // Show the sampled data.
        Serial.print("Sensor data: [");
        for (int i = 0; i < NUM_FEATURES; i++) {{
            Serial.print(features[i], 4);
            if (i < NUM_FEATURES - 1) Serial.print(", ");
        }}
        Serial.println("]");
        
        // Run predictions with all models.
        Serial.println("--- Predictions ---");
{chr(10).join([f'        Serial.print("{name}: ");' + chr(10) + f'        Serial.println(predict_{get_func_name(name)}(features));' for name in model_c_codes.keys()])}
        Serial.println();
    }}
}}

void initSensors() {{
    // TODO: Initialize your sensors.
    // Example: configure pins or call sensor.begin().
    
    Serial.println("✅ Sensors initialized");
}}

void readSensors() {{
    // TODO: Read data from real sensors.
    // Sample code for {n_features} features.
    
    {chr(10).join([f"    features[{i}] = analogRead(A{i}) * (5.0 / 1023.0); // Sensor {i+1}" for i in range(min(n_features, 6))])}
    
    // Or read from I2C/SPI sensors.
    // features[0] = sensor.readTemperature();
    // features[1] = sensor.readHumidity();
    // ...
}}
"""
    
    with open(f"{example2_dir}/SensorPredict.ino", 'w') as f:
        f.write(example2_content)
    print(f"✓ Created: examples/SensorPredict/SensorPredict.ino")
    
    print(f"\\n✅ Created {2} example files in library")


# ============================================================================
# PACKAGE AND DOWNLOAD
# ============================================================================

def package_and_download():
    """N?n t?t c? v? t?i xu?ng"""
    print("\n" + "=" * 80)
    print("COMPRESS AND DOWNLOAD")
    print("=" * 80)
    
    # Tạo file RAR/ZIP
    archive_name = f"{OUTPUT_DIR}_Results"
    
    try:
        # Sử dụng shutil để tạo zip (RAR yêu cầu WinRAR)
        shutil.make_archive(archive_name, 'zip', OUTPUT_DIR)
        print(f"📦 Compressed to: {archive_name}.zip")
        
        # Tải xuống nếu đang ở Google Colab
        if IN_COLAB:
            files.download(f"{archive_name}.zip")
            print("⬇️  Downloading file...")
        else:
            print(f"✅ File ready at: {archive_name}.zip")
            
    except Exception as e:
        print(f" Error compressing: {e}")

# ============================================================================
# GRADIO WEB INTERFACE
# ============================================================================

# Global variables cho Gradio
gradio_data = {}
gradio_results = {}

def gradio_process_single_file(file, train_ratio, random_state, progress=None):
    """Process unsplit data file for Gradio"""
    try:
        if progress is None:
            if GRADIO_AVAILABLE and gr is not None:
                progress = gr.Progress()
            else:
                progress = lambda *args, **kwargs: None
        progress(0.1, desc=" Loading file...")
        global SINGLE_FILE_PATH, TRAIN_FILE_PATH, TEST_FILE_PATH, TRAIN_RATIO, RANDOM_STATE, gradio_data
        
        # Delete old OUTPUT_DIR if exists (to start fresh)
        if os.path.exists(OUTPUT_DIR):
            import shutil
            try:
                shutil.rmtree(OUTPUT_DIR)
                print(f"🗑️  Deleted old folder: {OUTPUT_DIR}")
            except Exception as e:
                print(f" Cannot delete old folder: {e}")
        
        # Save temp file
        progress(0.2, desc=" Saving temporary file...")
        SINGLE_FILE_PATH = file.name
        TRAIN_FILE_PATH = ""
        TEST_FILE_PATH = ""
        TRAIN_RATIO = train_ratio
        RANDOM_STATE = int(random_state)
        
        # Load data
        progress(0.4, desc=" Loading and analyzing data...")
        X_train, X_test, y_train, y_test, feature_names = load_data()
        
        # Save to global
        gradio_data['X_train'] = X_train
        gradio_data['X_test'] = X_test
        gradio_data['y_train'] = y_train
        gradio_data['y_test'] = y_test
        gradio_data['feature_names'] = feature_names
        
        # T?o th?ng b?o
        class_counts = pd.Series(y_train).value_counts().sort_index()
        class_names = sorted(np.unique(y_train))
        
        info = f"""✅ **File uploaded successfully!**

📊 **Data information:**
- Total samples: {len(X_train) + len(X_test)}
- Number of features: {len(feature_names)}
- Feature names: {', '.join(feature_names)}
- Number of classes: {len(class_names)}
- Class names: {', '.join(map(str, class_names))}

📈 **Class distribution:**
{chr(10).join([f"- Class {cls}: {count} samples (train)" for cls, count in class_counts.items()])}

🔀 **Data split using SPXY algorithm:**
- Train: {len(X_train)} samples ({train_ratio*100:.0f}%)
- Test: {len(X_test)} samples ({(1-train_ratio)*100:.0f}%)

---

### ✅ **STEP 1 COMPLETED! GO TO STEP 2 FOR TRAINING**
"""
        
        # V? bi?u d? v?o file
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        plot_data_distribution(X_train, X_test, y_train, y_test, feature_names)
        
        plot_path = f"{OUTPUT_DIR}/data_distribution.png"
        return info, plot_path
        
    except Exception as e:
        return f" Error: {str(e)}", None

def gradio_process_split_files(train_file, test_file, progress=None):
    """Process pre-split files for Gradio"""
    try:
        if progress is None:
            if GRADIO_AVAILABLE and gr is not None:
                progress = gr.Progress()
            else:
                progress = lambda *args, **kwargs: None
        progress(0.1, desc=" Loading files...")
        global SINGLE_FILE_PATH, TRAIN_FILE_PATH, TEST_FILE_PATH, gradio_data
        
        # Delete old OUTPUT_DIR if exists (to start fresh)
        if os.path.exists(OUTPUT_DIR):
            import shutil
            try:
                shutil.rmtree(OUTPUT_DIR)
                print(f"🗑️  Deleted old folder: {OUTPUT_DIR}")
            except Exception as e:
                print(f" Cannot delete old folder: {e}")
        
        progress(0.2, desc=" Saving temporary files...")
        SINGLE_FILE_PATH = ""
        TRAIN_FILE_PATH = train_file.name
        TEST_FILE_PATH = test_file.name
        
        # Load data
        progress(0.4, desc=" Loading and checking data...")
        X_train, X_test, y_train, y_test, feature_names = load_data()
        
        # Luu v?o global
        gradio_data['X_train'] = X_train
        gradio_data['X_test'] = X_test
        gradio_data['y_train'] = y_train
        gradio_data['y_test'] = y_test
        gradio_data['feature_names'] = feature_names
        
        # T?o th?ng b?o
        train_counts = pd.Series(y_train).value_counts().sort_index()
        test_counts = pd.Series(y_test).value_counts().sort_index()
        class_names = sorted(np.unique(np.concatenate([y_train, y_test])))
        
        info = f"""✅ **Files loaded successfully!**

📊 **Data information:**
- Train: {len(X_train)} samples
- Test: {len(X_test)} samples
- Number of features: {len(feature_names)}
- Feature names: {', '.join(feature_names)}
- Number of classes: {len(class_names)}
- Class names: {', '.join(map(str, class_names))}

📊 **Class distribution in Train:**
{chr(10).join([f"- Class {cls}: {count} samples" for cls, count in train_counts.items()])}

📊 **Class distribution in Test:**
{chr(10).join([f"- Class {cls}: {count} samples" for cls, count in test_counts.items()])}

---

### ✅ **STEP 1 COMPLETED! GO TO STEP 2 FOR TRAINING**
"""
        
        # V? bi?u d? v?o file
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        plot_data_distribution(X_train, X_test, y_train, y_test, feature_names)
        
        # Trả về cả info và path ảnh
        plot_path = f"{OUTPUT_DIR}/data_distribution.png"
        return info, plot_path
        
    except Exception as e:
        return f" Error: {str(e)}", None

def gradio_train_models(target_device, n_trials, cv_folds, optuna_timeout, output_dir, selected_groups, progress=None):
    """Train models for Gradio with progress bar"""
    try:
        if progress is None:
            if GRADIO_AVAILABLE and gr is not None:
                progress = gr.Progress()
            else:
                progress = lambda *args, **kwargs: None
        global N_TRIALS, CV_FOLDS, OPTUNA_TIMEOUT, OUTPUT_DIR, gradio_results, TRAIN_FOR_MEGA

        # Map UI selection -> training constraints
        target_device = str(target_device or "ESP32").strip()
        TRAIN_FOR_MEGA = (target_device.lower() == "arduino mega")
        global ARDUINO_OPTIMIZE
        ARDUINO_OPTIMIZE = TRAIN_FOR_MEGA
        gradio_results['target_device'] = target_device
        
        if not gradio_data:
            return "❌ No data available. Please upload file first!", None, ""

        mode_note = (
            "⚙️ Target: Arduino Mega (size-aware training ON)" if TRAIN_FOR_MEGA
            else "🎯 Target: ESP32 (size-aware training OFF)"
        )
        
        # Update config
        N_TRIALS = int(n_trials)
        CV_FOLDS = int(cv_folds)
        OPTUNA_TIMEOUT = int(optuna_timeout)
        OUTPUT_DIR = output_dir
        
        # Create folder if not exists
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        
        # Map groups to family
        group_to_family = {
            "SVM (9 models)": "SVM",
            "Tree (5 models)": "Tree",
            "Neural Network (3 models)": "Neural Network",
            "KNN (2 models)": "KNN",
            "Discriminant (2 models)": "Discriminant",
            "Naive Bayes (2 models)": "Naive Bayes"
        }
        
        # Filter selected families
        selected_families = [group_to_family[g] for g in selected_groups if g in group_to_family]
        
        if not selected_families:
            return "❌ Please select at least 1 model group!", None, ""
        
        X_train = gradio_data['X_train']
        X_test = gradio_data['X_test']
        y_train = gradio_data['y_train']
        y_test = gradio_data['y_test']

        progress(0.05, desc=mode_note)
        
        import time
        start_time = time.time()
        
        # T?t optuna logging
        optuna.logging.set_verbosity(optuna.logging.WARNING)
        
        # Normalize data
        scaler = StandardScaler()
        X_train_scaled = scaler.fit_transform(X_train)
        X_test_scaled = scaler.transform(X_test)
        
        # Get all model configs and filter by selected groups
        all_model_configs = get_model_configs()
        model_configs = {name: config for name, config in all_model_configs.items() 
                        if config.get('family') in selected_families}
        total_models = len(model_configs)
        
        all_results = []
        trained_models = {}
        confusion_matrices = {}
        
        # Train each model with progress
        for idx, (model_name, config) in enumerate(model_configs.items()):
            progress_pct = idx / total_models
            progress(progress_pct, desc=f"🔧 Training {model_name} ({idx+1}/{total_models})...")
            
            try:
                model, results, cm = train_single_model(
                    model_name, config, X_train_scaled, y_train, X_test_scaled, y_test, scaler
                )
                
                if model is not None and results is not None:
                    all_results.append(results)
                    trained_models[model_name] = {'model': model, 'scaler': scaler, 'family': config.get('family', 'SVM')}
                    confusion_matrices[model_name] = cm
            except Exception as e:
                print(f" Error training {model_name}: {e}")
                continue
        
        progress(0.9, desc="💾 Saving results...")
        
        # Save results
        save_results(all_results, confusion_matrices)
        
        # Save to global
        gradio_results['all_results'] = all_results
        gradio_results['trained_models'] = trained_models
        gradio_results['confusion_matrices'] = confusion_matrices
        
        # Calculate time
        elapsed_time = time.time() - start_time
        hours = int(elapsed_time // 3600)
        minutes = int((elapsed_time % 3600) // 60)
        seconds = int(elapsed_time % 60)
        time_str = f"{hours}h {minutes}m {seconds}s" if hours > 0 else f"{minutes}m {seconds}s" if minutes > 0 else f"{seconds}s"
        
        progress(1.0, desc=f"✅ Completed! ({time_str})")
        
        # Create results DataFrame
        results_df = pd.DataFrame(all_results)
        results_df = results_df.sort_values('Test Accuracy', ascending=False)
        
        # Hide unnecessary columns for Gradio UI
        columns_to_hide = ['Algorithm Family', 'Kernel/Type', 'Best params (JSON)', 
                          'CV Train Recall', 'CV Train F1', 'CV Val Recall', 'CV Val F1',
                          'Test Recall', 'Test F1', 
                          'Prediction speed (obs/sec)', 'Model size (bytes)']
        display_df = results_df.drop(columns=[col for col in columns_to_hide if col in results_df.columns], errors='ignore')
        
        # Copy for formatting, keep original
        display_df = display_df.copy()
        
        # Format % columns for readability
        for col in display_df.columns:
            if 'Accuracy' in col or 'Precision' in col:
                if col in display_df.columns:
                    display_df[col] = display_df[col].apply(lambda x: f"{x*100:.2f}%" if pd.notna(x) else "")
        
        # Save training results to single Excel file
        excel_path = f"{OUTPUT_DIR}/results.xlsx"
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            # Sheet 1: Training results (numeric for sorting)
            results_df.to_excel(writer, sheet_name='Training Results', index=False)
            # Sheet 2: Summary (% format for readability)
            display_df.to_excel(writer, sheet_name='Summary', index=False)
        print(f"✓ Saved results: {excel_path}")
        
        # Create summary
        summary = f"""✅ **Training completed!**

⏱️ **Time:** {time_str}

📊 **Overview:**
- Trained: {len(all_results)} models
- Saved results to: {OUTPUT_DIR}/

📊 **Top 10 Models (by Test Accuracy):**
"""
        
        top_10 = results_df.head(10)
        for idx, (i, row) in enumerate(top_10.iterrows(), 1):
            summary += f"\n{idx}. **{row['Model']}**: Acc={row['Test Accuracy']*100:.2f}%, Prec={row['Test Precision']*100:.2f}%"
        
        summary += "\n\n---\n\n### ✅ **STEP 2 COMPLETED! GO TO STEP 3 TO SELECT MODELS**\n"
        
        # Create full list for manual selection
        models_list_md = "\n📋 **All models list (sorted by Test Accuracy):**\n\n"
        for idx, (i, row) in enumerate(results_df.iterrows(), 1):
            models_list_md += f"{idx}. {row['Model']} - Acc: {row['Test Accuracy']*100:.2f}%\n"
        
        return summary, display_df, models_list_md
        
    except Exception as e:
        import traceback
        return f"❌ Error: {str(e)}\n\n```\n{traceback.format_exc()}\n```", None, ""

def gradio_select_models(selection_mode, model_indices, auto_count):
    """Select models for Gradio"""
    try:
        global gradio_results
        
        if 'all_results' not in gradio_results:
            return "❌ No training results! Please train models first.", None
        
        all_results = gradio_results['all_results']
        trained_models = gradio_results['trained_models']
        
        selected_models = {}
        num_models = int(auto_count) if auto_count else 10
        
        if selection_mode == "Top N (Test Accuracy)":
            # Sort by Test Accuracy
            sorted_results = sorted(all_results, key=lambda x: x['Test Accuracy'], reverse=True)
            for i in range(min(num_models, len(sorted_results))):
                model_name = sorted_results[i]['Model']
                selected_models[model_name] = trained_models[model_name]
                
        elif selection_mode == "Top N (Train Accuracy)":
            # Sort by CV Train Accuracy
            sorted_results = sorted(all_results, key=lambda x: x['CV Train Accuracy'], reverse=True)
            for i in range(min(num_models, len(sorted_results))):
                model_name = sorted_results[i]['Model']
                selected_models[model_name] = trained_models[model_name]
            
        elif selection_mode == "Top N (Train Precision)":
            # Sort by CV Train Precision
            sorted_results = sorted(all_results, key=lambda x: x['CV Train Precision'], reverse=True)
            for i in range(min(num_models, len(sorted_results))):
                model_name = sorted_results[i]['Model']
                selected_models[model_name] = trained_models[model_name]
        
        elif selection_mode == "Top N (Val Accuracy)":
            # Sort by CV Val Accuracy
            sorted_results = sorted(all_results, key=lambda x: x['CV Val Accuracy'], reverse=True)
            for i in range(min(num_models, len(sorted_results))):
                model_name = sorted_results[i]['Model']
                selected_models[model_name] = trained_models[model_name]
        
        elif selection_mode == "Top N (Val Precision)":
            # Sort by CV Val Precision
            sorted_results = sorted(all_results, key=lambda x: x['CV Val Precision'], reverse=True)
            for i in range(min(num_models, len(sorted_results))):
                model_name = sorted_results[i]['Model']
                selected_models[model_name] = trained_models[model_name]
            
        elif selection_mode == "Top N (Test Precision)":
            # Sort by Test Precision
            sorted_results = sorted(all_results, key=lambda x: x['Test Precision'], reverse=True)
            for i in range(min(num_models, len(sorted_results))):
                model_name = sorted_results[i]['Model']
                selected_models[model_name] = trained_models[model_name]
            
        else:  # Manual
            try:
                indices = [int(x.strip()) for x in model_indices.split(',')]
                sorted_results = sorted(all_results, key=lambda x: x['Test Accuracy'], reverse=True)
                for idx in indices:
                    if 1 <= idx <= len(sorted_results):
                        model_name = sorted_results[idx-1]['Model']
                        selected_models[model_name] = trained_models[model_name]
            except:
                return " Error: Enter numbers in format: 1,2,3", None
        
        if not selected_models:
            return "⚠️ No model selected!", None
        
        gradio_results['selected_models'] = selected_models
        
        info = f"✅ **Selected {len(selected_models)} models:**\n"
        for name in selected_models.keys():
            info += f"- {name}\n"
        
        info += "\n---\n\n###  **STEP 3 COMPLETED! GO TO STEP 4 TO GENERATE ARDUINO LIBRARY**\n"
        
        return info, None
        
    except Exception as e:
        return f" Error: {str(e)}", None

def gradio_generate_arduino(lib_name, func_suffix, separate_files):
    """Generate Arduino library for Gradio"""
    try:
        global ARDUINO_LIB_NAME, FUNC_NAME_SUFFIX, SEPARATE_MODEL_FILES, gradio_results
        
        if 'selected_models' not in gradio_results:
            return " No models selected!", None
        
        ARDUINO_LIB_NAME = lib_name
        # Clean function suffix: remove spaces, ensure safe C identifier
        FUNC_NAME_SUFFIX = func_suffix.replace(' ', '_').replace('-', '_') if func_suffix else ""
        SEPARATE_MODEL_FILES = separate_files
        
        selected_models = gradio_results['selected_models']
        X_train = gradio_data['X_train']
        X_test = gradio_data['X_test']
        y_train = gradio_data['y_train']
        feature_names = gradio_data['feature_names']
        class_labels = sorted(np.unique(y_train))
        
        # Generate Arduino library
        lib_dir = generate_arduino_library(selected_models, feature_names, class_labels, X_train, X_test)
        
        # Note: Cannot calculate memory from Python - not accurate!
        memory_info = {}
        for model_name in selected_models.keys():
            memory_info[model_name] = {
                'rom_bytes': 0,
                'ram_bytes': 0,
                'rom_kb': 0,
                'ram_kb': 0,
                'details': 'Compile to measure'
            }
        
        # Create MCU compatibility table
        mcu_table = "\n🔌 **Microcontroller Compatibility:**\n\n"
        mcu_table += "| Microcontroller | Flash (KB) | SRAM (KB) |\n"
        mcu_table += "|--------------|-------|------|\n"
        
        mcu_specs = [
            ("Arduino Uno", 32, 2),
            ("Arduino Nano", 32, 2),
            ("Arduino Pro Mini", 32, 2),
            ("Arduino Mega 2560", 256, 8),
            ("ESP8266", (1024, 16384), 80),
            ("ESP32", (4096, 16384), 520),
            ("ESP32-S2", (4096, 16384), 320),
            ("ESP32-S3", (4096, 16384), 512),
            ("ESP32-C3", (4096, 16384), 400),
            ("Raspberry Pi Pico", (2048, 16384), 264),
            ("STM32F103", (64, 512), (20, 64)),
            ("STM32F4", (256, 2048), (64, 256)),
            ("STM32F7", (512, 2048), (256, 512)),
            ("STM32H7", (1024, 2048), (512, 1024)),
            ("STM32L", (64, 1024), (16, 320)),
        ]
        
        for mcu_name, flash, sram in mcu_specs:
            # Handle Flash range or single value
            if isinstance(flash, tuple):
                flash_min = flash[0]
                flash_str = f"{flash[0]}?{flash[1]} KB"
            else:
                flash_min = flash
                flash_str = f"{flash} KB"
            
            # Handle SRAM range or single value
            if isinstance(sram, tuple):
                sram_min = sram[0]
                sram_str = f"{sram[0]}?{sram[1]} KB"
            else:
                sram_min = sram
                sram_str = f"{sram} KB"
            
            # Display size only
            mcu_table += f"| {mcu_name} | {flash_str} | {sram_str} |\n"
        
        # Append MCU compatibility and memory info to results.xlsx
        # Note: user may click generate multiple times -> sheet already exists
        excel_path = f"{OUTPUT_DIR}/results.xlsx"

        # Prepare sheets to write
        sheet_library_info = 'Library Info'
        sheet_model_list = 'Model List'
        sheet_how_to_measure = 'How to Measure'
        sheets_to_write = [sheet_library_info, sheet_model_list, sheet_how_to_measure]

        target_device = gradio_results.get('target_device', 'ESP32')
        lib_info_df = pd.DataFrame([{
            'Library Name': lib_name,
            'Number of Models': len(memory_info),
            'Target Device': target_device,
            'Note': ' Compile example in Arduino IDE to measure actual Flash/RAM usage'
        }])

        mem_df = pd.DataFrame([
            {'Model': name, 'Note': 'Compile to see size'}
            for name in memory_info.keys()
        ])

        instructions = pd.DataFrame([
            {'Step': 1, 'Action': 'Open Arduino IDE'},
            {'Step': 2, 'Action': 'Open your sketch using the library'},
            {'Step': 3, 'Action': 'Select your board (ESP32/ESP32-S3/...)'},
            {'Step': 4, 'Action': 'Click Verify to see Flash/RAM in compile output'},
            {'Step': 5, 'Action': 'Upload and open Serial Monitor (115200)'},
            {'Step': 6, 'Action': 'View detailed memory report and benchmarks'},
        ])

        def _write_arduino_info_sheets(_writer):
            lib_info_df.to_excel(_writer, sheet_name=sheet_library_info, index=False)
            mem_df.to_excel(_writer, sheet_name=sheet_model_list, index=False)
            instructions.to_excel(_writer, sheet_name=sheet_how_to_measure, index=False)

        writer_kwargs = {'engine': 'openpyxl'}
        if os.path.exists(excel_path):
            writer_kwargs['mode'] = 'a'
            # pandas>=1.3 supports if_sheet_exists with mode='a'
            writer_kwargs['if_sheet_exists'] = 'replace'
        else:
            writer_kwargs['mode'] = 'w'

        try:
            with pd.ExcelWriter(excel_path, **writer_kwargs) as writer:
                _write_arduino_info_sheets(writer)
        except TypeError:
            # Fallback for older pandas without if_sheet_exists
            with pd.ExcelWriter(excel_path, engine='openpyxl', mode='w') as writer:
                _write_arduino_info_sheets(writer)
        except ValueError as e:
            # Safe fallback if sheet exists error
            if 'already exists' in str(e):
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(excel_path)
                    for s in sheets_to_write:
                        if s in wb.sheetnames:
                            wb.remove(wb[s])
                    wb.save(excel_path)
                    with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a') as writer:
                        _write_arduino_info_sheets(writer)
                except Exception:
                    # Last resort: overwrite file
                    with pd.ExcelWriter(excel_path, engine='openpyxl', mode='w') as writer:
                        _write_arduino_info_sheets(writer)
            else:
                raise
        
        print(f"✓ Saved Arduino info to: {excel_path}")
        
        # Check debug log file
        debug_log_exists = os.path.exists(f"{OUTPUT_DIR}/generation_debug.log")
        debug_note = ""
        if debug_log_exists:
            debug_note = f"""
 **NOTE**: Debug log file available at:
 `{OUTPUT_DIR}/generation_debug.log`
 Check this file if some models were skipped (code generation failed)
"""
        
        info = f"""✅ **Arduino library created: {lib_name}**

📂 **Library created at:** `{OUTPUT_DIR}/{lib_name}/`
{debug_note}
📏 **TO MEASURE ACTUAL SIZE:**

🔍 **Step 1: Compile to see Flash/RAM**
1. Open Arduino IDE
2. Open sketch using the library
3. Select board (ESP32/ESP32-S3/...)
4. Click **Verify** (Compile)
5. See final output:
   ```
   Sketch uses XXXXX bytes (XX%) of program storage space
   Global variables use XXXX bytes (XX%) of dynamic memory
   ```

⬆️ **Step 2: Upload to see details**
6. Upload sketch to board
7. Open Serial Monitor (115200 baud)
8. View full report:
   - Actual Flash Size
   - Heap/RAM usage
   - Execution time per model

{mcu_table}

📝 **Available sketches:**
1. **SerialPredict**: Read data from Serial Monitor
2. **SerialPredict**: Input data via Serial to predict
3. **SensorPredict**: Template for sensor reading

🔧 **Selected models ({len(selected_models)}):**
"""
        for name, mem in memory_info.items():
            func_name_display = get_func_name(name)
            info += f"- {name} -> `predict_{func_name_display}()`\n"
        
        info += f"\n📂 **Mode:** {'Separate files (optimized Flash)' if separate_files else 'Single file'}\n"
        if FUNC_NAME_SUFFIX:
            info += f" **Function suffix:** `{FUNC_NAME_SUFFIX}`\n"
        info += f" **Saved Excel:** {excel_path}\n"
        info += "\n---\n\n###  **STEP 4 COMPLETED! GO TO STEP 5 TO DOWNLOAD**\n"
        
        return info, None
        
    except Exception as e:
        import traceback
        return f" Error: {str(e)}\n\n```\n{traceback.format_exc()}\n```", None

def gradio_create_download():
    """Auto-compress and download ZIP file"""
    try:
        # Auto-compress all results
        package_and_download()
        
        zip_file = f"{OUTPUT_DIR}_Results.zip"
        if os.path.exists(zip_file):
            file_size = os.path.getsize(zip_file) / (1024 * 1024)  # MB
            return zip_file, f"✅ Compressed successfully! File: {zip_file} ({file_size:.1f} MB)"
        else:
            return None, "❌ ZIP file not found. Please run Training first!"
            
    except Exception as e:
        return None, f"❌ Error while compressing: {str(e)}"

def create_gradio_interface():
    """Create Gradio interface"""
    
    # English-only interface
    
    # CSS để ẩn footer Gradio và thêm animation cho buttons
    custom_css = """
    footer {display: none !important;}
    
    /* ========== NÚT CHUYỂN NGÔN NGỮ - CỐ ĐỊNH TRÊN ĐẦU ========== */
    .lang-switch-container {
        position: fixed !important;
        top: 10px !important;
        right: 20px !important;
        z-index: 99999 !important;
    }
    
    .lang-switch button {
        background: linear-gradient(135deg, #f59e0b 0%, #ef4444 50%, #f59e0b 100%) !important;
        background-size: 200% 200% !important;
        animation: gradient-shift 3s ease infinite !important;
        color: white !important;
        border: 3px solid #fbbf24 !important;
        padding: 12px 24px !important;
        border-radius: 50px !important;
        font-weight: 800 !important;
        font-size: 16px !important;
        cursor: pointer !important;
        box-shadow: 0 4px 20px rgba(245, 158, 11, 0.5), 0 0 30px rgba(239, 68, 68, 0.3) !important;
        transition: all 0.3s ease !important;
        text-transform: uppercase !important;
        letter-spacing: 1px !important;
    }
    
    .lang-switch button:hover {
        transform: scale(1.1) translateY(-3px) !important;
        box-shadow: 0 8px 30px rgba(245, 158, 11, 0.7), 0 0 50px rgba(239, 68, 68, 0.5) !important;
        border-color: #fcd34d !important;
    }
    
    @keyframes gradient-shift {
        0% { background-position: 0% 50%; }
        50% { background-position: 100% 50%; }
        100% { background-position: 0% 50%; }
    }
    
    /* Pulse animation for attention */
    .lang-switch::after {
        content: '🌐';
        position: absolute;
        top: -8px;
        right: -8px;
        font-size: 20px;
        animation: bounce 1s infinite;
    }
    
    @keyframes bounce {
        0%, 100% { transform: translateY(0); }
        50% { transform: translateY(-5px); }
    }
    
    @keyframes pulse-ring {
        0% { transform: scale(1); opacity: 0.5; }
        100% { transform: scale(1.2); opacity: 0; }
    }
    
    /* Card-like sections */
    .gradio-container {
        max-width: 1400px !important;
        margin: auto !important;
        padding-top: 60px !important;  /* để tránh che nút ngôn ngữ */
    }
    
    /* Tab styling */
    .tabs {
        border-radius: 8px;
        overflow: hidden;
    }
    
    /* Better spacing */
    .gradio-row {
        gap: 16px !important;
    }
    
    .gradio-column {
        padding: 12px !important;
        background: #1e293b !important;
        border-radius: 8px;
    }
    
    /* Button hover effect - ph?ng to khi hover */
    .pulse-btn button {
        transition: all 0.3s ease !important;
        font-weight: 600 !important;
        border-radius: 8px !important;
    }
    .pulse-btn button:hover {
        transform: scale(1.05) !important;
        box-shadow: 0 0 20px rgba(99, 102, 241, 0.5) !important;
    }
    
    /* Loading state - m? di khi click */
    .pulse-btn button:active {
        opacity: 0.4 !important;
        transform: scale(0.98) !important;
    }

    /* Link-like help buttons (ti?u d? xanh c? th? b?m) */
    .link-btn button {
        background: linear-gradient(135deg, #1e3a8a 0%, #3b82f6 100%) !important;
        color: white !important;
        border: none !important;
        padding: 8px 16px !important;
        border-radius: 6px !important;
        box-shadow: 0 2px 8px rgba(30, 58, 138, 0.3) !important;
        font-weight: 600 !important;
        font-size: 13px !important;
        transition: all 0.3s ease !important;
    }
    .link-btn button:hover {
        background: linear-gradient(135deg, #1e40af 0%, #60a5fa 100%) !important;
        transform: translateY(-2px) !important;
        box-shadow: 0 4px 12px rgba(30, 58, 138, 0.4) !important;
    }
    
    /* Training config inputs - nh? g?n hon */
    .training-input input,
    .training-input textarea {
        font-size: 14px !important;
        padding: 8px 12px !important;
        background: #1e293b !important;
        border: 1px solid #475569 !important;
        color: #e2e8f0 !important;
        border-radius: 6px !important;
    }
    
    .training-input input:focus,
    .training-input textarea:focus {
        border-color: #3b82f6 !important;
        box-shadow: 0 0 0 2px rgba(59, 130, 246, 0.2) !important;
    }
    
    /* Progress status */
    .status-box {
        padding: 16px;
        border-radius: 8px;
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        font-weight: 600;
    }
    
    /* Markdown headings */
    .markdown-text h3 {
        color: #1f2937;
        border-bottom: 2px solid #e5e7eb;
        padding-bottom: 8px;
        margin-top: 24px;
    }
    
    /* File upload areas */
    .file-upload {
        border: 2px dashed #475569 !important;
        border-radius: 8px !important;
        padding: 20px !important;
        transition: all 0.3s ease !important;
        background: #0f172a !important;
        color: #94a3b8 !important;
    }
    
    .file-upload:hover {
        border-color: #64748b !important;
        background: #1e293b !important;
        box-shadow: 0 4px 12px rgba(71, 85, 105, 0.3) !important;
    }
    
    /* Progress bar - make it LARGE and prominent */
    .progress-container {
        position: fixed !important;
        top: 50% !important;
        left: 50% !important;
        transform: translate(-50%, -50%) !important;
        z-index: 9999 !important;
        background: rgba(0, 0, 0, 0.85) !important;
        padding: 40px 60px !important;
        border-radius: 16px !important;
        box-shadow: 0 20px 60px rgba(0, 0, 0, 0.5) !important;
    }
    
    .progress-text {
        color: white !important;
        font-size: 24px !important;
        font-weight: 700 !important;
        text-align: center !important;
        margin-bottom: 20px !important;
    }
    
    .progress-bar {
        height: 12px !important;
        border-radius: 6px !important;
        background: linear-gradient(90deg, #3b82f6, #8b5cf6) !important;
    }
    """
    
    with gr.Blocks(title="ML Pipeline - Arduino Library Generator", theme=gr.themes.Soft(), css=custom_css) as app:
        # Header
        gr.Markdown("""
        #  Machine Learning Pipeline - Arduino Library Generator
        ### Train ML models and generate libraries for Arduino/ESP32
        ---
        """)
        
        with gr.Tabs():
            # ===== TAB 1: UPLOAD DATA =====
            with gr.Tab("📁 1. Upload Data") as tab1:
                tab1_title = gr.Markdown("""
                # 📁 Step 1: Upload Data
                Choose one of the two upload options below:
                """)
                
                data_info = gr.Markdown("""
                ### 📋 Instructions:
                - **Option 1** (Recommended): Upload 1 Excel/CSV file with 'Class' column → Auto-split train/test using SPXY
                - **Option 2**: Upload 2 pre-split files (train.xlsx + test.xlsx)
                - After processing → Check the data distribution chart
                """, elem_classes=["info-box"])
                
                with gr.Row():
                    with gr.Column(scale=1, elem_classes=["upload-section"]):
                        option1_title = gr.Markdown("### 📁 Option 1: Single File")
                        option1_desc = gr.Markdown("*Auto-split train/test using SPXY algorithm*")
                        single_file = gr.File(label="📊 Select Excel or CSV file", file_types=['.xlsx', '.csv'], show_label=True, elem_classes=["file-upload"])
                        
                        # Help panel for upload params
                        upload_help = gr.Markdown("💡 *Click blue titles to see detailed explanation*", visible=True)
                        
                        with gr.Row():
                            with gr.Column(scale=1):
                                btn_help_train_ratio = gr.Button("❓ Train Ratio", elem_classes=["link-btn"])
                                train_ratio = gr.Slider(0.5, 0.9, value=0.7, step=0.05, label="", show_label=False)
                            with gr.Column(scale=1):
                                btn_help_random_seed = gr.Button("🎲 Random Seed", elem_classes=["link-btn"])
                                random_state = gr.Number(value=42, label="", precision=0, show_label=False)
                        
                        hide_upload_help = gr.Button("Hide explanation", variant="secondary", size="sm")
                        btn_single = gr.Button("▶️ Process File", variant="primary", size="lg", elem_classes="pulse-btn")
                    
                    with gr.Column(scale=1, elem_classes=["upload-section"]):
                        gr.Markdown("### 📄 Option 2: Pre-split Files")
                        gr.Markdown("*Use existing train.xlsx and test.xlsx*")
                        train_file = gr.File(label="📄 Train File", file_types=['.xlsx', '.csv'], show_label=True, elem_classes=["file-upload"])
                        test_file = gr.File(label="📄 Test File", file_types=['.xlsx', '.csv'], show_label=True, elem_classes=["file-upload"])
                        gr.Markdown("")  # Spacer
                        btn_split = gr.Button("▶️ Process Files", variant="primary", size="lg", elem_classes="pulse-btn")
                
                # Results and download info
                gr.Markdown("---")
                gr.Markdown("## 📊 Processing Results")
                data_result = gr.Markdown("*Results will be displayed here after processing*", elem_classes=["status-message"])
                
                gr.Markdown("### 📊 Data Distribution Chart")
                plot_image = gr.Image(label="Train/Test Distribution", type="filepath", interactive=False, show_label=True)
                
                # Explanation functions for upload parameters
                def explain_train_ratio():
                    return (
                        """### ❓ Train Ratio

                        - Percentage of data used for training (the rest for testing).
                        - **Recommended**: 0.7-0.8 (70-80%) for most tasks.
                        - Uses **SPXY** algorithm to split data evenly.
                        """
                    )
                
                def explain_random_seed():
                    return (
                        """### 🎲 What is Random Seed?

                        **Simple definition:**
                        - Random Seed is ANY number (42, 99, 5, 123, 9999... any number works!)
                        - Code uses this number to **initialize the data shuffler**
                        - Different seed → Different shuffle → Different results

                        ### 📌 Example:

                        **You have 10 samples:** [A, B, C, D, E, F, G, H, I, J]

                        **When seed = 42:**
                        1. Code shuffles to: [D, A, H, B, I, C, J, E, F, G]
                        2. Take 70% = 7 samples for Train: [D, A, H, B, I, C, J]
                        3. Take 30% = 3 samples for Test: [E, F, G]
                        4. Run model → Test Accuracy = **93.2%**

                        **When seed = 99:**
                        1. Code shuffles to: [G, C, A, I, E, B, D, F, H, J]
                        2. Take 70% = 7 samples for Train: [G, C, A, I, E, B, D]
                        3. Take 30% = 3 samples for Test: [F, H, J]
                        4. Run model → Test Accuracy = **92.8%** (DIFFERENT because test samples differ!)

                        ---
                     
                        """
                    )
                
                # Wire click events for upload params
                btn_help_train_ratio.click(fn=explain_train_ratio, outputs=upload_help)
                btn_help_random_seed.click(fn=explain_random_seed, outputs=upload_help)
                hide_upload_help.click(fn=lambda: "💡 *Click blue titles to see detailed explanation*", outputs=upload_help)
            
            # ===== TAB 2: TRAINING CONFIG =====
            with gr.Tab("⚙️ 2. Config & Training"):
                training_info = gr.Markdown("⏳ Not trained yet")
                
                gr.Markdown("### Configure training parameters:")

                gr.Markdown("### 🎯 Select target device:")
                target_device = gr.Radio(
                    choices=["ESP32", "Arduino Mega"],
                    value="ESP32",
                    label="Target",
                    info="ESP32: prioritize accuracy. Arduino Mega: Optuna prioritizes smaller models to fit Flash."
                )

                # Help box for parameter explanations
                param_help = gr.Markdown("💡 Click parameter title to see explanation.")
                hide_help_btn = gr.Button("🔽 Hide explanation", variant="secondary", size="sm")
                
                with gr.Row():
                    with gr.Column(scale=1):
                        btn_help_n_trials = gr.Button("🎯 N_TRIALS", elem_classes=["link-btn"])
                        n_trials = gr.Number(value=50, label="", precision=0, show_label=False, elem_classes=["training-input"])
                    with gr.Column(scale=1):
                        btn_help_cv_folds = gr.Button("🔄 CV_FOLDS", elem_classes=["link-btn"])
                        cv_folds = gr.Number(value=5, label="", precision=0, show_label=False, elem_classes=["training-input"])
                
                with gr.Row():
                    with gr.Column(scale=1):
                        btn_help_optuna_timeout = gr.Button("⏱️ TIMEOUT (seconds)", elem_classes=["link-btn"])
                        optuna_timeout = gr.Number(value=180, label="", precision=0, show_label=False, elem_classes=["training-input"])
                    with gr.Column(scale=1):
                        btn_help_output_dir = gr.Button("📁 OUTPUT_DIR", elem_classes=["link-btn"])
                        output_dir_name = gr.Textbox(value="ML_Output", label="", show_label=False, elem_classes=["training-input"])
                
                # Parameter explanation functions
                def explain_n_trials():
                    return (
                        """### ❓ What is N_TRIALS?

                        - Number of trials for Optuna to find the best hyperparameters for EACH model.
                        - Each trial = 1 different parameter set. More trials → easier to find better config but takes longer.
                        - Suggestion: 20-50 (fast, demo), 100-300 (good quality), >300 (for large data and thorough optimization).

                        Example:
                        - N_TRIALS = 50 → each model will try 50 different configurations.
                        - 6 models × 50 trials = up to ~300 configurations tested.
                        """
                    )

                def explain_cv_folds():
                    return (
                        """### ❓ What is CV_FOLDS?

                        - Number of folds when splitting train set for cross-validation (K-Fold CV).
                        - Larger → more stable results but takes more time. Common: 5 or 10.

                        Example:
                        - CV_FOLDS = 5 → split train into 5 parts, rotate 5 times train/validate, take average score.
                        - Small data → use 5 to balance speed and stability.
                        """
                    )

                def explain_optuna_timeout():
                    return (
                        """### ❓ What is OPTUNA_TIMEOUT?

                        - Maximum time (seconds) to optimize EACH model. Even if N_TRIALS not finished, Optuna stops when time's up.
                        - Used to limit total training time.

                        Example:
                        - OPTUNA_TIMEOUT = 180 → each model optimizes for max 3 minutes.
                        - If N_TRIALS=200 but 180 seconds elapsed → stops early, takes best config found.
                        """
                    )

                def explain_output_dir():
                    return (
                        """### ❓ What is OUTPUT_DIR?

                        - Folder name to save all results: results.xlsx, charts, and generated Arduino library.
                        - Can be named by project for easy management.

                        Example:
                        - OUTPUT_DIR = "ML_Output" → creates ML_Output/ folder containing all results.
                        - Each new run should use different name to avoid overwriting (e.g.: ML_Output_2025_12_29).
                        """
                    )

                # Wire click events to show/hide parameter explanations
                btn_help_n_trials.click(fn=explain_n_trials, outputs=param_help)
                btn_help_cv_folds.click(fn=explain_cv_folds, outputs=param_help)
                btn_help_optuna_timeout.click(fn=explain_optuna_timeout, outputs=param_help)
                btn_help_output_dir.click(fn=explain_output_dir, outputs=param_help)
                hide_help_btn.click(fn=lambda: "💡 Click parameter title to see explanation.", outputs=param_help)

                gr.Markdown("---")
                gr.Markdown("### Select model groups to train:")
                
                model_groups = gr.CheckboxGroup(
                    choices=[
                        "SVM (9 models)",
                        "Tree (5 models)",
                        "Neural Network (3 models)",
                        "KNN (2 models)",
                        "Discriminant (2 models)",
                        "Naive Bayes (2 models)"
                    ],
                    value=[
                        "SVM (9 models)",
                        "Tree (5 models)",
                        "Neural Network (3 models)",
                        "KNN (2 models)",
                        "Discriminant (2 models)",
                        "Naive Bayes (2 models)"
                    ],
                    label="Model Groups",
                    interactive=True
                )
                
                with gr.Row():
                    btn_select_all = gr.Button("✅ Select All", size="sm")
                    btn_deselect_all = gr.Button("❌ Deselect All", size="sm")
                
                gr.Markdown("---")
                btn_train = gr.Button("🚀 Start Training", variant="primary", size="lg", elem_classes="pulse-btn")
                
                results_table = gr.Dataframe(
                    label="Training Results",
                    interactive=False,
                    wrap=True
                )
            
            # ===== TAB 3: MODEL SELECTION =====
            with gr.Tab(" 3. Select Models"):
                gr.Markdown("""
                # 📋 Step 3: Select Best Models
                After training, select models to generate Arduino/ESP32 library:
                """)
                
                selection_info = gr.Markdown("*Results will be displayed after selecting models*", elem_classes=["status-message"])
                
                gr.Markdown("### 📖 How to select models:")
                
                with gr.Row():
                    with gr.Column(scale=1):
                        selection_mode = gr.Radio(
                            choices=[
                                "Top N (Test Accuracy)",
                                "Top N (Train Accuracy)",
                                "Top N (Val Accuracy)",
                                "Top N (Train Precision)",
                                "Top N (Val Precision)",
                                "Top N (Test Precision)",
                                "Manual (enter numbers)"
                            ],
                            value="Top N (Test Accuracy)",
                            label=" Sorting Criteria",
                            info="Select criteria to filter top N models"
                        )
                    with gr.Column(scale=1):
                        auto_count = gr.Number(
                            value=10,
                            label="🔢 Number of models",
                            precision=0,
                            minimum=1,
                            maximum=50,
                            info="How many models to select (default 10)",
                            elem_classes=["training-input"]
                        )
                
                model_indices = gr.Textbox(
                    label="✏️ Enter model indices (Manual mode)",
                    placeholder="Example: 1,3,5,7,10",
                    visible=False,
                    info="Enter numbers separated by commas"
                )
                
                models_list = gr.Markdown("ℹï¸Â Model list will appear after training", visible=False)
                
                btn_select = gr.Button("✅ Confirm model selection", variant="primary", elem_classes="pulse-btn")
            
            # ===== TAB 4: ARDUINO LIBRARY =====
            with gr.Tab(" 4. Generate Arduino Library"):
                gr.Markdown("""
                # 📦 Step 4: Generate Arduino/ESP32 Library
                Convert selected models to C/C++ code to run on microcontrollers:
                """)
                
                generation_info = gr.Markdown("*Results will be displayed after generating library*", elem_classes=["status-message"])
                
                gr.Markdown("### ⚙️ Library configuration:")
                
                with gr.Row():
                    with gr.Column(scale=1):
                        lib_name = gr.Textbox(value="MLPredictor", label="📚 Library name", info="Folder name and C++ class")
                    with gr.Column(scale=1):
                        func_suffix = gr.Textbox(value="", label="🔧 Function name suffix", info="Added after model name (e.g. '_v1' → predict_SVM_RBF_v1)")
                
                with gr.Row():
                    with gr.Column(scale=1):
                        separate_files = gr.Checkbox(
                            value=True, 
                            label="✅ Separate each model into individual file",
                            info="Recommended: Optimize Flash ROM for ESP32/Arduino"
                        )
                
                gr.Markdown("---")
                btn_generate = gr.Button("🔨 Generate Arduino Library", variant="primary", size="lg", elem_classes="pulse-btn")
            
            # ===== TAB 5: DOWNLOAD =====
            with gr.Tab("💾 5. Download"):
                gr.Markdown("""
                # 💾 Step 5: Download Results
                All results are packaged into a single ZIP file:
                """)
                
                download_info = gr.Markdown("*Click button below to auto-compress and download*", elem_classes=["status-message"])
                
                gr.Markdown("""
                ### 📦 ZIP file contents:
                
                **ZIP file will auto-compress and include:**
                - 📄 **train_data.xlsx & test_data.xlsx** - Split data
                - 📊 **data_distribution.png** - Data distribution chart
                - 📋 **results.xlsx** - ALL results (7 sheets):
                  - Training Results (full numerical results)
                  - Summary (results in % format)
                  - Library Info (overview: name, model count, Flash/SRAM size, recommended MCU)
                  - Model Memory (Flash/SRAM per model detail)
                  - MCU Compatibility (compatibility with 15 MCUs with % usage)
                  - How to Measure (step-by-step instructions)
                - 🖼️ **confusion_matrices/** - All confusion matrices (PNG)
                - 📦 **Complete Arduino library** (library folder with .h, .cpp, examples)
                
**Note:** Clicking button will auto-compress and download immediately!
                """)
                
                download_file = gr.File(label="Result ZIP file")
                btn_download = gr.Button("📥 Download All (ZIP)", variant="primary", size="lg", elem_classes="pulse-btn")
        
        # ===== EVENT HANDLERS =====
        
        # Select/Deselect all model groups
        btn_select_all.click(
            lambda: [
                "SVM (9 models)",
                "Tree (5 models)",
                "Neural Network (3 models)",
                "KNN (2 models)",
                "Discriminant (2 models)",
                "Naive Bayes (2 models)"
            ],
            outputs=[model_groups]
        )
        
        btn_deselect_all.click(
            lambda: [],
            outputs=[model_groups]
        )
        
        # Upload single file
        btn_single.click(
            gradio_process_single_file,
            inputs=[single_file, train_ratio, random_state],
            outputs=[data_result, plot_image]
        )
        
        # Upload split files
        btn_split.click(
            gradio_process_split_files,
            inputs=[train_file, test_file],
            outputs=[data_result, plot_image]
        )
        
        # Train models
        btn_train.click(
            gradio_train_models,
            inputs=[target_device, n_trials, cv_folds, optuna_timeout, output_dir_name, model_groups],
            outputs=[training_info, results_table, models_list]
        )
        
        # Toggle visibility for manual mode
        def toggle_visibility(mode):
            if mode == "Manual (enter numbers)":
                return gr.update(visible=True), gr.update(visible=True)
            else:
                return gr.update(visible=False), gr.update(visible=False)
        
        selection_mode.change(
            toggle_visibility,
            inputs=[selection_mode],
            outputs=[model_indices, models_list]
        )
        
        # Select models
        btn_select.click(
            gradio_select_models,
            inputs=[selection_mode, model_indices, auto_count],
            outputs=[selection_info, generation_info]
        )
        
        # Generate Arduino library
        btn_generate.click(
            gradio_generate_arduino,
            inputs=[lib_name, func_suffix, separate_files],
            outputs=[generation_info, download_info]
        )
        
        # Create download package
        btn_download.click(
            gradio_create_download,
            outputs=[download_file, download_info]
        )
    
    return app

# ============================================================================
# MAIN EXECUTION
# ============================================================================

def main():
    """Main function to execute the entire pipeline"""
    print("\n" + "=" * 80)
    print(" " * 20 + "MACHINE LEARNING PIPELINE")
    print(" " * 15 + "SVM Models for Arduino/ESP32")
    print("=" * 80)
    print(f"Started at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    try:
        # 1. Load data
        X_train, X_test, y_train, y_test, feature_names = load_data()
        
        # 2. Plot distribution chart
        plot_data_distribution(X_train, X_test, y_train, y_test, feature_names)
        
        # 3. Train models
        all_results, trained_models, confusion_matrices = train_all_models(
            X_train, y_train, X_test, y_test
        )
        
        # 4. Save results
        save_results(all_results, confusion_matrices)
        
        # 5. Select models to generate Arduino library
        selected_models = select_models_for_arduino(all_results, trained_models)
        
        # 6. Create Arduino library (only for selected models)
        class_labels = sorted(np.unique(y_train))
        generate_arduino_library(selected_models, feature_names, class_labels, X_train, X_test)
        
        print("\n" + "=" * 80)
        print(" " * 25 + "✅ COMPLETED!")
        print("=" * 80)
        print(f"Finished at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"\nResults saved at: {OUTPUT_DIR}/")
        print("\nOutput files:")
        print("  • train_data.xlsx, test_data.xlsx - Split data")
        print("  • data_distribution.png - Distribution chart")
        print("  • results.xlsx - Training results (7 sheets: Results, Summary, Library Info, Model Memory, MCU Compatibility, How to Measure)")
        print("  • confusion_matrices/ - Confusion matrices")
        print(f"  • {ARDUINO_LIB_NAME}/ - Arduino/ESP32 library")
        print("\n💡 Tip: Run package_and_download() if you need to create ZIP file")
        
    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback
        traceback.print_exc()

# ============================================================================
# RUN
# ============================================================================

if __name__ == "__main__":
    # If running on Colab/Kaggle, automatically use Gradio
    if IN_CLOUD:
        print("\n" + "=" * 60)
        print(f"🌐 {'Google Colab' if IN_COLAB else 'Kaggle'} - ML Pipeline for Arduino/ESP32")
        print("=" * 60)
        
        if not GRADIO_AVAILABLE:
            print("📦 Installing Gradio...")
            import subprocess
            subprocess.run([sys.executable, "-m", "pip", "install", "-q", "gradio"])
            import gradio as gr
            GRADIO_AVAILABLE = True
        
        print("\n🚀 Starting web interface...\n")
        
        app = create_gradio_interface()
        
        print("\n" + "=" * 60)
        print("👉 CLICK THE LINK BELOW TO OPEN INTERFACE:")
        print("=" * 60 + "\n")
        
        app.launch(
            share=True,
            debug=False,
            show_error=True,
            max_file_size="100mb"
        )
        
    else:
        # Local machine - allow mode selection
        import argparse
        parser = argparse.ArgumentParser(description='ML Pipeline - Train models and generate Arduino libraries')
        parser.add_argument('--mode', type=str, default='gradio', choices=['cli', 'gradio'],
                           help='Run mode: cli (command line) or gradio (web interface)')
        parser.add_argument('--share', action='store_true', help='Create public link for Gradio (share=True)')
        parser.add_argument('--port', type=int, default=7860, help='Port for Gradio server')
        
        args = parser.parse_args()
        
        if args.mode == 'gradio':
            # Gradio Web Interface mode
            if not GRADIO_AVAILABLE:
                print("⚠️  Gradio not installed!")
                print("Install with: pip install gradio")
                print("Or run in CLI mode: python generate15.py --mode cli")
                sys.exit(1)
            
            print("\n" + "=" * 80)
            print(" " * 20 + "🌐 GRADIO WEB INTERFACE")
            print(" " * 15 + "ML Pipeline for Arduino/ESP32")
            print("=" * 80)
            print(f"\n🚀 Starting Gradio server on port {args.port}...")
            
            app = create_gradio_interface()
            app.launch(
                share=args.share,
                server_name="0.0.0.0",
                server_port=args.port,
                show_error=True
            )
            
        else:
            # CLI (Command Line) mode - original code
            # Mount Google Drive if needed
            if IN_COLAB and (SINGLE_FILE_PATH.startswith('/content/drive') or 
                             TRAIN_FILE_PATH.startswith('/content/drive') or 
                             TEST_FILE_PATH.startswith('/content/drive')):
                print("📂 Mounting Google Drive...")
                from google.colab import drive
                drive.mount('/content/drive')
            
            # Run pipeline CLI
            main()





